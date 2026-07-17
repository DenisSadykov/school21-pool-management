import os
import json
import time
import threading
import secrets
import re
import hashlib
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from io import BytesIO
from functools import wraps
from datetime import datetime, date, timedelta, timezone

from flask import Flask, jsonify, request, g, send_file, abort
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy.pool import NullPool
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature

def _env_flag(name, default='false'):
    return os.getenv(name, default).strip().lower() in {'1', 'true', 'yes', 'on'}


def _is_production_runtime():
    return _env_flag('VERCEL') or os.getenv('FLASK_ENV', '').lower() == 'production'


load_dotenv()
if not _env_flag('SKIP_LOCAL_DOTENV', 'false'):
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env.local'), override=True)


def database_engine_options():
    """Avoid persistent per-instance pools in ephemeral serverless runtimes."""
    if not _env_flag('VERCEL') or not _env_flag('SERVERLESS_DB_NULL_POOL', 'true'):
        return {}
    return {
        'poolclass': NullPool,
        'connect_args': {
            'connect_timeout': max(1, int(os.getenv('DATABASE_CONNECT_TIMEOUT', '10'))),
        },
    }


def _naive_utcnow():
    return datetime.now(timezone.utc).replace(tzinfo=None)


def get_model_or_404(model, obj_id):
    obj = db.session.get(model, obj_id)
    if obj is None:
        abort(404)
    return obj


app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///pool.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = database_engine_options()
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-me')
app.config['TESTING'] = _env_flag('TESTING', 'false')

frontend_urls = [
    value.strip()
    for value in os.getenv(
        'FRONTEND_URL',
        'http://localhost:3000,http://localhost:3001,http://localhost:3005,'
        'http://127.0.0.1:3000,http://127.0.0.1:3001,http://127.0.0.1:3005',
    ).split(',')
    if value.strip()
]
frontend_url_set = set(frontend_urls)
CORS(
    app,
    resources={r"/api/.*": {"origins": frontend_urls or "*"}},
    allow_headers=['Content-Type', 'Authorization'],
    methods=['GET', 'POST', 'PATCH', 'DELETE', 'OPTIONS'],
)


@app.after_request
def add_api_cors_headers(response):
    if not request.path.startswith('/api/'):
        return response
    origin = request.headers.get('Origin', '')
    if not origin:
        return response
    if frontend_url_set and '*' not in frontend_url_set and origin not in frontend_url_set:
        return response
    response.headers['Access-Control-Allow-Origin'] = origin if origin else '*'
    response.headers['Access-Control-Allow-Headers'] = 'Authorization, Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PATCH, DELETE, OPTIONS'
    response.headers['Vary'] = 'Origin'
    return response

db = SQLAlchemy(app)
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'], salt='auth')

# Роли, которым нужен пароль для входа
ROLES_WITH_PASSWORD = {'team_lead', 'admin'}
ALL_ROLES = {'volunteer', 'tribe_master', 'team_lead', 'admin'}
VOLUNTEER_PROFILE_ROLES = {'volunteer', 'tribe_master', 'team_lead'}
TRIBES = ['Ленты', 'Короны', 'Олени']
TRIBE_ALIASES = {
    '1': 'Ленты',
    'a': 'Ленты',
    'ленты': 'Ленты',
    '2': 'Короны',
    'b': 'Короны',
    'короны': 'Короны',
    '3': 'Олени',
    'c': 'Олени',
    'олени': 'Олени',
}
STUDENT_EVENT_POINTS = {
    'entertainment': 2,
    'education': 4,
}
STUDENT_EVENT_STATUSES = {'pending', 'confirmed', 'rejected'}
PENALTY_STATUSES = {'pending', 'in_workoff', 'overdue', 'awaiting_unlock', 'unlocked', 'done'}
PENALTY_DUPLICATE_WINDOW = timedelta(minutes=3)
ANNOUNCEMENT_DUPLICATE_WINDOW = timedelta(minutes=1)
TIME_VALUE_RE = re.compile(r'^(?:[01]\d|2[0-3]):[0-5]\d$')
REWARD_RATES = {
    'first_day_hour': 15,
    'exam_hour': 15,
    'first_week_or_weekend_hour': 8,
    'subsequent_weekday_hour': 5,
    'group_review': 25,
    'confession': 25,
    'tribe_master_event': 30,
    'team_lead': 450,
}
REWARD_EVENT_TYPES = {
    'confession': {'label': 'Исповедь', 'coins': REWARD_RATES['confession']},
}

# Односторонняя выгрузка platform -> Google Sheets через Apps Script.
SYNC_SECRET = os.getenv('SYNC_SECRET', '')
BACKUP_DIR = os.getenv('BACKUP_DIR', os.path.join(os.path.dirname(__file__), 'backups'))
BACKUP_INTERVAL = int(os.getenv('BACKUP_INTERVAL', str(60 * 60)))  # сек
_backup_lock = threading.Lock()
_runtime_started = False
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '').strip()
TELEGRAM_API_TIMEOUT = max(1.0, float(os.getenv('TELEGRAM_API_TIMEOUT', '8')))
TELEGRAM_BOT_USERNAME = os.getenv('TELEGRAM_BOT_USERNAME', '').strip().lstrip('@')
TELEGRAM_TEST_MODE = os.getenv('TELEGRAM_TEST_MODE', 'true').lower() == 'true'
TELEGRAM_POLL_INTERVAL = float(os.getenv('TELEGRAM_POLL_INTERVAL', '2'))
TELEGRAM_LONG_POLL_TIMEOUT = int(os.getenv('TELEGRAM_LONG_POLL_TIMEOUT', '20'))
TELEGRAM_QUIET_HOURS_START = int(os.getenv('TELEGRAM_QUIET_HOURS_START', '23'))
TELEGRAM_QUIET_HOURS_END = int(os.getenv('TELEGRAM_QUIET_HOURS_END', '7'))
TELEGRAM_WEBHOOK_SECRET = os.getenv('TELEGRAM_WEBHOOK_SECRET', '').strip()
TELEGRAM_SUPPORT_CONTACT = os.getenv('TELEGRAM_SUPPORT_CONTACT', '@odessabu').strip()
INTERNAL_API_SECRET = os.getenv('INTERNAL_API_SECRET', '').strip()
NOTIFICATION_DISPATCH_URL = os.getenv(
    'NOTIFICATION_DISPATCH_URL',
    'https://school21-pool-management-api.vercel.app/api/notifications/dispatch',
).strip()
NOTIFICATION_DISPATCH_TIMEOUT_MS = max(
    1000,
    int(os.getenv('NOTIFICATION_DISPATCH_TIMEOUT_MS', '60000')),
)
SCHOOL_RULES_URL = os.getenv('SCHOOL_RULES_URL', 'https://applicant.21-school.ru/rules')
EXAM_BRIEF_URL = os.getenv('EXAM_BRIEF_URL', '')
MOSCOW_OFFSET = timedelta(hours=3)
_telegram_bot_avatar_cache = {
    'expires_at': None,
    'bytes': None,
    'content_type': None,
}


def _normalized_dedupe_text(value):
    return ' '.join((value or '').split()).casefold()


def _dedupe_lock_key(*parts):
    payload = ':'.join(str(part) for part in parts)
    return hashlib.sha256(payload.encode('utf-8')).hexdigest()


def _acquire_dedupe_lock(*parts):
    if db.engine.dialect.name != 'postgresql':
        return
    db.session.execute(
        db.text('SELECT pg_advisory_xact_lock(hashtextextended(:key, 0))'),
        {'key': _dedupe_lock_key(*parts)},
    )


def _moscow_today():
    return (datetime.now(timezone.utc) + MOSCOW_OFFSET).date()

# ==================== Модели ====================


class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    nick = db.Column(db.String(100), unique=True, nullable=False)
    name = db.Column(db.String(100))
    role = db.Column(db.String(20), nullable=False, default='volunteer')
    is_group_reviewer = db.Column(db.Boolean, default=False)
    has_confession = db.Column(db.Boolean, default=False)
    coins_adjustment = db.Column(db.Integer, default=0)
    password_hash = db.Column(db.String(255))  # только для team_lead / admin
    telegram = db.Column(db.String(100))
    tribe = db.Column(db.String(50))
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'nick': self.nick,
            'name': self.name or self.nick,
            'role': self.role,
            'is_group_reviewer': bool(self.is_group_reviewer),
            'has_confession': bool(self.has_confession),
            'coins_adjustment': self.coins_adjustment or 0,
            'telegram': self.telegram,
            'tribe': self.tribe,
            'active': self.active,
            'has_password': bool(self.password_hash),
            'avatar_url': _avatar_url_for_user(self),
        }


class Pool(db.Model):
    __tablename__ = 'pools'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    start_date = db.Column(db.Date)
    active = db.Column(db.Boolean, default=True)
    archived = db.Column(db.Boolean, default=False)
    google_sheet_url = db.Column(db.Text)
    google_sheet_webhook_url = db.Column(db.Text)
    google_sheet_enabled = db.Column(db.Boolean, default=False)
    google_sheet_last_export_at = db.Column(db.DateTime)
    google_sheet_last_error = db.Column(db.Text)
    self_signup_enabled = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    __table_args__ = (
        db.Index(
            'uq_single_active_pool',
            'active',
            unique=True,
            postgresql_where=db.text('active = true'),
            sqlite_where=db.text('active = 1'),
        ),
    )

    def to_dict(self):
        if self.archived:
            state = 'archived'
        elif self.active:
            state = 'active'
        else:
            state = 'ended'
        generations_count = ScheduleGeneration.query.filter_by(pool_id=self.id).count()
        last_generation = (
            ScheduleGeneration.query
            .filter_by(pool_id=self.id)
            .order_by(ScheduleGeneration.created_at.desc())
            .first()
        )
        return {
            'id': self.id,
            'name': self.name,
            'start_date': self.start_date.isoformat() if self.start_date else None,
            'active': self.active,
            'archived': bool(self.archived),
            'state': state,
            'schedule_generations_count': generations_count,
            'has_schedule_generation': generations_count > 0,
            'last_schedule_generation_at': last_generation.created_at.isoformat() if last_generation and last_generation.created_at else None,
            'self_signup_enabled': self.self_signup_enabled is not False,
        }


class ShiftBlock(db.Model):
    """Тайм-блок смены: день + интервал времени. На него записывается N волонтёров."""
    __tablename__ = 'shift_blocks'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    time_start = db.Column(db.String(5), nullable=False)  # "10:00"
    time_end = db.Column(db.String(5), nullable=False)    # "14:00"
    label = db.Column(db.String(50), default='')          # '', 'EXAM', ...
    capacity = db.Column(db.Integer)                       # None = без лимита
    generation_id = db.Column(db.Integer, db.ForeignKey('schedule_generations.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class ScheduleGeneration(db.Model):
    __tablename__ = 'schedule_generations'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)
    end_date = db.Column(db.Date)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class Signup(db.Model):
    """Запись одного волонтёра на один тайм-блок (одна строка = один человек)."""
    __tablename__ = 'signups'
    id = db.Column(db.Integer, primary_key=True)
    block_id = db.Column(db.Integer, db.ForeignKey('shift_blocks.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    __table_args__ = (db.UniqueConstraint('block_id', 'user_id', name='uq_block_user'),)


class PoolVolunteer(db.Model):
    """Привязка волонтёра к бассейну. Без записи волонтёр не видит бассейн."""
    __tablename__ = 'pool_volunteers'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    tribe = db.Column(db.String(50))
    pool_role = db.Column(db.String(40), default='volunteer')
    notifications_enabled = db.Column(db.Boolean, default=True, nullable=False)
    has_confession = db.Column(db.Boolean, default=False)
    coins_adjustment = db.Column(db.Integer, default=0)
    assigned_at = db.Column(db.DateTime, default=_naive_utcnow)
    __table_args__ = (db.UniqueConstraint('pool_id', 'user_id', name='uq_pool_volunteer'),)


class PoolInviteLink(db.Model):
    __tablename__ = 'pool_invite_links'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False, unique=True)
    token = db.Column(db.String(128), nullable=False, unique=True)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    uses_count = db.Column(db.Integer, default=0, nullable=False)
    max_uses = db.Column(db.Integer)
    expires_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    updated_at = db.Column(db.DateTime, default=_naive_utcnow, onupdate=_naive_utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'pool_id': self.pool_id,
            'token': self.token,
            'is_active': bool(self.is_active),
            'uses_count': self.uses_count or 0,
            'max_uses': self.max_uses,
            'expires_at': self.expires_at.isoformat() if self.expires_at else None,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
        }


class RewardEvent(db.Model):
    __tablename__ = 'reward_events'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=True)
    event_type = db.Column(db.String(50), nullable=False)
    event_date = db.Column(db.Date)
    quantity = db.Column(db.Integer, default=1)
    coins = db.Column(db.Integer, nullable=False)
    comment = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class Tribe(db.Model):
    __tablename__ = 'tribes'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'))
    name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


STANDARD_TRIBES_NNV = ['Ленты', 'Короны', 'Олени']


class GroupReview(db.Model):
    __tablename__ = 'group_reviews'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)
    review_date = db.Column(db.Date, nullable=False)
    time_start = db.Column(db.String(5), nullable=False)
    flow = db.Column(db.String(50), nullable=False)
    quantity = db.Column(db.Integer, default=1)
    reviewer_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class Student(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    nick = db.Column(db.String(100), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    tribe = db.Column(db.String(50))
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)
    total_penalty_hours = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    __table_args__ = (db.UniqueConstraint('pool_id', 'nick', name='uq_student_pool_nick'),)


class StudentEvent(db.Model):
    __tablename__ = 'student_events'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    event_type = db.Column(db.String(30), nullable=False)  # entertainment, education
    event_date = db.Column(db.Date)
    post_url = db.Column(db.String(500))
    proof_url = db.Column(db.String(500))
    points = db.Column(db.Integer, default=0)
    status = db.Column(db.String(20), default='pending')
    comment = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class TribeEvent(db.Model):
    __tablename__ = 'tribe_events'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)
    tribe = db.Column(db.String(50), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    event_date = db.Column(db.Date, nullable=False)
    time_start = db.Column(db.String(5))
    location = db.Column(db.String(200))
    comment = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class StudentPenalty(db.Model):
    __tablename__ = 'student_penalties'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'))
    student_name = db.Column(db.String(100), nullable=False)
    volunteer_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    volunteer_name = db.Column(db.String(100))
    hours = db.Column(db.Integer, default=2)
    multiplier = db.Column(db.Integer, default=1)
    workoff_status = db.Column(db.String(20), default='pending')  # pending, done, overdue
    description = db.Column(db.Text)
    date_issued = db.Column(db.DateTime, default=_naive_utcnow)
    date_worked_off = db.Column(db.DateTime)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'), nullable=False)


class PenaltyHistory(db.Model):
    __tablename__ = 'penalty_history'
    id = db.Column(db.Integer, primary_key=True)
    penalty_id = db.Column(db.Integer, db.ForeignKey('student_penalties.id', ondelete='CASCADE'), nullable=False)
    old_status = db.Column(db.String(20))
    new_status = db.Column(db.String(20), nullable=False)
    old_hours = db.Column(db.Integer)
    new_hours = db.Column(db.Integer)
    actor_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    actor_nick = db.Column(db.String(100))
    actor_name = db.Column(db.String(100))
    comment = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class ActionLog(db.Model):
    __tablename__ = 'action_logs'
    id = db.Column(db.Integer, primary_key=True)
    actor_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    actor_nick = db.Column(db.String(100))
    actor_name = db.Column(db.String(100))
    action = db.Column(db.String(80), nullable=False)
    entity = db.Column(db.String(50), nullable=False)
    entity_id = db.Column(db.Integer)
    description = db.Column(db.Text)
    payload = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class TelegramAccount(db.Model):
    __tablename__ = 'telegram_accounts'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, unique=True)
    telegram_username = db.Column(db.String(100), nullable=False)
    telegram_user_id = db.Column(db.String(100), unique=True)
    telegram_chat_id = db.Column(db.String(100))
    is_linked = db.Column(db.Boolean, default=False)
    linked_at = db.Column(db.DateTime)
    last_photo_sync_at = db.Column(db.DateTime)
    photo_file_id = db.Column(db.String(255))
    photo_url = db.Column(db.Text)
    delivery_enabled = db.Column(db.Boolean, default=True)
    last_delivery_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    updated_at = db.Column(db.DateTime, default=_naive_utcnow, onupdate=_naive_utcnow)


class NotificationEvent(db.Model):
    __tablename__ = 'notification_events'
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(50), nullable=False)
    priority = db.Column(db.String(20), default='normal')
    status = db.Column(db.String(20), default='draft')
    scheduled_for = db.Column(db.DateTime)
    sent_at = db.Column(db.DateTime)
    cancelled_at = db.Column(db.DateTime)
    processing_started_at = db.Column(db.DateTime)
    recipient_user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'))
    payload = db.Column(db.Text)
    dedupe_key = db.Column(db.String(255))
    source_entity = db.Column(db.String(50))
    source_entity_id = db.Column(db.Integer)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class NotificationDelivery(db.Model):
    __tablename__ = 'notification_deliveries'
    id = db.Column(db.Integer, primary_key=True)
    notification_id = db.Column(db.Integer, db.ForeignKey('notification_events.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    telegram_chat_id = db.Column(db.String(100))
    delivery_status = db.Column(db.String(20), default='pending')
    error = db.Column(db.Text)
    message_id = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=_naive_utcnow)


class Broadcast(db.Model):
    __tablename__ = 'broadcasts'
    id = db.Column(db.Integer, primary_key=True)
    author_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'))
    text = db.Column(db.Text, nullable=False)
    filters = db.Column(db.Text)
    priority = db.Column(db.String(20), default='normal')
    status = db.Column(db.String(20), default='draft')
    is_anonymous = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    updated_at = db.Column(db.DateTime, default=_naive_utcnow, onupdate=_naive_utcnow)


class DashboardNote(db.Model):
    __tablename__ = 'dashboard_notes'
    id = db.Column(db.Integer, primary_key=True)
    author_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    pool_id = db.Column(db.Integer, db.ForeignKey('pools.id'))
    text = db.Column(db.Text, nullable=False)
    is_pinned = db.Column(db.Boolean, default=False)
    is_highlighted = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    is_anonymous = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=_naive_utcnow)
    updated_at = db.Column(db.DateTime, default=_naive_utcnow, onupdate=_naive_utcnow)


class AppSetting(db.Model):
    __tablename__ = 'app_settings'
    key = db.Column(db.String(100), primary_key=True)
    value = db.Column(db.Text)
    updated_at = db.Column(db.DateTime, default=_naive_utcnow, onupdate=_naive_utcnow)


def _actor_snapshot(user):
    if not user:
        return None, None, None
    return user.id, user.nick, user.name or user.nick


def log_action(action, entity, entity_id=None, description='', payload=None, actor=None):
    actor = actor or getattr(g, 'user', None)
    actor_id, actor_nick, actor_name = _actor_snapshot(actor)
    contextual_payload = dict(payload or {})
    contextual_payload.setdefault('pool_id', active_pool_id())
    db.session.add(ActionLog(
        actor_id=actor_id,
        actor_nick=actor_nick,
        actor_name=actor_name,
        action=action,
        entity=entity,
        entity_id=entity_id,
        description=description,
        payload=json.dumps(contextual_payload, ensure_ascii=False),
    ))


TELEGRAM_SETTINGS_DEFAULTS = {
    'test_mode': TELEGRAM_TEST_MODE,
    'quiet_hours_start': TELEGRAM_QUIET_HOURS_START,
    'quiet_hours_end': TELEGRAM_QUIET_HOURS_END,
}


def _parse_bool_setting(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}


def _parse_int_setting(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _get_setting_row(key):
    return AppSetting.query.filter_by(key=key).first()


def get_app_setting(key, default=None):
    row = _get_setting_row(key)
    if not row:
        return default
    return row.value


def set_app_setting(key, value):
    row = _get_setting_row(key)
    if not row:
        row = AppSetting(key=key)
        db.session.add(row)
    row.value = '' if value is None else str(value)
    return row


def get_telegram_settings():
    return {
        'test_mode': _parse_bool_setting(
            get_app_setting('telegram.test_mode', TELEGRAM_SETTINGS_DEFAULTS['test_mode']),
            TELEGRAM_SETTINGS_DEFAULTS['test_mode'],
        ),
        'quiet_hours_start': _parse_int_setting(
            get_app_setting('telegram.quiet_hours_start', TELEGRAM_SETTINGS_DEFAULTS['quiet_hours_start']),
            TELEGRAM_SETTINGS_DEFAULTS['quiet_hours_start'],
        ),
        'quiet_hours_end': _parse_int_setting(
            get_app_setting('telegram.quiet_hours_end', TELEGRAM_SETTINGS_DEFAULTS['quiet_hours_end']),
            TELEGRAM_SETTINGS_DEFAULTS['quiet_hours_end'],
        ),
    }


def update_telegram_settings(payload):
    current = get_telegram_settings()
    next_settings = current.copy()
    changes = {}

    if 'test_mode' in payload:
        new_value = bool(payload.get('test_mode'))
        if current['test_mode'] != new_value:
            changes['test_mode'] = {'from': current['test_mode'], 'to': new_value}
            next_settings['test_mode'] = new_value

    for field in ('quiet_hours_start', 'quiet_hours_end'):
        if field not in payload:
            continue
        try:
            new_value = int(payload.get(field))
        except (TypeError, ValueError):
            raise ValueError('Тихие часы должны быть целыми часами от 0 до 23')
        if new_value < 0 or new_value > 23:
            raise ValueError('Тихие часы должны быть в диапазоне от 0 до 23')
        if current[field] != new_value:
            changes[field] = {'from': current[field], 'to': new_value}
            next_settings[field] = new_value

    for key, value in next_settings.items():
        set_app_setting(f'telegram.{key}', value)

    return next_settings, changes


def add_penalty_history(penalty, old_status, new_status, old_hours, comment='', actor=None):
    actor = actor or getattr(g, 'user', None)
    actor_id, actor_nick, actor_name = _actor_snapshot(actor)
    db.session.add(PenaltyHistory(
        penalty_id=penalty.id,
        old_status=old_status,
        new_status=new_status,
        old_hours=old_hours,
        new_hours=penalty.hours * penalty.multiplier,
        actor_id=actor_id,
        actor_nick=actor_nick,
        actor_name=actor_name,
        comment=comment,
    ))


def _clean_telegram_username(value):
    username = (value or '').strip().replace('@', '')
    return username


def _normalize_telegram_value(value):
    username = _clean_telegram_username(value)
    return f'@{username}' if username else None


def _telegram_account_for_user(user_id):
    if not user_id:
        return None
    return TelegramAccount.query.filter_by(user_id=user_id, is_linked=True).first()


def _telegram_account_any(user_id):
    if not user_id:
        return None
    return TelegramAccount.query.filter_by(user_id=user_id).first()


def _avatar_url_for_user(user):
    if not user:
        return None
    account = _telegram_account_any(user.id)
    if not account or not (account.photo_file_id or account.photo_url):
        return None
    version = int(account.last_photo_sync_at.timestamp()) if account.last_photo_sync_at else 0
    return f'/api/users/{user.id}/avatar?v={version}'


def _telegram_is_configured():
    return bool(TELEGRAM_BOT_TOKEN)


def _telegram_file_download_url(file_path):
    if not TELEGRAM_BOT_TOKEN or not file_path:
        return None
    return f'https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}'


def _download_telegram_photo_bytes(account):
    if not account:
        return None, None
    if account.photo_url and account.photo_url.startswith('data:'):
        header, _, data = account.photo_url.partition(',')
        if not data:
            return None, None
        mime = 'image/jpeg'
        if ';base64' in header:
            mime = header[5:].split(';', 1)[0] or mime
        try:
            import base64
            return base64.b64decode(data), mime
        except Exception:
            return None, None
    if not account.photo_file_id:
        return None, None
    import requests

    file_info = telegram_get('getFile', {'file_id': account.photo_file_id})
    file_path = (file_info or {}).get('file_path')
    if not file_path:
        return None, None
    download_url = _telegram_file_download_url(file_path)
    if not download_url:
        return None, None
    response = requests.get(download_url, timeout=30)
    response.raise_for_status()
    account.photo_url = file_path
    return response.content, response.headers.get('Content-Type') or 'image/jpeg'


def _telegram_link_status(user):
    if not user:
        return {
            'username': '',
            'linked': False,
            'delivery_enabled': False,
            'last_photo_sync_at': None,
            'linked_at': None,
        }
    username = _clean_telegram_username(user.telegram)
    account = _telegram_account_any(user.id)
    linked = bool(account and account.is_linked)
    return {
        'username': f'@{account.telegram_username}' if account and account.telegram_username else (f'@{username}' if username else ''),
        'linked': linked,
        'configured': _telegram_is_configured(),
        'delivery_enabled': bool(account.delivery_enabled) if account else False,
        'last_photo_sync_at': account.last_photo_sync_at.isoformat() if account and account.last_photo_sync_at else None,
        'linked_at': account.linked_at.isoformat() if account and account.linked_at else None,
        'photo_url': account.photo_url if account else None,
        'needs_username': not username,
        'bot_username': f'@{TELEGRAM_BOT_USERNAME}' if TELEGRAM_BOT_USERNAME else '',
        'bot_avatar_url': '/api/telegram/bot-avatar' if _telegram_is_configured() and TELEGRAM_BOT_USERNAME else '',
    }


def _telegram_bot_avatar_payload():
    now = _utcnow()
    expires_at = _telegram_bot_avatar_cache.get('expires_at')
    if (
        _telegram_bot_avatar_cache.get('bytes')
        and expires_at
        and expires_at > now
    ):
        return _telegram_bot_avatar_cache.get('bytes'), _telegram_bot_avatar_cache.get('content_type') or 'image/jpeg'

    if not _telegram_is_configured():
        return None, None

    bot_info = telegram_get('getMe')
    bot_id = bot_info.get('id') if isinstance(bot_info, dict) else None
    if not bot_id:
        return None, None

    photos = telegram_get('getUserProfilePhotos', {'user_id': bot_id, 'limit': 1})
    total = photos.get('total_count', 0) if isinstance(photos, dict) else 0
    if not total or not photos.get('photos'):
        return None, None

    best = photos['photos'][0][-1]
    file_id = best.get('file_id')
    if not file_id:
        return None, None

    file_info = telegram_get('getFile', {'file_id': file_id})
    file_path = file_info.get('file_path') if isinstance(file_info, dict) else None
    download_url = _telegram_file_download_url(file_path)
    if not download_url:
        return None, None

    response = requests.get(download_url, timeout=30)
    response.raise_for_status()
    payload = response.content
    content_type = response.headers.get('Content-Type') or 'image/jpeg'
    _telegram_bot_avatar_cache['bytes'] = payload
    _telegram_bot_avatar_cache['content_type'] = content_type
    _telegram_bot_avatar_cache['expires_at'] = now + timedelta(hours=1)
    return payload, content_type


def _dashboard_note_to_dict(note):
    author = db.session.get(User, note.author_id) if note.author_id else None
    return {
        'id': note.id,
        'pool_id': note.pool_id,
        'text': note.text,
        'is_pinned': bool(note.is_pinned),
        'is_highlighted': bool(note.is_highlighted),
        'is_active': bool(note.is_active),
        'is_anonymous': bool(note.is_anonymous),
        'author_nick': '' if note.is_anonymous else (author.nick if author else ''),
        'author_name': '' if note.is_anonymous else (author.name if author else ''),
        'created_at': note.created_at.isoformat() if note.created_at else None,
        'updated_at': note.updated_at.isoformat() if note.updated_at else None,
    }


def _broadcast_runtime_status(broadcast):
    events = NotificationEvent.query.filter_by(source_entity='broadcast', source_entity_id=broadcast.id).all()
    if not events:
        return broadcast.status or 'draft'
    statuses = [event.status for event in events]
    if any(status == 'error' for status in statuses):
        return 'error'
    if any(status in ('queued', 'pending') for status in statuses):
        return 'queued'
    if all(status == 'cancelled' for status in statuses):
        return 'cancelled'
    if any(status == 'sent' for status in statuses):
        return 'sent'
    if all(status == 'skipped' for status in statuses):
        return 'skipped'
    return broadcast.status or 'draft'


def _broadcast_to_dict(broadcast):
    author = db.session.get(User, broadcast.author_id) if broadcast.author_id else None
    filters = json.loads(broadcast.filters or '{}')
    return {
        'id': broadcast.id,
        'text': broadcast.text,
        'filters': filters,
        'priority': broadcast.priority,
        'status': _broadcast_runtime_status(broadcast),
        'is_anonymous': bool(broadcast.is_anonymous),
        'author_nick': '' if broadcast.is_anonymous else (author.nick if author else ''),
        'author_name': '' if broadcast.is_anonymous else (author.name if author else ''),
        'created_at': broadcast.created_at.isoformat() if broadcast.created_at else None,
        'updated_at': broadcast.updated_at.isoformat() if broadcast.updated_at else None,
    }


def _unlinked_users_for_pool(pool_id):
    linked_ids = {
        row.user_id for row in TelegramAccount.query.filter_by(is_linked=True).with_entities(TelegramAccount.user_id).all()
    }
    volunteer_ids = {
        row.user_id for row in PoolVolunteer.query.filter_by(pool_id=pool_id).with_entities(PoolVolunteer.user_id).all()
    } if pool_id else set()
    candidate_ids = volunteer_ids
    if not candidate_ids:
        return []
    users = User.query.filter(User.id.in_(candidate_ids)).order_by(User.role, User.nick).all()
    result = []
    for user in users:
        username = _clean_telegram_username(user.telegram)
        if user.id in linked_ids:
            continue
        result.append({
            'id': user.id,
            'nick': user.nick,
            'name': user.name or user.nick,
            'role': _effective_access_role(user, pool_id),
            'telegram': f'@{username}' if username else '',
            'needs_username': not bool(username),
            'avatar_url': _avatar_url_for_user(user),
        })
    return result


def _linked_users_for_pool(pool_id):
    volunteer_ids = {
        row.user_id for row in PoolVolunteer.query.filter_by(pool_id=pool_id).with_entities(PoolVolunteer.user_id).all()
    } if pool_id else set()
    candidate_ids = volunteer_ids
    if not candidate_ids:
        return []
    rows = (
        db.session.query(User, TelegramAccount)
        .join(TelegramAccount, TelegramAccount.user_id == User.id)
        .filter(
            User.id.in_(candidate_ids),
            TelegramAccount.is_linked.is_(True),
        )
        .order_by(User.role, User.nick)
        .all()
    )
    return [{
        'id': user.id,
        'nick': user.nick,
        'name': user.name or user.nick,
        'role': _effective_access_role(user, pool_id),
        'telegram': f'@{account.telegram_username}' if account.telegram_username else (user.telegram or ''),
        'linked_at': account.linked_at.isoformat() if account.linked_at else None,
        'delivery_enabled': bool(account.delivery_enabled),
        'avatar_url': _avatar_url_for_user(user),
    } for user, account in rows]


def telegram_api(method, payload=None):
    if not TELEGRAM_BOT_TOKEN:
        raise RuntimeError('TELEGRAM_BOT_TOKEN не задан')
    import requests
    response = requests.post(
        f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/{method}',
        json=payload or {},
        timeout=TELEGRAM_API_TIMEOUT,
    )
    response.raise_for_status()
    data = response.json()
    if not data.get('ok'):
        raise RuntimeError(f'Telegram API error in {method}: {data}')
    return data.get('result')


def telegram_get(method, params=None):
    if not TELEGRAM_BOT_TOKEN:
        raise RuntimeError('TELEGRAM_BOT_TOKEN не задан')
    import requests
    response = requests.get(
        f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/{method}',
        params=params or {},
        timeout=TELEGRAM_API_TIMEOUT,
    )
    response.raise_for_status()
    data = response.json()
    if not data.get('ok'):
        raise RuntimeError(f'Telegram API error in {method}: {data}')
    return data.get('result')


def telegram_is_quiet_hours(now=None):
    now = now or datetime.now()
    settings = get_telegram_settings()
    quiet_start = settings['quiet_hours_start']
    quiet_end = settings['quiet_hours_end']
    hour = now.hour
    if quiet_start == quiet_end:
        return False
    if quiet_start < quiet_end:
        return quiet_start <= hour < quiet_end
    return hour >= quiet_start or hour < quiet_end


def telegram_send_message(chat_id, text, disable_notification=False, reply_markup=None):
    payload = {
        'chat_id': chat_id,
        'text': text,
        'disable_notification': disable_notification,
    }
    if reply_markup:
        payload['reply_markup'] = reply_markup
    return telegram_api('sendMessage', payload)


def telegram_sync_commands():
    commands = [
        {'command': 'start', 'description': 'Привязать бота к платформе'},
        {'command': 'myshifts', 'description': 'Посмотреть свои смены'},
        {'command': 'rules', 'description': 'Правила школы'},
        {'command': 'penalties', 'description': 'Ученики с нарушениями'},
        {'command': 'responsibles', 'description': 'Ответственные за бассейн'},
        {'command': 'help', 'description': 'Помощь'},
    ]
    return telegram_api('setMyCommands', {'commands': commands})


def telegram_delete_message(chat_id, message_id):
    return telegram_api('deleteMessage', {
        'chat_id': chat_id,
        'message_id': message_id,
    })


def telegram_answer_callback(callback_query_id, text='Готово'):
    return telegram_api('answerCallbackQuery', {
        'callback_query_id': callback_query_id,
        'text': text,
    })


def _utcnow():
    return _naive_utcnow()


def _moscow_now():
    return _utcnow() + MOSCOW_OFFSET


def _moscow_to_utc(value):
    return value - MOSCOW_OFFSET


def _frontend_base_url():
    return (frontend_urls[0] if frontend_urls else '').rstrip('/')


def _exam_brief_url():
    if EXAM_BRIEF_URL:
        return EXAM_BRIEF_URL
    base = _frontend_base_url()
    return f'{base}/exam-brief' if base else '/exam-brief'


def _tg_link(user):
    username = _clean_telegram_username(user.telegram)
    return f'@{username}' if username else f'@{user.nick}'


def _tg_url(value):
    username = _clean_telegram_username(value)
    return f'https://t.me/{username}' if username else ''


def _telegram_main_keyboard(user):
    if not user or user.role == 'admin':
        return {'remove_keyboard': True}
    return {
        'keyboard': [
            [{'text': 'Мои смены'}],
        ],
        'resize_keyboard': True,
        'persistent': True,
    }


def _format_telegram_shift_day(date_value, today):
    label = date_value.strftime('%A %d.%m.%Y').capitalize()
    if date_value == today:
        return f'{label} · сегодня'
    if date_value == today + timedelta(days=1):
        return f'{label} · завтра'
    return label


def _format_shift(block):
    label = f' · {block.label}' if block.label else ''
    return f'{block.date.strftime("%d.%m")} {block.time_start}-{block.time_end}{label}'


def _admin_team_leads(pool_id=None):
    query = User.query.filter(User.active.is_(True), User.role.in_(['admin', 'team_lead']))
    return query.order_by(User.role, User.nick).all()


def _pool_responsible_users(pool_id):
    if not pool_id:
        return []
    return (
        db.session.query(User)
        .join(PoolVolunteer, PoolVolunteer.user_id == User.id)
        .filter(
            PoolVolunteer.pool_id == pool_id,
            PoolVolunteer.pool_role.in_(['responsible_admin', 'responsible_team_lead']),
            PoolVolunteer.notifications_enabled.isnot(False),
            User.active.is_(True),
        )
        .order_by(User.role, User.nick)
        .all()
    )


def _pool_responsibles(pool_id):
    if not pool_id:
        return []
    rows = (
        db.session.query(User, PoolVolunteer)
        .join(PoolVolunteer, PoolVolunteer.user_id == User.id)
        .filter(
            PoolVolunteer.pool_id == pool_id,
            PoolVolunteer.pool_role.in_(['responsible_admin', 'responsible_team_lead']),
            User.active.is_(True),
        )
        .order_by(User.role, User.nick)
        .all()
    )
    return [{
        'id': user.id,
        'nick': user.nick,
        'name': user.name or user.nick,
        'role': user.role,
        'telegram': user.telegram or '',
        'telegram_url': _tg_url(user.telegram),
        'avatar_url': _avatar_url_for_user(user),
        'pool_role': pv.pool_role,
        'notifications_enabled': pv.notifications_enabled is not False,
    } for user, pv in rows]


def _pool_notifications_enabled_for_user(user_id, pool_id):
    if not user_id or not pool_id:
        return True
    relation = PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=user_id).first()
    if not relation or relation.pool_role not in ('responsible_admin', 'responsible_team_lead'):
        return True
    return relation.notifications_enabled is not False


def _signed_users_for_block(block):
    return (
        db.session.query(User)
        .join(Signup, Signup.user_id == User.id)
        .filter(Signup.block_id == block.id, User.active.is_(True))
        .order_by(User.nick)
        .all()
    )


def _users_on_shift(pool_id, moment_msk=None):
    moment_msk = moment_msk or _moscow_now()
    today = moment_msk.date()
    current_time = moment_msk.strftime('%H:%M')
    blocks = (
        ShiftBlock.query
        .filter(
            ShiftBlock.pool_id == pool_id,
            ShiftBlock.date == today,
            ShiftBlock.time_start <= current_time,
            ShiftBlock.time_end >= current_time,
        )
        .all()
    )
    users = {}
    for block in blocks:
        for user in _signed_users_for_block(block):
            users[user.id] = user
    return list(users.values())


def _pool_start_date(pool_id):
    pool = db.session.get(Pool, pool_id) if pool_id else None
    return pool.start_date if pool and pool.start_date else None


def _tribe_masters_for_pool(pool_id, tribe):
    normalized_tribe = normalize_tribe(tribe)
    if not pool_id or not normalized_tribe:
        return []
    rows = (
        db.session.query(User)
        .join(PoolVolunteer, PoolVolunteer.user_id == User.id)
        .filter(
            User.active.is_(True),
            PoolVolunteer.pool_id == pool_id,
            PoolVolunteer.pool_role == 'tribe_master',
            PoolVolunteer.tribe == normalized_tribe,
        )
        .order_by(User.nick)
        .all()
    )
    return rows


def _queue_notification(user, event_type, text, dedupe_key, pool_id=None, priority='normal',
                        scheduled_for=None, payload=None, source_entity=None, source_entity_id=None,
                        created_by=None, action_buttons=None):
    if not user or not dedupe_key:
        return None
    if pool_id and not _pool_notifications_enabled_for_user(user.id, pool_id):
        return None
    if pool_id is not None and not dedupe_key.startswith(f'pool:{pool_id}:'):
        dedupe_key = f'pool:{pool_id}:{dedupe_key}'
    existing = NotificationEvent.query.filter_by(dedupe_key=dedupe_key).first()
    if existing:
        return existing
    full_payload = payload.copy() if isinstance(payload, dict) else {}
    full_payload['text'] = text
    if action_buttons:
        full_payload['action_buttons'] = action_buttons
    event = NotificationEvent(
        type=event_type,
        priority=priority,
        status='queued',
        scheduled_for=scheduled_for,
        recipient_user_id=user.id,
        pool_id=pool_id,
        payload=json.dumps(full_payload, ensure_ascii=False),
        dedupe_key=dedupe_key,
        source_entity=source_entity,
        source_entity_id=source_entity_id,
        created_by=created_by,
    )
    db.session.add(event)
    return event


def _cancel_pending_notifications(source_entity, source_entity_id, event_types=None):
    query = NotificationEvent.query.filter(
        NotificationEvent.source_entity == source_entity,
        NotificationEvent.source_entity_id == source_entity_id,
        NotificationEvent.status.in_(['queued', 'pending', 'processing']),
    )
    if event_types:
        query = query.filter(NotificationEvent.type.in_(event_types))
    for event in query.all():
        event.status = 'cancelled'
        event.cancelled_at = _utcnow()


def sync_telegram_photo(account, telegram_user_id):
    try:
        photos = telegram_get('getUserProfilePhotos', {'user_id': telegram_user_id, 'limit': 1})
    except Exception:
        account.last_photo_sync_at = _utcnow()
        return

    total = photos.get('total_count', 0) if isinstance(photos, dict) else 0
    if total and photos.get('photos'):
        best = photos['photos'][0][-1]
        account.photo_file_id = best.get('file_id')
        account.photo_url = f'/api/users/{account.user_id}/avatar'
    account.last_photo_sync_at = _utcnow()


def normalize_tg_username(value):
    return _clean_telegram_username(value).lower()


def telegram_username_conflict(value, exclude_user_id=None):
    normalized = normalize_tg_username(value)
    if not normalized:
        return None
    variants = [normalized, f'@{normalized}']
    query = User.query.filter(db.func.lower(User.telegram).in_(variants))
    if exclude_user_id is not None:
        query = query.filter(User.id != exclude_user_id)
    return query.first()


def find_platform_user_by_username(username):
    normalized = normalize_tg_username(username)
    if not normalized:
        return None
    variants = [normalized, f'@{normalized}']
    matches = User.query.filter(
        db.func.lower(User.telegram).in_(variants)
    ).limit(2).all()
    return matches[0] if len(matches) == 1 else None


def upsert_telegram_account(user, tg_user, chat_id):
    telegram_user_id = str(tg_user.get('id') or '') or None
    if telegram_user_id:
        conflict = TelegramAccount.query.filter(
            TelegramAccount.telegram_user_id == telegram_user_id,
            TelegramAccount.user_id != user.id,
        ).first()
        if conflict:
            raise ValueError('Этот Telegram аккаунт уже привязан к другому пользователю')
    account = TelegramAccount.query.filter_by(user_id=user.id).first()
    if not account:
        account = TelegramAccount(user_id=user.id)
        db.session.add(account)
    account.telegram_username = normalize_tg_username(tg_user.get('username'))
    account.telegram_user_id = telegram_user_id
    account.telegram_chat_id = str(chat_id)
    account.is_linked = True
    account.linked_at = account.linked_at or _utcnow()
    account.delivery_enabled = True
    sync_telegram_photo(account, tg_user.get('id'))
    user.telegram = f'@{account.telegram_username}'
    return account


def telegram_link_account(chat_id, tg_user):
    username = normalize_tg_username(tg_user.get('username'))
    if not username:
        telegram_send_message(
            chat_id,
            'У тебя не указан Telegram username. Пожалуйста, добавь @username в настройках Telegram и повтори /start.',
        )
        return {'linked': False, 'reason': 'missing_username'}

    user = find_platform_user_by_username(username)
    if not user:
        telegram_send_message(
            chat_id,
            'Я не нашел тебя в системе бассейна по этому username. Попроси администратора или Team Lead указать твой Telegram username на платформе.',
        )
        return {'linked': False, 'reason': 'user_not_found'}

    account = upsert_telegram_account(user, tg_user, chat_id)
    log_action(
        'link',
        'telegram_account',
        user.id,
        'Пользователь привязал Telegram-бота',
        {'telegram_username': f'@{account.telegram_username}'},
        actor=user,
    )
    greeting = [
        'Привязка прошла успешно.',
        f'Теперь я буду присылать уведомления для @{user.nick}.',
        '',
        'Команды:',
        '/myshifts — мои смены',
        '/rules — правила школы',
        '/responsibles — ответственные за бассейн',
        '/help — помощь',
    ]
    telegram_send_message(chat_id, '\n'.join(greeting), reply_markup=_telegram_main_keyboard(user))
    return {'linked': True, 'user_id': user.id, 'telegram_username': f'@{account.telegram_username}'}


def telegram_handle_help(chat_id, tg_user=None):
    user = _telegram_user_from_tg(tg_user) if tg_user else None
    pool_id = active_pool_id()
    responsibles = _pool_responsibles(pool_id)
    lines = [
        'Если что-то не работает, напиши:',
        TELEGRAM_SUPPORT_CONTACT,
        '',
        'Шаблон заявки на фикс:',
        '1. Что именно не работает.',
        '2. Где это произошло: платформа или Telegram-бот.',
        '3. Что ты нажал перед ошибкой.',
        '4. Какой результат ожидался.',
        '5. Что произошло по факту.',
        '6. Пришли скриншот или запись экрана.',
    ]
    if responsibles:
        lines.extend(['', 'Ответственные за бассейн:'])
        for person in responsibles:
            contact = person['telegram_url'] or f'@{person["nick"]}'
            lines.append(f'• @{person["nick"]} — {contact}')
    lines.extend([
        '',
        'Команды:',
        '/start — привязать бота к платформе',
        '/myshifts — показать мои смены',
        '/rules — открыть правила школы',
        '/penalties — ученики с нарушениями по статусам',
        '/responsibles — ответственные за бассейн',
        '/help — показать это сообщение',
    ])
    telegram_send_message(
        chat_id,
        '\n'.join(lines),
        reply_markup=_telegram_main_keyboard(user),
    )
    return {'ok': True}


def telegram_handle_responsibles(chat_id):
    pool_id = active_pool_id()
    responsibles = _pool_responsibles(pool_id)
    if not responsibles:
        telegram_send_message(chat_id, 'Для активного бассейна пока не назначены ответственные.')
        return {'ok': True, 'count': 0}

    lines = ['Ответственные за текущий бассейн:']
    for person in responsibles:
        role = 'Админ' if person.get('role') == 'admin' else 'Тимлид'
        contact = person.get('telegram_url') or person.get('telegram') or 'tg не указан'
        lines.append(f'• @{person["nick"]} · {role} · {contact}')
    telegram_send_message(chat_id, '\n'.join(lines))
    return {'ok': True, 'count': len(responsibles)}


def _telegram_user_from_tg(tg_user):
    username = normalize_tg_username(tg_user.get('username'))
    if not username:
        return None
    account = TelegramAccount.query.filter_by(telegram_username=username, is_linked=True).first()
    if account:
        return db.session.get(User, account.user_id)
    return find_platform_user_by_username(username)


def telegram_handle_penalties(chat_id, tg_user):
    user = _telegram_user_from_tg(tg_user)
    if not user:
        telegram_send_message(chat_id, 'Сначала привяжи бота командой /start.')
        return {'ok': False, 'reason': 'not_linked'}
    pool_id = active_pool_id()
    if not pool_id:
        telegram_send_message(chat_id, 'Сейчас нет активного бассейна, поэтому список штрафов недоступен.')
        return {'ok': False, 'reason': 'no_active_pool'}
    if not _can_access_pool_id(user, pool_id):
        telegram_send_message(chat_id, 'У тебя нет доступа к активному бассейну, поэтому список штрафов недоступен.')
        return {'ok': False, 'reason': 'no_pool_access'}
    penalties = StudentPenalty.query.filter_by(pool_id=pool_id).order_by(StudentPenalty.student_name).all()
    groups = {
        'pending': ('Нарушили, но ещё не отрабатывают', []),
        'in_workoff': ('На отработке', []),
        'awaiting_unlock': ('Ждут разблокировки', []),
    }
    for penalty in penalties:
        if penalty.workoff_status in groups:
            groups[penalty.workoff_status][1].append(f'• {penalty.student_name} ({penalty.hours * penalty.multiplier}h)')
    lines = ['Ученики с нарушениями:']
    for _, (title, items) in groups.items():
        lines.append('')
        lines.append(title + ':')
        lines.extend(items[:20] or ['нет'])
        if len(items) > 20:
            lines.append(f'...и ещё {len(items) - 20}')
    telegram_send_message(chat_id, '\n'.join(lines))
    return {'ok': True, 'action': 'penalties'}


def telegram_handle_my_shifts(chat_id, tg_user):
    user = _telegram_user_from_tg(tg_user)
    if not user:
        telegram_send_message(chat_id, 'Сначала привяжи бота командой /start.')
        return {'ok': False, 'reason': 'not_linked'}
    pool_id = active_pool_id()
    if not pool_id:
        telegram_send_message(chat_id, 'Сейчас нет активного бассейна.')
        return {'ok': False, 'reason': 'no_active_pool'}
    if not _can_access_pool_id(user, pool_id):
        telegram_send_message(chat_id, 'У тебя нет доступа к активному бассейну.')
        return {'ok': False, 'reason': 'no_pool_access'}
    if _effective_access_role(user, pool_id) == 'admin':
        telegram_send_message(
            chat_id,
            'Для админа кнопка «Мои смены» не нужна.',
            reply_markup=_telegram_main_keyboard(user),
        )
        return {'ok': False, 'reason': 'admin_not_supported'}

    today = _moscow_now().date()
    rows = (
        db.session.query(ShiftBlock, Pool)
        .join(Signup, Signup.block_id == ShiftBlock.id)
        .join(Pool, Pool.id == ShiftBlock.pool_id)
        .filter(
            Signup.user_id == user.id,
            ShiftBlock.pool_id == pool_id,
            ShiftBlock.date >= today,
            Pool.active.is_(True),
        )
        .order_by(ShiftBlock.date, ShiftBlock.time_start)
        .all()
    )

    if not rows:
        telegram_send_message(
            chat_id,
            'У тебя пока нет записей на будущие смены.',
            reply_markup=_telegram_main_keyboard(user),
        )
        return {'ok': True, 'count': 0, 'action': 'my_shifts'}

    grouped = {}
    for block, pool in rows[:20]:
        grouped.setdefault(block.date, []).append((block, pool))

    lines = ['Твои ближайшие смены:']
    for shift_date, day_rows in grouped.items():
        lines.append('')
        lines.append(_format_telegram_shift_day(shift_date, today))
        for block, pool in day_rows:
            label = f' · {block.label}' if block.label else ''
            pool_name = f' · {pool.name}' if pool and pool.name else ''
            lines.append(f'• {block.time_start}-{block.time_end}{label}{pool_name}')
    if len(rows) > 20:
        lines.append(f'...и ещё {len(rows) - 20}')

    telegram_send_message(
        chat_id,
        '\n'.join(lines),
        reply_markup=_telegram_main_keyboard(user),
    )
    return {'ok': True, 'count': len(rows), 'action': 'my_shifts'}


def telegram_handle_message(message):
    chat = message.get('chat') or {}
    chat_id = chat.get('id')
    tg_user = message.get('from') or {}
    text = (message.get('text') or '').strip()
    if not chat_id or not text:
        return {'ok': False, 'reason': 'empty_message'}
    normalized_text = text.lower()

    if text.startswith('/start'):
        return telegram_link_account(chat_id, tg_user)
    if text.startswith('/myshifts') or normalized_text == 'мои смены':
        return telegram_handle_my_shifts(chat_id, tg_user)
    if text.startswith('/rules'):
        telegram_send_message(chat_id, f'Правила школы: {SCHOOL_RULES_URL}')
        return {'ok': True, 'action': 'rules'}
    if text.startswith('/penalties') or text.startswith('/violations'):
        return telegram_handle_penalties(chat_id, tg_user)
    if text.startswith('/responsibles'):
        return telegram_handle_responsibles(chat_id)
    if text.startswith('/help'):
        return telegram_handle_help(chat_id, tg_user)

    telegram_send_message(chat_id, 'Я понимаю команды /start, /myshifts, /rules, /penalties, /responsibles и /help.')
    return {'ok': True, 'action': 'unknown_command'}


def build_notification_reply_markup(event):
    payload = json.loads(event.payload or '{}')
    buttons = payload.get('action_buttons') or []
    if not buttons:
        return None
    telegram_buttons = []
    for button in buttons:
        telegram_button = {'text': button.get('text') or 'Открыть'}
        if button.get('url'):
            telegram_button['url'] = button['url']
        elif button.get('callback_data'):
            telegram_button['callback_data'] = button['callback_data']
        else:
            continue
        telegram_buttons.append(telegram_button)
    return {'inline_keyboard': [telegram_buttons]} if telegram_buttons else None


def build_notification_text(event):
    payload = json.loads(event.payload or '{}')
    text = (payload.get('text') or '').strip()
    if not text:
        text = f'Новое уведомление типа {event.type}.'
    if event.source_entity == 'broadcast' and event.source_entity_id:
        broadcast = db.session.get(Broadcast, event.source_entity_id)
        if broadcast and not broadcast.is_anonymous:
            author = db.session.get(User, broadcast.author_id) if broadcast.author_id else None
            if author:
                text = f'{text}\n\nОтправил: {author.name or author.nick} (@{author.nick})'
    if get_telegram_settings()['test_mode']:
        text = f'[TEST MODE]\n{text}'
    return text


def _delete_related_penalty_messages(penalty_id, event_type, exclude_event_id=None):
    events = NotificationEvent.query.filter_by(
        source_entity='penalty',
        source_entity_id=penalty_id,
        type=event_type,
    ).all()
    for event in events:
        if exclude_event_id and event.id == exclude_event_id:
            pass
        if event.status in ('queued', 'pending'):
            event.status = 'cancelled'
            event.cancelled_at = _utcnow()
        deliveries = NotificationDelivery.query.filter_by(notification_id=event.id, delivery_status='sent').all()
        for delivery in deliveries:
            if not delivery.telegram_chat_id or not delivery.message_id:
                continue
            try:
                telegram_delete_message(delivery.telegram_chat_id, delivery.message_id)
            except Exception:
                pass


def _set_penalty_status_from_bot(penalty, new_status, actor=None, comment=''):
    old_status = penalty.workoff_status
    old_hours = penalty.hours * penalty.multiplier
    penalty.workoff_status = new_status
    if new_status == 'in_workoff':
        penalty.date_worked_off = _utcnow()
    if new_status in ('done', 'awaiting_unlock'):
        penalty.date_worked_off = _utcnow()
    if new_status == 'pending':
        penalty.date_worked_off = None
    if old_status != new_status or old_hours != penalty.hours * penalty.multiplier:
        add_penalty_history(penalty, old_status, new_status, old_hours, comment, actor=actor)
        log_action(
            'telegram_update',
            'penalty',
            penalty.id,
            f'Telegram: штраф {penalty.student_name}: {old_status} → {new_status}',
            {
                'student': penalty.student_name,
                'old_status': old_status,
                'new_status': new_status,
                'actor_nick': actor.nick if actor else None,
            },
            actor=actor,
        )
def _notify_admins_penalty_created(penalty):
    events = []
    for user in _pool_responsible_users(penalty.pool_id):
        event = _queue_notification(
            user,
            'penalty_admin_block',
            (
                f'Ученик {penalty.student_name} получил штраф.\n'
                f'Выдал: {penalty.volunteer_name}.\n'
                'Нужно заблокировать ученика на учебной платформе.'
            ),
            f'penalty:{penalty.id}:admin-block:user:{user.id}',
            pool_id=penalty.pool_id,
            priority='urgent',
            source_entity='penalty',
            source_entity_id=penalty.id,
        )
        if event:
            events.append(event)
    return events


def _notify_admins_penalty_awaiting_unlock(penalty):
    events = []
    for user in _pool_responsible_users(penalty.pool_id):
        event = _queue_notification(
            user,
            'penalty_admin_unlock',
            (
                f'Ученик {penalty.student_name} отработал пенальти.\n'
                'Нужно снять ограничения на учебной платформе.'
            ),
            f'penalty:{penalty.id}:admin-unlock:user:{user.id}',
            pool_id=penalty.pool_id,
            priority='urgent',
            source_entity='penalty',
            source_entity_id=penalty.id,
        )
        if not event:
            continue
        db.session.flush()
        payload = json.loads(event.payload or '{}')
        payload['action_buttons'] = [{
            'text': 'Разблокирован',
            'callback_data': f'p:{penalty.id}:u:y:{event.id}',
        }]
        event.payload = json.dumps(payload, ensure_ascii=False)
        events.append(event)
    return events


def _queue_penalty_method_question(penalty, scheduled_for=None, suffix='initial'):
    users = _users_on_shift(penalty.pool_id)
    for user in users:
        event = _queue_notification(
            user,
            'penalty_method_question',
            (
                f'Ученик {penalty.student_name} получил пенальти.\n'
                'Получил ли он метод отработки?'
            ),
            f'penalty:{penalty.id}:method:{suffix}:user:{user.id}',
            pool_id=penalty.pool_id,
            scheduled_for=scheduled_for,
            source_entity='penalty',
            source_entity_id=penalty.id,
            payload={'penalty_id': penalty.id, 'question': 'method'},
            action_buttons=[],
        )
        if event:
            db.session.flush()
            payload = json.loads(event.payload or '{}')
            payload['action_buttons'] = [
                {'text': 'Да', 'callback_data': f'p:{penalty.id}:m:y:{event.id}'},
                {'text': 'Нет', 'callback_data': f'p:{penalty.id}:m:n:{event.id}'},
                {'text': 'Пропустить', 'callback_data': f'p:{penalty.id}:m:s:{event.id}'},
            ]
            event.payload = json.dumps(payload, ensure_ascii=False)


def _queue_penalty_workoff_check(penalty, scheduled_for=None, suffix='initial'):
    users = _users_on_shift(penalty.pool_id)
    for user in users:
        event = _queue_notification(
            user,
            'penalty_workoff_check',
            f'Проверь пенальти: {penalty.student_name} отработал?',
            f'penalty:{penalty.id}:complete:{suffix}:user:{user.id}',
            pool_id=penalty.pool_id,
            scheduled_for=scheduled_for,
            source_entity='penalty',
            source_entity_id=penalty.id,
            payload={'penalty_id': penalty.id, 'question': 'complete'},
            action_buttons=[],
        )
        if event:
            db.session.flush()
            payload = json.loads(event.payload or '{}')
            payload['action_buttons'] = [
                {'text': 'Да', 'callback_data': f'p:{penalty.id}:c:y:{event.id}'},
                {'text': 'Нет', 'callback_data': f'p:{penalty.id}:c:n:{event.id}'},
                {'text': 'Пропустить', 'callback_data': f'p:{penalty.id}:c:s:{event.id}'},
            ]
            event.payload = json.dumps(payload, ensure_ascii=False)


def telegram_handle_callback(callback):
    callback_id = callback.get('id')
    message = callback.get('message') or {}
    tg_user = callback.get('from') or {}
    data = callback.get('data') or ''
    parts = data.split(':')
    if len(parts) != 5 or parts[0] != 'p':
        if callback_id:
            telegram_answer_callback(callback_id, 'Неизвестное действие')
        return {'ok': False, 'reason': 'unknown_callback'}
    _, penalty_id_raw, question, answer, event_id_raw = parts
    try:
        penalty_id = int(penalty_id_raw)
        event_id = int(event_id_raw)
    except ValueError:
        if callback_id:
            telegram_answer_callback(callback_id, 'Некорректное действие')
        return {'ok': False, 'reason': 'bad_callback'}

    account = TelegramAccount.query.filter_by(
        telegram_user_id=str(tg_user.get('id') or ''),
        is_linked=True,
    ).first()
    if not account:
        account = TelegramAccount.query.filter_by(
            telegram_username=normalize_tg_username(tg_user.get('username')),
            is_linked=True,
        ).first()
    actor = db.session.get(User, account.user_id) if account else None
    penalty = db.session.get(StudentPenalty, penalty_id)
    if not penalty:
        if callback_id:
            telegram_answer_callback(callback_id, 'Пенальти уже не найдено')
        return {'ok': False, 'reason': 'penalty_not_found'}
    event = db.session.get(NotificationEvent, event_id)
    current_pool_id = active_pool_id()
    if (
        not actor
        or not event
        or event.recipient_user_id != actor.id
        or penalty.pool_id != current_pool_id
        or event.pool_id != current_pool_id
    ):
        if callback_id:
            telegram_answer_callback(callback_id, 'Действие уже не относится к активному бассейну')
        return {'ok': False, 'reason': 'inactive_pool_context'}

    expected_event_type = {
        'm': 'penalty_method_question',
        'c': 'penalty_workoff_check',
        'u': 'penalty_admin_unlock',
    }.get(question)
    if (
        not expected_event_type
        or event.type != expected_event_type
        or event.source_entity != 'penalty'
        or event.source_entity_id != penalty.id
    ):
        if callback_id:
            telegram_answer_callback(callback_id, 'Действие больше недоступно')
        return {'ok': False, 'reason': 'invalid_event_context'}

    immediate_events = []

    if question == 'm':
        event_type = 'penalty_method_question'
        if answer == 'y' and penalty.workoff_status in ('pending', 'overdue'):
            _set_penalty_status_from_bot(penalty, 'in_workoff', actor, 'Telegram: метод отработки выдан')
            _delete_related_penalty_messages(penalty.id, event_type, event_id)
            if callback_id:
                telegram_answer_callback(callback_id, 'Статус: отрабатывает')
        elif answer == 'n' and penalty.workoff_status in ('pending', 'overdue'):
            _delete_related_penalty_messages(penalty.id, event_type, event_id)
            _queue_penalty_method_question(penalty, scheduled_for=_utcnow() + timedelta(minutes=5), suffix=f'retry-{int(time.time())}')
            if callback_id:
                telegram_answer_callback(callback_id, 'Спросим ещё раз через 5 минут')
        else:
            _delete_related_penalty_messages(penalty.id, event_type, event_id)
            if callback_id:
                telegram_answer_callback(callback_id, 'Вопрос пропущен')
    elif question == 'c':
        event_type = 'penalty_workoff_check'
        if answer == 'y' and penalty.workoff_status == 'in_workoff':
            _set_penalty_status_from_bot(penalty, 'awaiting_unlock', actor, 'Telegram: пенальти отработан')
            _delete_related_penalty_messages(penalty.id, event_type, event_id)
            immediate_events = _notify_admins_penalty_awaiting_unlock(penalty)
            if callback_id:
                telegram_answer_callback(callback_id, 'Статус: ожидает разблокировки')
        elif answer == 'n' and penalty.workoff_status == 'in_workoff':
            _delete_related_penalty_messages(penalty.id, event_type, event_id)
            _queue_penalty_workoff_check(penalty, scheduled_for=_utcnow() + timedelta(minutes=5), suffix=f'retry-{int(time.time())}')
            if callback_id:
                telegram_answer_callback(callback_id, 'Спросим ещё раз через 5 минут')
        else:
            _delete_related_penalty_messages(penalty.id, event_type, event_id)
            if callback_id:
                telegram_answer_callback(callback_id, 'Вопрос пропущен')
    elif question == 'u':
        is_responsible = PoolVolunteer.query.filter(
            PoolVolunteer.pool_id == penalty.pool_id,
            PoolVolunteer.user_id == actor.id,
            PoolVolunteer.pool_role.in_(['responsible_admin', 'responsible_team_lead']),
        ).first() is not None
        if not is_responsible:
            if callback_id:
                telegram_answer_callback(callback_id, 'Только ответственный может подтвердить разблокировку')
            return {'ok': False, 'reason': 'not_pool_responsible'}
        if answer == 'y' and penalty.workoff_status == 'awaiting_unlock':
            _set_penalty_status_from_bot(
                penalty,
                'unlocked',
                actor,
                'Telegram: ответственный подтвердил разблокировку',
            )
            _delete_related_penalty_messages(penalty.id, 'penalty_admin_unlock')
            if callback_id:
                telegram_answer_callback(callback_id, 'Статус на платформе: разблокирован')
        elif penalty.workoff_status == 'unlocked':
            if callback_id:
                telegram_answer_callback(callback_id, 'Ученик уже отмечен разблокированным')
        else:
            if callback_id:
                telegram_answer_callback(callback_id, 'Пенальти уже изменилось, обнови данные')
    db.session.commit()
    if immediate_events:
        queue_notification_dispatch(immediate_events)
    return {'ok': True}


def _queue_shift_change_notifications(block, target_user, action):
    now_msk = _moscow_now()
    if block.date != now_msk.date() + timedelta(days=1) or now_msk.hour < 14:
        return
    action_text = 'записан на' if action == 'create' else 'снят со'
    if action == 'create':
        coworkers = [u for u in _signed_users_for_block(block) if u.id != target_user.id]
        text = f'Ты дежуришь завтра: {_format_shift(block)}.'
        if coworkers:
            text += '\nС тобой на смене: ' + ', '.join(_tg_link(u) for u in coworkers)
        if block.label == 'EXAM':
            text += f'\nБриф экзамена: {_exam_brief_url()}'
        _queue_notification(
            target_user,
            'shift_change_volunteer',
            text,
            f'shift-change:{block.id}:{target_user.id}:{action}:{int(time.time())}',
            pool_id=block.pool_id,
            scheduled_for=_utcnow() + timedelta(minutes=5),
            source_entity='shift_block',
            source_entity_id=block.id,
        )

    for user in _pool_responsible_users(block.pool_id):
        _queue_notification(
            user,
            'shift_change_staff',
            (
                f'Изменение в завтрашней смене: {target_user.name or target_user.nick} '
                f'({_tg_link(target_user)}) {action_text} смены {_format_shift(block)}.'
            ),
            f'shift-change-staff:{block.id}:{target_user.id}:{action}:{user.id}:{int(time.time())}',
            pool_id=block.pool_id,
            scheduled_for=_utcnow() + timedelta(minutes=5),
            source_entity='shift_block',
            source_entity_id=block.id,
        )


def mark_delivery_failed(event, delivery, error_text):
    delivery.delivery_status = 'error'
    delivery.error = (error_text or '')[:500]
    event.status = 'queued'
    event.scheduled_for = _utcnow() + timedelta(minutes=5)


def _schedule_daily_shift_notifications():
    now_msk = _moscow_now()
    if now_msk.hour != 14:
        return
    pool_id = active_pool_id()
    if not pool_id:
        return
    target_date = now_msk.date() + timedelta(days=1)
    blocks = ShiftBlock.query.filter_by(pool_id=pool_id, date=target_date).order_by(ShiftBlock.time_start).all()
    if not blocks:
        return
    users_to_blocks = {}
    for block in blocks:
        for user in _signed_users_for_block(block):
            users_to_blocks.setdefault(user.id, {'user': user, 'blocks': []})['blocks'].append(block)
    for item in users_to_blocks.values():
        user = item['user']
        block_lines = []
        coworker_lines = []
        has_exam = False
        for block in item['blocks']:
            has_exam = has_exam or (block.label == 'EXAM')
            block_lines.append(f'• {_format_shift(block)}')
            coworkers = [u for u in _signed_users_for_block(block) if u.id != user.id]
            if coworkers:
                coworker_lines.append(f'{block.time_start}-{block.time_end}: ' + ', '.join(_tg_link(u) for u in coworkers))
        text = 'Завтра ты дежуришь на бассейне:\n' + '\n'.join(block_lines)
        if coworker_lines:
            text += '\n\nС тобой на смене:\n' + '\n'.join(coworker_lines)
        if has_exam:
            text += f'\n\nБриф экзамена: {_exam_brief_url()}'
        _queue_notification(
            user,
            'shift_reminder_volunteer',
            text,
            f'shift-reminder:{target_date.isoformat()}:user:{user.id}',
            pool_id=pool_id,
        )

    summary_lines = []
    for block in blocks:
        volunteers = _signed_users_for_block(block)
        people = ', '.join(f'{u.name or u.nick} ({_tg_link(u)})' for u in volunteers) or 'никто не записан'
        summary_lines.append(f'• {_format_shift(block)}: {people}')
    for user in _pool_responsible_users(pool_id):
        _queue_notification(
            user,
            'shift_reminder_staff',
            'Кто дежурит завтра:\n' + '\n'.join(summary_lines),
            f'shift-reminder-staff:{target_date.isoformat()}:user:{user.id}',
            pool_id=pool_id,
        )


def _schedule_tribe_notifications():
    now_msk = _moscow_now()
    pool_id = active_pool_id()
    if not pool_id:
        return
    pool_start_date = _pool_start_date(pool_id)
    tomorrow = now_msk.date() + timedelta(days=1)
    tomorrow_events = TribeEvent.query.filter_by(pool_id=pool_id, event_date=tomorrow).all()
    for event in tomorrow_events:
        masters = _tribe_masters_for_pool(pool_id, event.tribe)
        for user in masters:
            _queue_notification(
                user,
                'tribe_event_tomorrow',
                (
                    f'Завтра мероприятие твоего трайба {event.tribe}.\n'
                    f'{event.title} в {event.time_start or "время не указано"}'
                    + (f'\nМесто: {event.location}' if event.location else '')
                ),
                f'tribe-event:{event.id}:tomorrow:user:{user.id}',
                pool_id=pool_id,
                scheduled_for=_moscow_to_utc(datetime.combine(tomorrow - timedelta(days=1), datetime.strptime('14:00', '%H:%M').time())),
                source_entity='tribe_event',
                source_entity_id=event.id,
            )

    today = now_msk.date()
    if pool_start_date != today:
        return
    today_events = TribeEvent.query.filter_by(pool_id=pool_id, event_date=today).all()
    for event in today_events:
        if not event.time_start:
            continue
        event_dt_msk = datetime.combine(event.event_date, datetime.strptime(event.time_start, '%H:%M').time())
        scheduled_for = _moscow_to_utc(event_dt_msk - timedelta(minutes=10))
        if scheduled_for > _utcnow() + timedelta(minutes=10):
            continue
        if scheduled_for < _utcnow() - timedelta(minutes=15):
            continue
        for user in _users_on_shift(pool_id, event_dt_msk - timedelta(minutes=10)):
            _queue_notification(
                user,
                'tribe_event_first_day_shift',
                (
                    f'Через 10 минут мероприятие трайба {event.tribe}: {event.title}.\n'
                    'Пожалуйста, направь ребят на мероприятие.'
                    + (f'\nМесто: {event.location}' if event.location else '')
                ),
                f'tribe-event:{event.id}:shift-reminder:user:{user.id}',
                pool_id=pool_id,
                scheduled_for=scheduled_for,
                source_entity='tribe_event',
                source_entity_id=event.id,
            )


def _schedule_penalty_checks():
    now = _utcnow()
    pool_id = active_pool_id()
    if not pool_id:
        return
    penalties = StudentPenalty.query.filter(
        StudentPenalty.pool_id == pool_id,
        StudentPenalty.workoff_status == 'in_workoff',
        StudentPenalty.date_worked_off.isnot(None),
    ).all()
    for penalty in penalties:
        scheduled_for = penalty.date_worked_off + timedelta(hours=1, minutes=55)
        if scheduled_for <= now + timedelta(minutes=5):
            _queue_penalty_workoff_check(
                penalty,
                scheduled_for=scheduled_for,
                suffix=f'auto-{scheduled_for.strftime("%Y%m%d%H%M")}',
            )


def enqueue_scheduled_notifications():
    _schedule_daily_shift_notifications()
    _schedule_tribe_notifications()
    _schedule_penalty_checks()


def _claim_pending_notification_ids(limit, event_ids=None):
    now = _utcnow()
    stale_before = now - timedelta(minutes=10)
    NotificationEvent.query.filter(
        NotificationEvent.status == 'processing',
        NotificationEvent.processing_started_at < stale_before,
    ).update({
        NotificationEvent.status: 'queued',
        NotificationEvent.processing_started_at: None,
    }, synchronize_session=False)

    candidate_query = (
        NotificationEvent.query
        .with_entities(NotificationEvent.id)
        .filter(NotificationEvent.status.in_(['queued', 'pending']))
        .filter(db.or_(NotificationEvent.scheduled_for.is_(None), NotificationEvent.scheduled_for <= now))
    )
    if event_ids is not None:
        candidate_query = candidate_query.filter(NotificationEvent.id.in_(event_ids))
    candidate_query = (
        candidate_query
        .order_by(NotificationEvent.created_at.asc(), NotificationEvent.id.asc())
        .limit(limit)
    )
    if db.engine.dialect.name == 'postgresql':
        candidate_query = candidate_query.with_for_update(skip_locked=True)

    candidate_ids = [row.id for row in candidate_query.all()]
    if candidate_ids:
        NotificationEvent.query.filter(
            NotificationEvent.id.in_(candidate_ids),
            NotificationEvent.status.in_(['queued', 'pending']),
        ).update({
            NotificationEvent.status: 'processing',
            NotificationEvent.processing_started_at: now,
        }, synchronize_session=False)
    db.session.commit()
    return candidate_ids


def process_pending_notifications(limit=20, event_ids=None, schedule_events=True):
    if schedule_events:
        enqueue_scheduled_notifications()
    candidate_ids = _claim_pending_notification_ids(limit, event_ids=event_ids)
    sent = 0
    failed = 0
    skipped = 0
    processed = 0
    for event_id in candidate_ids:
        processed += 1
        event = db.session.get(NotificationEvent, event_id)
        if event.pool_id is not None and event.pool_id != active_pool_id():
            event.status = 'cancelled'
            event.cancelled_at = _utcnow()
            event.processing_started_at = None
            skipped += 1
            db.session.commit()
            continue
        user = db.session.get(User, event.recipient_user_id) if event.recipient_user_id else None
        account = TelegramAccount.query.filter_by(
            user_id=user.id,
            is_linked=True,
            delivery_enabled=True,
        ).first() if user and user.active else None
        delivery = NotificationDelivery.query.filter_by(notification_id=event.id).order_by(NotificationDelivery.id.desc()).first()

        if not delivery:
            delivery = NotificationDelivery(
                notification_id=event.id,
                user_id=user.id if user else None,
                telegram_chat_id=account.telegram_chat_id if account else None,
                delivery_status='pending',
            )
            db.session.add(delivery)

        if not user or not account or not account.telegram_chat_id:
            delivery.delivery_status = 'skipped'
            delivery.error = 'Доставка Telegram отключена или аккаунт не привязан'
            event.status = 'skipped'
            event.processing_started_at = None
            skipped += 1
            db.session.commit()
            continue

        try:
            result = telegram_send_message(
                account.telegram_chat_id,
                build_notification_text(event),
                disable_notification=telegram_is_quiet_hours(_moscow_now()) and event.priority != 'urgent',
                reply_markup=build_notification_reply_markup(event),
            )
            delivery.telegram_chat_id = account.telegram_chat_id
            delivery.delivery_status = 'sent'
            delivery.message_id = str(result.get('message_id'))
            delivery.error = None
            event.status = 'sent'
            event.sent_at = _utcnow()
            event.scheduled_for = None
            event.processing_started_at = None
            account.last_delivery_at = _utcnow()
            log_action(
                'send',
                'notification_event',
                event.id,
                'Telegram уведомление отправлено',
                {
                    'recipient_nick': user.nick,
                    'event_type': event.type,
                    'priority': event.priority,
                    'message_id': delivery.message_id,
                },
            )
            sent += 1
        except Exception as exc:
            mark_delivery_failed(event, delivery, str(exc))
            event.processing_started_at = None
            log_action(
                'delivery_error',
                'notification_event',
                event.id,
                'Ошибка отправки Telegram уведомления',
                {
                    'recipient_nick': user.nick,
                    'event_type': event.type,
                    'error': str(exc),
                },
            )
            failed += 1

        db.session.commit()

    return {
        'processed': processed,
        'sent': sent,
        'failed': failed,
        'skipped': skipped,
    }


def queue_notification_dispatch(events):
    """Ask Postgres to start delivery after commit without blocking the UI request."""
    event_ids = sorted({event.id for event in events if event and event.id})
    if (
        not event_ids
        or not _telegram_is_configured()
        or not INTERNAL_API_SECRET
        or not NOTIFICATION_DISPATCH_URL
        or db.engine.dialect.name != 'postgresql'
    ):
        return None
    try:
        request_id = db.session.execute(
            db.text('''
                SELECT net.http_post(
                    url := :url,
                    body := CAST(:body AS jsonb),
                    headers := CAST(:headers AS jsonb),
                    timeout_milliseconds := :timeout_ms
                )
            '''),
            {
                'url': NOTIFICATION_DISPATCH_URL,
                'body': json.dumps({'event_ids': event_ids}),
                'headers': json.dumps({
                    'Authorization': f'Bearer {INTERNAL_API_SECRET}',
                    'Content-Type': 'application/json',
                }),
                'timeout_ms': NOTIFICATION_DISPATCH_TIMEOUT_MS,
            },
        ).scalar()
        db.session.commit()
        return request_id
    except Exception as exc:
        db.session.rollback()
        app.logger.warning('Could not queue immediate Telegram dispatch: %s', exc)
        return None


def verify_internal_api_secret():
    if not INTERNAL_API_SECRET:
        return False
    auth = request.headers.get('Authorization', '')
    bearer = auth[7:] if auth.startswith('Bearer ') else ''
    alt = request.headers.get('X-Internal-Secret', '')
    query_secret = request.args.get('secret', '')
    return INTERNAL_API_SECRET in {bearer, alt, query_secret}


# ==================== Авторизация ====================


def make_token(user):
    return serializer.dumps({'id': user.id, 'role': user.role})


def _pool_membership_for_user(user, pool_id=None):
    if not user:
        return None
    target_pool_id = pool_id or active_pool_id()
    if not target_pool_id:
        return None
    return PoolVolunteer.query.filter_by(pool_id=target_pool_id, user_id=user.id).first()


def _effective_access_role(user, pool_id=None):
    if user.role in {'team_lead', 'admin'}:
        return user.role
    membership = _pool_membership_for_user(user, pool_id)
    if membership and membership.pool_role in {'volunteer', 'tribe_master'}:
        return membership.pool_role
    return 'volunteer'


def _effective_access_tribe(user, pool_id=None):
    if user.role in {'team_lead', 'admin'}:
        return user.tribe
    membership = _pool_membership_for_user(user, pool_id)
    if membership and membership.pool_role == 'tribe_master' and membership.tribe:
        return membership.tribe
    return None


def _session_user_dict(user, pool_id=None):
    pool_id = pool_id or active_pool_id()
    payload = user.to_dict()
    payload['role'] = _effective_access_role(user, pool_id)
    payload['tribe'] = _effective_access_tribe(user, pool_id)
    payload['active_pool_id'] = pool_id
    return payload


def _new_pool_invite_token():
    return secrets.token_urlsafe(24)


def _invite_is_expired(invite):
    return bool(invite and invite.expires_at and invite.expires_at <= _naive_utcnow())


def _invite_limit_reached(invite):
    return bool(invite and invite.max_uses is not None and invite.max_uses > 0 and (invite.uses_count or 0) >= invite.max_uses)


def _invite_availability_error(invite, pool):
    if not invite or not invite.is_active:
        return 'Ссылка недействительна', 404
    if not pool or pool.archived or not pool.active:
        return 'Этот бассейн уже недоступен', 404
    if _invite_is_expired(invite):
        return 'Срок действия ссылки истёк', 410
    if _invite_limit_reached(invite):
        return 'Лимит входов по этой ссылке уже исчерпан', 410
    return None, None


def _pool_invite_payload(invite):
    return {
        **invite.to_dict(),
        'invite_url': f'/join/{invite.token}',
        'is_expired': _invite_is_expired(invite),
        'is_limit_reached': _invite_limit_reached(invite),
        'is_available': bool(invite.is_active) and not _invite_is_expired(invite) and not _invite_limit_reached(invite),
        'remaining_uses': None if invite.max_uses is None else max(invite.max_uses - (invite.uses_count or 0), 0),
    }


def _bind_request_user(user, pool_id=None):
    g.user = user
    g.current_role = _effective_access_role(user, pool_id)
    g.current_tribe = _effective_access_tribe(user, pool_id)
    return user


def load_user_from_request():
    auth = request.headers.get('Authorization', '')
    token = request.args.get('token')
    if not token:
        if not auth.startswith('Bearer '):
            return None
        token = auth[7:]
    try:
        data = serializer.loads(token, max_age=60 * 60 * 24 * 30)  # 30 дней
    except BadSignature:
        return None
    user = db.session.get(User, data.get('id'))
    return user if user and user.active else None


def require_auth(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = load_user_from_request()
        if not user:
            return jsonify({'error': 'Не авторизован'}), 401
        _bind_request_user(user)
        return fn(*args, **kwargs)
    return wrapper


def require_role(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = load_user_from_request()
            if not user:
                return jsonify({'error': 'Не авторизован'}), 401
            effective_role = _effective_access_role(user)
            if effective_role not in roles:
                return jsonify({'error': 'Недостаточно прав'}), 403
            _bind_request_user(user)
            return fn(*args, **kwargs)
        return wrapper
    return deco


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'timestamp': _utcnow().isoformat()})


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json(silent=True) or {}
    nick = (data.get('nick') or '').strip()
    password = data.get('password') or ''
    if not nick:
        return jsonify({'error': 'Укажите ник'}), 400

    user = User.query.filter(db.func.lower(User.nick) == nick.lower()).first()
    if not user or not user.active:
        return jsonify({'error': 'Тебя нет в системе. Обратись к тимлиду, чтобы внесли твой ник.'}), 403

    if user.role in ROLES_WITH_PASSWORD:
        if not user.password_hash or not check_password_hash(user.password_hash, password):
            return jsonify({'error': 'Неверный пароль'}), 403

    return jsonify({'token': make_token(user), 'user': _session_user_dict(user)})


@app.route('/api/auth/me', methods=['GET'])
@require_auth
def me():
    return jsonify(_session_user_dict(g.user))


@app.route('/api/me', methods=['PATCH'])
@require_auth
def update_me():
    user = g.user
    data = request.get_json(silent=True) or {}
    changes = {}

    if 'name' in data:
        new_name = (data.get('name') or '').strip() or user.nick
        old_name = user.name or user.nick
        if old_name != new_name:
            changes['name'] = {'from': old_name, 'to': new_name}
        user.name = new_name

    if 'nick' in data:
        new_nick = (data.get('nick') or '').strip()
        if not new_nick:
            return jsonify({'error': 'Укажите ник'}), 400
        existing = User.query.filter(db.func.lower(User.nick) == new_nick.lower(), User.id != user.id).first()
        if existing:
            return jsonify({'error': 'Такой ник уже есть'}), 409
        if user.nick != new_nick:
            changes['nick'] = {'from': user.nick, 'to': new_nick}
        user.nick = new_nick

    if 'telegram' in data:
        raw_telegram = (data.get('telegram') or '').strip()
        new_telegram = raw_telegram if not raw_telegram or raw_telegram.startswith('@') else f'@{raw_telegram}'
        if telegram_username_conflict(new_telegram, user.id):
            return jsonify({'error': 'Этот Telegram username уже используется'}), 409
        old_telegram = user.telegram or ''
        if old_telegram != (new_telegram or ''):
            changes['telegram'] = {'from': old_telegram or None, 'to': new_telegram or None}
        user.telegram = new_telegram or None
        account = _telegram_account_any(user.id)
        if account and not account.is_linked:
            account.telegram_username = normalize_tg_username(user.telegram or user.nick)

    log_action('update', 'profile', user.id, 'Пользователь обновил личные данные', {'changes': changes}, actor=user)
    db.session.commit()
    return jsonify(_session_user_dict(user))


@app.route('/api/users/<int:user_id>/avatar', methods=['GET'])
@require_auth
def user_avatar(user_id):
    user = get_model_or_404(User, user_id)
    account = _telegram_account_any(user.id)
    if not account or not (account.photo_file_id or account.photo_url):
        return jsonify({'error': 'Фото не найдено'}), 404
    try:
        photo_bytes, content_type = _download_telegram_photo_bytes(account)
        if not photo_bytes:
            return jsonify({'error': 'Фото не найдено'}), 404
        db.session.commit()
        return send_file(
            BytesIO(photo_bytes),
            mimetype=content_type or 'image/jpeg',
            max_age=60 * 60,
            download_name=f'user-{user.id}-avatar.jpg',
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': f'Не удалось получить фото: {exc}'}), 502


@app.route('/api/me/avatar', methods=['POST'])
@require_auth
def upload_my_avatar():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Загрузите изображение'}), 400
    mime = (file.mimetype or '').lower()
    if mime not in {'image/jpeg', 'image/png', 'image/webp', 'image/jpg'}:
        return jsonify({'error': 'Поддерживаются только JPG, PNG и WEBP'}), 400
    content = file.read()
    if not content:
        return jsonify({'error': 'Файл пустой'}), 400
    if len(content) > 3 * 1024 * 1024:
        return jsonify({'error': 'Файл слишком большой. Максимум 3 МБ'}), 400
    import base64
    user = g.user
    account = _telegram_account_any(user.id)
    if not account:
        account = TelegramAccount(
            user_id=user.id,
            telegram_username=normalize_tg_username(user.telegram or user.nick),
            is_linked=False,
            delivery_enabled=False,
        )
        db.session.add(account)
    account.photo_file_id = None
    account.photo_url = f'data:{mime};base64,{base64.b64encode(content).decode("ascii")}'
    account.last_photo_sync_at = _utcnow()
    db.session.commit()
    log_action('upload', 'profile_avatar', user.id, 'Пользователь загрузил фото профиля', actor=user)
    return jsonify({'ok': True, 'avatar_url': _avatar_url_for_user(user), 'user': _session_user_dict(user)})


# ==================== Пользователи (волонтёры) ====================


@app.route('/api/users', methods=['GET'])
@require_role('team_lead', 'admin')
def list_users():
    role = request.args.get('role')
    q = User.query
    if role:
        q = q.filter_by(role=role)
    users = q.order_by(User.nick).all()
    return jsonify([u.to_dict() for u in users])


@app.route('/api/users', methods=['POST'])
@require_role('team_lead', 'admin')
def create_user():
    data = request.get_json(silent=True) or {}
    nick = (data.get('nick') or '').strip()
    role = data.get('role', 'volunteer')
    if not nick:
        return jsonify({'error': 'Укажите ник'}), 400
    if role not in ALL_ROLES:
        return jsonify({'error': 'Неизвестная роль'}), 400
    # тимлид может заводить только волонтёров и трайб-мастеров
    if g.user.role == 'team_lead' and role in ROLES_WITH_PASSWORD:
        return jsonify({'error': 'Только админ может создавать тимлидов и админов'}), 403

    existing = User.query.filter(db.func.lower(User.nick) == nick.lower()).first()
    if existing:
        if role in ROLES_WITH_PASSWORD and existing.role not in ROLES_WITH_PASSWORD:
            password = data.get('password') or ''
            if len(password) < 4:
                return jsonify({'error': 'Для этой роли нужен пароль (мин. 4 символа)'}), 400
            existing.name = data.get('name') or existing.name or nick
            existing.role = role
            existing.password_hash = generate_password_hash(password)
            db.session.commit()
            log_action(
                'update',
                'user',
                existing.id,
                f'Повышен @{existing.nick} до роли {role}',
                {'target_nick': existing.nick, 'role': role},
            )
            return jsonify(existing.to_dict()), 200
        return jsonify({'error': 'Такой ник уже есть'}), 409

    telegram = _normalize_telegram_value(data.get('telegram'))
    if telegram_username_conflict(telegram):
        return jsonify({'error': 'Этот Telegram username уже используется'}), 409
    user = User(
        nick=nick,
        name=data.get('name') or nick,
        role=role,
        telegram=telegram,
        tribe=normalize_tribe(data.get('tribe')),
    )
    if role in ROLES_WITH_PASSWORD:
        password = data.get('password') or ''
        if len(password) < 4:
            return jsonify({'error': 'Для этой роли нужен пароль (мин. 4 символа)'}), 400
        user.password_hash = generate_password_hash(password)
    db.session.add(user)
    db.session.commit()
    return jsonify(user.to_dict()), 201


@app.route('/api/users/<int:user_id>', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_user(user_id):
    user = get_model_or_404(User, user_id)
    if user.role in ROLES_WITH_PASSWORD and g.user.role != 'admin':
        return jsonify({'error': 'Только админ может редактировать тимлидов и админов'}), 403
    data = request.json or {}
    changes = {}

    if 'name' in data:
        new_name = (data.get('name') or '').strip() or user.nick
        old_name = user.name or user.nick
        if old_name != new_name:
            changes['name'] = {'from': old_name, 'to': new_name}
        user.name = new_name

    if 'nick' in data:
        new_nick = (data.get('nick') or '').strip()
        if not new_nick:
            return jsonify({'error': 'Укажите ник'}), 400
        existing = User.query.filter(db.func.lower(User.nick) == new_nick.lower(), User.id != user.id).first()
        if existing:
            return jsonify({'error': 'Такой ник уже есть'}), 409
        if user.nick != new_nick:
            changes['nick'] = {'from': user.nick, 'to': new_nick}
        user.nick = new_nick

    if 'telegram' in data:
        new_telegram = _normalize_telegram_value(data.get('telegram'))
        if telegram_username_conflict(new_telegram, user.id):
            return jsonify({'error': 'Этот Telegram username уже используется'}), 409
        old_telegram = user.telegram or ''
        if old_telegram != (new_telegram or ''):
            changes['telegram'] = {'from': old_telegram or None, 'to': new_telegram or None}
        user.telegram = new_telegram or None
        account = _telegram_account_any(user.id)
        if account and not account.is_linked:
            account.telegram_username = normalize_tg_username(user.telegram or user.nick)

    if 'password' in data and user.role in ROLES_WITH_PASSWORD:
        password = data.get('password') or ''
        if password:
            if len(password) < 4:
                return jsonify({'error': 'Пароль должен быть не короче 4 символов'}), 400
            user.password_hash = generate_password_hash(password)
            changes['password'] = {'from': 'set', 'to': 'updated'}

    log_action('update', 'user', user.id, 'Обновлён пользователь', {'changes': changes}, actor=g.user)
    db.session.commit()
    return jsonify(user.to_dict())


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_user(user_id):
    user = get_model_or_404(User, user_id)
    if (user.nick or '').strip().lower() == 'admin':
        return jsonify({'error': 'Пользователя @admin удалить нельзя'}), 400
    if user.role in ROLES_WITH_PASSWORD and g.user.role != 'admin':
        return jsonify({'error': 'Только админ может удалять тимлидов и админов'}), 403
    if user.id == g.user.id:
        return jsonify({'error': 'Нельзя удалить самого себя'}), 400

    NotificationDelivery.query.filter_by(user_id=user.id).delete()
    NotificationEvent.query.filter_by(recipient_user_id=user.id).delete()
    NotificationEvent.query.filter_by(created_by=user.id).update({'created_by': None})
    TelegramAccount.query.filter_by(user_id=user.id).delete()
    PoolVolunteer.query.filter_by(user_id=user.id).delete()
    RewardEvent.query.filter(
        db.or_(RewardEvent.user_id == user.id, RewardEvent.created_by == user.id)
    ).delete(synchronize_session=False)
    GroupReview.query.filter(
        db.or_(GroupReview.reviewer_id == user.id, GroupReview.created_by == user.id)
    ).delete(synchronize_session=False)
    TribeEvent.query.filter_by(created_by=user.id).update({'created_by': None})
    StudentEvent.query.filter_by(created_by=user.id).update({'created_by': None})
    StudentPenalty.query.filter_by(volunteer_id=user.id).update({
        'volunteer_id': None,
        'volunteer_name': None,
    }, synchronize_session=False)
    Broadcast.query.filter_by(author_id=user.id).delete()
    DashboardNote.query.filter_by(author_id=user.id).delete()
    ActionLog.query.filter_by(actor_id=user.id).update({'actor_id': None})
    PenaltyHistory.query.filter_by(actor_id=user.id).update({'actor_id': None})
    ScheduleGeneration.query.filter_by(created_by=user.id).update({'created_by': None})
    Signup.query.filter_by(user_id=user.id).delete()
    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'Удалён'})


@app.route('/api/me/password', methods=['POST'])
@require_role('team_lead', 'admin')
def change_my_password():
    data = request.json or {}
    current_password = data.get('current_password') or ''
    new_password = data.get('new_password') or ''

    if not g.user.password_hash or not check_password_hash(g.user.password_hash, current_password):
        return jsonify({'error': 'Текущий пароль неверный'}), 400
    if len(new_password) < 4:
        return jsonify({'error': 'Новый пароль должен быть не короче 4 символов'}), 400

    g.user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    log_action('update', 'user', g.user.id, 'Сменён пароль администратора/тимлида', {'user_id': g.user.id})
    return jsonify({'message': 'Пароль обновлён'})


# ==================== Бассейны ====================


def can_access_pool(user, pool):
    """team_lead/admin видят всё. Волонтёр видит бассейн если назначен И не архивирован."""
    if user.role in ('team_lead', 'admin'):
        return True
    if pool.archived:
        return False
    return PoolVolunteer.query.filter_by(pool_id=pool.id, user_id=user.id).first() is not None


def _can_access_pool_id(user, pool_id):
    if not pool_id:
        return False
    pool = db.session.get(Pool, pool_id)
    return bool(pool and can_access_pool(user, pool))


@app.route('/api/pools', methods=['GET'])
@require_auth
def list_pools():
    pools = Pool.query.order_by(Pool.created_at.desc()).all()
    if g.user.role not in ('team_lead', 'admin'):
        pools = [p for p in pools if can_access_pool(g.user, p)]
    return jsonify([p.to_dict() for p in pools])


@app.route('/api/pools/active', methods=['GET'])
@require_auth
def active_pool():
    pool = Pool.query.filter_by(active=True).order_by(Pool.created_at.desc()).first()
    return jsonify(pool.to_dict() if pool else None)


@app.route('/api/pools', methods=['POST'])
@require_role('team_lead', 'admin')
def create_pool():
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Укажите название бассейна'}), 400
    start_date = None
    if data.get('start_date'):
        start_date = datetime.fromisoformat(data['start_date']).date()
    pool = Pool(name=name, start_date=start_date, active=False)
    db.session.add(pool)
    db.session.commit()
    return jsonify(pool.to_dict()), 201


@app.route('/api/pools/<int:pool_id>', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_pool(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Укажите название бассейна'}), 400
    pool.name = name
    if 'start_date' in data:
        pool.start_date = datetime.fromisoformat(data['start_date']).date() if data.get('start_date') else None
    db.session.commit()
    return jsonify(pool.to_dict())


@app.route('/api/pools/<int:pool_id>/activate', methods=['POST'])
@require_role('team_lead', 'admin')
def activate_pool(pool_id):
    Pool.query.with_for_update().all()
    pool = get_model_or_404(Pool, pool_id)
    Pool.query.update({Pool.active: False})
    db.session.flush()
    pool.active = True
    pool.archived = False
    NotificationEvent.query.filter(
        NotificationEvent.pool_id.isnot(None),
        NotificationEvent.pool_id != pool.id,
        NotificationEvent.status.in_(['queued', 'pending', 'processing']),
    ).update({
        NotificationEvent.status: 'cancelled',
        NotificationEvent.cancelled_at: _utcnow(),
    }, synchronize_session=False)
    db.session.commit()
    return jsonify(pool.to_dict())


@app.route('/api/pools/<int:pool_id>/archive', methods=['POST'])
@require_role('team_lead', 'admin')
def archive_pool(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    if pool.active:
        return jsonify({'error': 'Сначала активируйте другой бассейн'}), 409
    pool.archived = True
    db.session.commit()
    return jsonify(pool.to_dict())


@app.route('/api/pools/<int:pool_id>/unarchive', methods=['POST'])
@require_role('team_lead', 'admin')
def unarchive_pool(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    pool.archived = False
    db.session.commit()
    return jsonify(pool.to_dict())


@app.route('/api/pools/<int:pool_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_pool(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    if pool.active:
        return jsonify({'error': 'Сначала переведите активный бассейн в другой статус'}), 400
    try:
        block_ids = [row.id for row in ShiftBlock.query.with_entities(ShiftBlock.id).filter_by(pool_id=pool_id).all()]
        student_ids = [row.id for row in Student.query.with_entities(Student.id).filter_by(pool_id=pool_id).all()]
        penalty_ids = [row.id for row in StudentPenalty.query.with_entities(StudentPenalty.id).filter_by(pool_id=pool_id).all()]
        notification_ids = [row.id for row in NotificationEvent.query.with_entities(NotificationEvent.id).filter_by(pool_id=pool_id).all()]

        if block_ids:
            Signup.query.filter(Signup.block_id.in_(block_ids)).delete(synchronize_session=False)
        if student_ids:
            StudentEvent.query.filter(StudentEvent.student_id.in_(student_ids)).delete(synchronize_session=False)
        if penalty_ids:
            PenaltyHistory.query.filter(PenaltyHistory.penalty_id.in_(penalty_ids)).delete(synchronize_session=False)
        if notification_ids:
            NotificationDelivery.query.filter(NotificationDelivery.notification_id.in_(notification_ids)).delete(synchronize_session=False)

        ShiftBlock.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        ScheduleGeneration.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        PoolInviteLink.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        PoolVolunteer.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        Tribe.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        Broadcast.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        DashboardNote.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        NotificationEvent.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        RewardEvent.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        GroupReview.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        TribeEvent.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        StudentPenalty.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)
        Student.query.filter_by(pool_id=pool_id).delete(synchronize_session=False)

        db.session.delete(pool)
        db.session.commit()
        return jsonify({'message': 'Бассейн удалён'})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': f'Не удалось удалить бассейн: {exc}'}), 500


@app.route('/api/pools/<int:pool_id>/volunteers', methods=['GET'])
@require_role('team_lead', 'admin')
def list_pool_volunteers(pool_id):
    get_model_or_404(Pool, pool_id)
    pvs = PoolVolunteer.query.filter_by(pool_id=pool_id).order_by(PoolVolunteer.assigned_at).all()
    result = []
    for pv in pvs:
        user = db.session.get(User, pv.user_id)
        if user:
            if pv.pool_role in ('responsible_admin', 'responsible_team_lead'):
                continue
            d = user.to_dict()
            d['role'] = pv.pool_role or 'volunteer'
            d['tribe'] = pv.tribe
            d['pool_tribe'] = pv.tribe
            d['assigned_at'] = pv.assigned_at.isoformat() if pv.assigned_at else None
            result.append(d)
    return jsonify(result)


@app.route('/api/pools/<int:pool_id>/volunteers', methods=['POST'])
@require_role('team_lead', 'admin')
def add_pool_volunteers(pool_id):
    get_model_or_404(Pool, pool_id)
    data = request.json or {}
    user_ids = data.get('user_ids') or ([data['user_id']] if data.get('user_id') else [])
    if not user_ids:
        return jsonify({'error': 'Укажите user_id или user_ids'}), 400
    added = 0
    for uid in user_ids:
        if not PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=uid).first():
            db.session.add(PoolVolunteer(pool_id=pool_id, user_id=int(uid)))
            added += 1
    db.session.commit()
    return jsonify({'added': added, 'message': f'Добавлено {added} волонтёров'})


@app.route('/api/pools/<int:pool_id>/volunteers/<int:user_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def remove_pool_volunteer(pool_id, user_id):
    pv = PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=user_id).first_or_404()
    block_ids = [b.id for b in ShiftBlock.query.filter_by(pool_id=pool_id).all()]
    if block_ids:
        Signup.query.filter(
            Signup.block_id.in_(block_ids), Signup.user_id == user_id
        ).delete(synchronize_session=False)
    db.session.delete(pv)
    db.session.commit()
    return jsonify({'message': 'Волонтёр удалён из бассейна, его смены очищены'})


@app.route('/api/pools/<int:pool_id>/responsibles', methods=['GET'])
@require_role('team_lead', 'admin')
def list_pool_responsibles(pool_id):
    get_model_or_404(Pool, pool_id)
    return jsonify(_pool_responsibles(pool_id))


@app.route('/api/pools/<int:pool_id>/responsibles', methods=['POST'])
@require_role('team_lead', 'admin')
def add_pool_responsible(pool_id):
    get_model_or_404(Pool, pool_id)
    data = request.json or {}
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'error': 'Укажите user_id'}), 400
    user = get_model_or_404(User, int(user_id))
    if user.role not in ('admin', 'team_lead'):
        return jsonify({'error': 'Ответственным может быть только админ или тимлид'}), 400
    pool_role = 'responsible_admin' if user.role == 'admin' else 'responsible_team_lead'
    pv = PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=user.id).first()
    if pv:
        pv.pool_role = pool_role
    else:
        db.session.add(PoolVolunteer(pool_id=pool_id, user_id=user.id, pool_role=pool_role))
    db.session.commit()
    return jsonify({'message': 'Ответственный добавлен', 'responsibles': _pool_responsibles(pool_id)})


@app.route('/api/pools/<int:pool_id>/responsibles/<int:user_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def remove_pool_responsible(pool_id, user_id):
    pv = PoolVolunteer.query.filter(
        PoolVolunteer.pool_id == pool_id,
        PoolVolunteer.user_id == user_id,
        PoolVolunteer.pool_role.in_(['responsible_admin', 'responsible_team_lead']),
    ).first_or_404()
    db.session.delete(pv)
    db.session.commit()
    return jsonify({'message': 'Ответственный удалён'})


@app.route('/api/pools/<int:pool_id>/responsibles/<int:user_id>/notifications', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_pool_responsible_notifications(pool_id, user_id):
    get_model_or_404(Pool, pool_id)
    error = _active_entity_error(pool_id)
    if error:
        return error
    data = request.json or {}
    enabled = data.get('enabled')
    if not isinstance(enabled, bool):
        return jsonify({'error': 'Поле enabled должно быть boolean'}), 400

    relation = PoolVolunteer.query.filter(
        PoolVolunteer.pool_id == pool_id,
        PoolVolunteer.user_id == user_id,
        PoolVolunteer.pool_role.in_(['responsible_admin', 'responsible_team_lead']),
    ).first_or_404()
    relation.notifications_enabled = enabled

    if not enabled:
        pending_events = NotificationEvent.query.filter(
            NotificationEvent.pool_id == pool_id,
            NotificationEvent.recipient_user_id == user_id,
            NotificationEvent.status.in_(['queued', 'pending', 'processing']),
        ).all()
        for event in pending_events:
            event.status = 'cancelled'
            event.cancelled_at = _utcnow()
            event.processing_started_at = None

    log_action(
        'update',
        'pool_responsible',
        user_id,
        'Изменена подписка ответственного на уведомления',
        {'pool_id': pool_id, 'notifications_enabled': enabled},
    )
    db.session.commit()
    return jsonify({
        'notifications_enabled': enabled,
        'message': 'Уведомления включены' if enabled else 'Уведомления выключены',
    })


@app.route('/api/pools/<int:pool_id>/invite-link', methods=['GET'])
@require_role('team_lead', 'admin')
def get_pool_invite_link(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    invite = PoolInviteLink.query.filter_by(pool_id=pool.id).first()
    payload = {
        'pool': pool.to_dict(),
        'invite': None,
    }
    if invite:
        payload['invite'] = _pool_invite_payload(invite)
    return jsonify(payload)


@app.route('/api/pools/<int:pool_id>/invite-link', methods=['POST'])
@require_role('team_lead', 'admin')
def create_pool_invite_link(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    if pool.archived or not pool.active:
        return jsonify({'error': 'Инвайт-ссылка доступна только для активного бассейна'}), 400
    data = request.get_json(silent=True) or {}

    raw_max_uses = data.get('max_uses')
    if raw_max_uses in ('', None):
        max_uses = None
    else:
        try:
            max_uses = int(raw_max_uses)
        except (TypeError, ValueError):
            return jsonify({'error': 'Лимит входов должен быть числом'}), 400
        if max_uses < 1:
            return jsonify({'error': 'Лимит входов должен быть не меньше 1'}), 400

    raw_expires_at = (data.get('expires_at') or '').strip()
    expires_at = None
    if raw_expires_at:
        try:
            expires_at = datetime.fromisoformat(raw_expires_at)
        except ValueError:
            return jsonify({'error': 'Некорректный срок действия'}), 400
        if expires_at <= _naive_utcnow():
            return jsonify({'error': 'Срок действия должен быть позже текущего времени'}), 400

    invite = PoolInviteLink.query.filter_by(pool_id=pool.id).first()
    if invite:
        invite.token = _new_pool_invite_token()
        invite.is_active = True
        invite.created_by = g.user.id
        invite.uses_count = 0
    else:
        invite = PoolInviteLink(
            pool_id=pool.id,
            token=_new_pool_invite_token(),
            created_by=g.user.id,
            is_active=True,
        )
        db.session.add(invite)
    invite.max_uses = max_uses
    invite.expires_at = expires_at

    db.session.commit()
    log_action(
        'create',
        'pool_invite_link',
        invite.id,
        'Создана или обновлена инвайт-ссылка бассейна',
        {'pool_id': pool.id, 'token': invite.token, 'max_uses': invite.max_uses, 'expires_at': invite.expires_at.isoformat() if invite.expires_at else None},
        actor=g.user,
    )
    return jsonify({
        'message': 'Инвайт-ссылка обновлена',
        'invite': _pool_invite_payload(invite),
    })


@app.route('/api/pools/<int:pool_id>/invite-link', methods=['DELETE'])
@require_role('team_lead', 'admin')
def disable_pool_invite_link(pool_id):
    get_model_or_404(Pool, pool_id)
    invite = PoolInviteLink.query.filter_by(pool_id=pool_id).first_or_404()
    invite.is_active = False
    db.session.commit()
    log_action(
        'delete',
        'pool_invite_link',
        invite.id,
        'Отключена инвайт-ссылка бассейна',
        {'pool_id': pool_id},
        actor=g.user,
    )
    return jsonify({'message': 'Инвайт-ссылка отключена'})


@app.route('/api/invites/<string:token>', methods=['GET'])
def get_invite_info(token):
    invite = PoolInviteLink.query.filter_by(token=token).first()
    pool = db.session.get(Pool, invite.pool_id) if invite else None
    error, status = _invite_availability_error(invite, pool)
    if error:
        return jsonify({'error': error}), status

    return jsonify({
        'pool': {
            'id': pool.id,
            'name': pool.name,
            'start_date': pool.start_date.isoformat() if pool.start_date else None,
        },
        'invite': _pool_invite_payload(invite),
    })


@app.route('/api/invites/<string:token>/accept', methods=['POST'])
@require_auth
def accept_pool_invite(token):
    invite = PoolInviteLink.query.filter_by(token=token).first()
    pool = db.session.get(Pool, invite.pool_id) if invite else None
    error, status = _invite_availability_error(invite, pool)
    if error:
        return jsonify({'error': error}), status

    existing = PoolVolunteer.query.filter_by(pool_id=pool.id, user_id=g.user.id).first()
    if existing:
        return jsonify({
            'message': 'Ты уже привязан к этому бассейну',
            'already_joined': True,
            'pool': pool.to_dict(),
            'user': _session_user_dict(g.user, pool.id),
        })

    db.session.add(PoolVolunteer(pool_id=pool.id, user_id=g.user.id, pool_role='volunteer'))
    invite.uses_count = (invite.uses_count or 0) + 1
    db.session.commit()
    log_action(
        'create',
        'pool_invite_join',
        invite.id,
        'Пользователь присоединился к бассейну по инвайт-ссылке',
        {'pool_id': pool.id, 'user_id': g.user.id},
        actor=g.user,
    )
    return jsonify({
        'message': 'Доступ к бассейну открыт',
        'already_joined': False,
        'pool': pool.to_dict(),
        'user': _session_user_dict(g.user, pool.id),
    })


@app.route('/api/pools/<int:pool_id>/volunteers/template', methods=['GET'])
@require_role('team_lead', 'admin')
def pool_volunteers_template(pool_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Волонтёры'
    headers = ['Имя', 'Ник школьный', 'Ник Telegram', 'Роль']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='000000')
    header_font = Font(bold=True, color='00FF41')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    ws.append(['Иван Петров', 'ivanpetrov', '@ivanpetrov', 'volunteer'])
    ws.append(['Анна Сидорова', 'sidoroanna', '@sidoanna', 'tribe_master'])
    for col, width in [('A', 20), ('B', 16), ('C', 18), ('D', 14)]:
        ws.column_dimensions[col].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='volunteers_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/volunteers/template', methods=['GET'])
@require_role('team_lead', 'admin')
def global_volunteers_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Волонтёры'
    headers = ['Имя', 'Ник школьный', 'Ник Telegram']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='000000')
    header_font = Font(bold=True, color='00FF41')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    ws.append(['Иван Петров', 'ivanpetrov', '@ivanpetrov'])
    ws.append(['Анна Сидорова', 'sidoroanna', '@sidoanna'])
    for col, width in [('A', 20), ('B', 16), ('C', 18)]:
        ws.column_dimensions[col].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='volunteers_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def save_pool_volunteer_rows(rows, pool_id):
    """Создать/обновить волонтёров и назначить их на бассейн из Excel-файла."""
    created = 0
    updated = 0
    assigned = 0
    skipped = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            skipped.append({'row': index, 'reason': 'Некорректная строка'})
            continue
        nick = (row.get('nick') or '').strip()
        name = (row.get('name') or '').strip() or nick
        telegram = _normalize_telegram_value(row.get('telegram'))
        if not nick:
            skipped.append({'row': index, 'reason': 'Нет ника'})
            continue
        try:
            role = _volunteer_role_from_payload(row)
        except ValueError as e:
            skipped.append({'row': index, 'nick': nick, 'reason': str(e)})
            continue
        tribe = normalize_tribe(row.get('tribe'))
        user = User.query.filter(db.func.lower(User.nick) == nick.lower()).first()
        if telegram and telegram_username_conflict(telegram, user.id if user else None):
            skipped.append({'row': index, 'nick': nick, 'reason': 'Telegram username уже используется'})
            continue
        conflict = validate_tribe_master_assignment(
            pool_id,
            role,
            tribe,
            exclude_user_id=user.id if user else None,
        )
        if conflict:
            skipped.append({'row': index, 'nick': nick, 'reason': conflict})
            continue
        if user:
            if user.role in ROLES_WITH_PASSWORD:
                skipped.append({'row': index, 'nick': nick, 'reason': 'Ник занят тимлидом/админом'})
                continue
            if name:
                user.name = name
            if telegram:
                user.telegram = telegram
            updated += 1
        else:
            user = User(nick=nick, name=name, role='volunteer', telegram=telegram or None)
            db.session.add(user)
            db.session.flush()
            created += 1
        membership = PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=user.id).first()
        if not membership:
            db.session.add(PoolVolunteer(
                pool_id=pool_id,
                user_id=user.id,
                pool_role=role,
                tribe=tribe if role == 'tribe_master' else None,
            ))
            assigned += 1
        else:
            membership.pool_role = role
            membership.tribe = tribe if role == 'tribe_master' else None
    db.session.commit()
    return {
        'created': created, 'updated': updated, 'assigned': assigned, 'skipped': skipped,
        'message': f'Новых {created}, обновлено {updated}, добавлено на бассейн {assigned}',
    }


@app.route('/api/pools/<int:pool_id>/volunteers/import-file', methods=['POST'])
@require_role('team_lead', 'admin')
def import_pool_volunteers_file(pool_id):
    get_model_or_404(Pool, pool_id)
    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({'error': 'Загрузите файл'}), 400
    try:
        rows = parse_xlsx_rows(uploaded)
        volunteers = rows_to_dicts(rows, ['name', 'nick', 'telegram', 'role'])
    except Exception as e:
        return jsonify({'error': f'Не удалось прочитать .xlsx: {e}'}), 400
    return jsonify(save_pool_volunteer_rows(volunteers, pool_id))


# ==================== Тайм-блоки и запись ====================


def block_to_dict(block, signups_by_block):
    rows = signups_by_block.get(block.id, [])
    return {
        'id': block.id,
        'pool_id': block.pool_id,
        'date': block.date.isoformat(),
        'time_start': block.time_start,
        'time_end': block.time_end,
        'label': block.label or '',
        'capacity': block.capacity,
        'volunteers': [{
            'user_id': uid,
            'nick': nick,
            'name': name or nick,
            'telegram': telegram,
            'role': role,
        } for uid, nick, name, telegram, role in rows],
        'count': len(rows),
    }


def _signups_index(pool_id):
    """Вернуть {block_id: [(user_id, nick, name, telegram, role), ...]} для бассейна."""
    rows = (
        db.session.query(
            Signup.block_id,
            User.id,
            User.nick,
            User.name,
            User.telegram,
            User.role,
            PoolVolunteer.pool_role,
        )
        .join(User, User.id == Signup.user_id)
        .join(ShiftBlock, ShiftBlock.id == Signup.block_id)
        .outerjoin(
            PoolVolunteer,
            db.and_(PoolVolunteer.user_id == User.id, PoolVolunteer.pool_id == pool_id),
        )
        .filter(ShiftBlock.pool_id == pool_id)
        .order_by(Signup.created_at)
        .all()
    )
    index = {}
    for block_id, uid, nick, name, telegram, global_role, pool_role in rows:
        role = global_role if global_role in ('team_lead', 'admin') else (pool_role or 'volunteer')
        index.setdefault(block_id, []).append((uid, nick, name, telegram, role))
    return index


@app.route('/api/schedule', methods=['GET'])
@require_auth
def schedule():
    """Сетка графика: дни -> тайм-блоки -> волонтёры."""
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    pool = get_model_or_404(Pool, pool_id)
    if not can_access_pool(g.user, pool):
        return jsonify({'error': 'Ты не добавлен на этот бассейн. Обратись к тимлиду.'}), 403

    blocks = ShiftBlock.query.filter_by(pool_id=pool_id).order_by(
        ShiftBlock.date, ShiftBlock.time_start
    ).all()
    index = _signups_index(pool_id)

    days = {}
    for b in blocks:
        key = b.date.isoformat()
        days.setdefault(key, []).append(block_to_dict(b, index))

    days_list = [{'date': d, 'blocks': blocks} for d, blocks in sorted(days.items())]
    return jsonify({'pool': pool.to_dict(), 'days': days_list})


@app.route('/api/schedule/settings', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_schedule_settings():
    pool_id, error = _active_pool_id_for_request()
    if error:
        return error
    data = request.get_json(silent=True) or {}
    enabled = data.get('self_signup_enabled')
    if not isinstance(enabled, bool):
        return jsonify({'error': 'self_signup_enabled должен быть true или false'}), 400

    pool = get_model_or_404(Pool, pool_id)
    previous = pool.self_signup_enabled is not False
    pool.self_signup_enabled = enabled
    log_action(
        'update',
        'schedule_settings',
        pool.id,
        'Самозапись на смены включена' if enabled else 'Самозапись на смены выключена',
        {'self_signup_enabled': enabled, 'previous_self_signup_enabled': previous},
    )
    db.session.commit()
    return jsonify({
        'self_signup_enabled': enabled,
        'message': 'Самозапись включена' if enabled else 'Самозапись выключена',
    })


def _self_signup_access_error(user, pool):
    if pool.self_signup_enabled is not False:
        return None
    if _effective_access_role(user, pool.id) in {'team_lead', 'admin'}:
        return None
    return jsonify({'error': 'Самозапись на смены временно выключена'}), 403


@app.route('/api/blocks', methods=['POST'])
@require_role('team_lead', 'admin')
def create_block():
    data = request.json or {}
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    try:
        d = datetime.fromisoformat(data['date']).date()
    except (KeyError, ValueError):
        return jsonify({'error': 'Некорректная дата'}), 400
    time_start = (data.get('time_start') or '10:00').strip()
    time_end = (data.get('time_end') or '14:00').strip()
    if not TIME_VALUE_RE.fullmatch(time_start) or not TIME_VALUE_RE.fullmatch(time_end):
        return jsonify({'error': 'Время должно быть в формате HH:MM'}), 400
    if time_start >= time_end:
        return jsonify({'error': 'Время окончания должно быть позже времени начала'}), 400
    capacity = data.get('capacity')
    if capacity is not None:
        try:
            capacity = int(capacity)
        except (TypeError, ValueError):
            return jsonify({'error': 'Вместимость должна быть целым числом'}), 400
        if capacity < 1:
            return jsonify({'error': 'Вместимость должна быть больше нуля'}), 400
    overlapping = ShiftBlock.query.filter(
        ShiftBlock.pool_id == pool_id,
        ShiftBlock.date == d,
        ShiftBlock.time_start < time_end,
        ShiftBlock.time_end > time_start,
    ).first()
    if overlapping:
        return jsonify({'error': 'Этот интервал пересекается с существующим тайм-блоком'}), 409
    block = ShiftBlock(
        pool_id=pool_id,
        date=d,
        time_start=time_start,
        time_end=time_end,
        label=data.get('label', ''),
        capacity=capacity,
    )
    db.session.add(block)
    db.session.commit()
    return jsonify({'id': block.id, 'message': 'Тайм-блок создан'}), 201


@app.route('/api/blocks/<int:block_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_block(block_id):
    block = get_model_or_404(ShiftBlock, block_id)
    error = _active_entity_error(block.pool_id)
    if error:
        return error
    Signup.query.filter_by(block_id=block.id).delete()
    db.session.delete(block)
    db.session.commit()
    return jsonify({'message': 'Тайм-блок удалён'})


@app.route('/api/blocks/<int:block_id>/capacity', methods=['PATCH'])
@require_role('team_lead', 'admin')
def patch_block_capacity(block_id):
    block = get_model_or_404(ShiftBlock, block_id)
    error = _active_entity_error(block.pool_id)
    if error:
        return error
    data = request.json or {}
    if 'capacity' in data:
        val = data['capacity']
        try:
            next_capacity = int(val) if val is not None else None
        except (TypeError, ValueError):
            return jsonify({'error': 'Вместимость должна быть целым числом'}), 400
        current_count = Signup.query.filter_by(block_id=block.id).count()
        if next_capacity is not None and next_capacity < max(1, current_count):
            return jsonify({'error': 'Вместимость не может быть меньше числа записанных'}), 409
        block.capacity = next_capacity
    else:
        delta = int(data.get('delta', 1))
        base = block.capacity if block.capacity is not None else Signup.query.filter_by(block_id=block.id).count()
        new_cap = base + delta
        if delta < 0:
            current_count = Signup.query.filter_by(block_id=block.id).count()
            new_cap = max(current_count, new_cap)
        block.capacity = max(1, new_cap)
    db.session.commit()
    return jsonify({'id': block.id, 'capacity': block.capacity})


# Стандартный шаблон расписания School21 pool из Google-таблицы.
# (время_от, время_до, метка, capacity). Capacity = сколько волонтёров может записаться.
_SCHEDULE_TPL = {
    0: [('09:00', '19:00', '', 7), ('19:00', '20:00', '', 2)],    # стартовый понедельник
    1: [('10:00', '14:00', '', 4), ('15:00', '19:00', '', 4)],
    2: [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)],
    3: [('11:00', '17:00', 'EXAM', 5)],
    4: [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)],
    5: [('10:00', '14:00', '', 1), ('15:00', '19:00', '', 1)],
    6: [('10:00', '14:00', '', 1), ('15:00', '19:00', '', 1)],
    7: [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)],
    8: [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)],
    9: [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)],
    10: [('11:00', '17:00', 'EXAM', 4)],
    11: [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)],
    12: [('10:00', '14:00', '', 1), ('15:00', '19:00', '', 1)],
    13: [('10:00', '14:00', '', 1), ('15:00', '19:00', '', 1)],
}


def _schedule_template_for_day(day_index):
    if day_index in _SCHEDULE_TPL:
        return _SCHEDULE_TPL[day_index]
    weekday = day_index % 7
    if weekday == 3:
        return [('11:00', '17:00', 'EXAM', 4)]
    if weekday in (5, 6):
        return [('10:00', '14:00', '', 1), ('15:00', '19:00', '', 1)]
    return [('10:00', '14:00', '', 2), ('15:00', '19:00', '', 2)]


@app.route('/api/pools/<int:pool_id>/generate-schedule', methods=['POST'])
@require_role('team_lead', 'admin')
def generate_schedule(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    error = _active_entity_error(pool.id)
    if error:
        return error
    if not pool.start_date:
        return jsonify({'error': 'У бассейна не задана дата начала'}), 400
    data = request.json or {}
    try:
        raw_end_date = data.get('end_date')
        end_date = (
            datetime.fromisoformat(raw_end_date).date()
            if raw_end_date
            else pool.start_date + timedelta(days=13)
        )
    except (KeyError, ValueError):
        return jsonify({'error': 'Укажите end_date (YYYY-MM-DD)'}), 400
    if end_date < pool.start_date:
        return jsonify({'error': 'Дата окончания раньше даты начала'}), 400

    created = 0
    updated = 0
    generation = ScheduleGeneration(pool_id=pool.id, end_date=end_date, created_by=g.user.id)
    db.session.add(generation)
    db.session.flush()
    current = pool.start_date
    day_index = 0
    while current <= end_date:
        tpl = _schedule_template_for_day(day_index)
        for t1, t2, label, cap in tpl:
            existing = ShiftBlock.query.filter_by(
                pool_id=pool_id,
                date=current,
                time_start=t1,
                time_end=t2,
                label=label,
            ).first()
            if existing:
                if existing.capacity != cap:
                    existing.capacity = cap
                    updated += 1
                continue
            db.session.add(
                ShiftBlock(
                    pool_id=pool_id,
                    date=current,
                    time_start=t1,
                    time_end=t2,
                    label=label,
                    capacity=cap,
                    generation_id=generation.id,
                )
            )
            created += 1
        current += timedelta(days=1)
        day_index += 1

    db.session.commit()
    return jsonify({
        'created': created,
        'updated': updated,
        'message': f'Создано {created} тайм-блоков, обновлено {updated}',
    })


@app.route('/api/pools/<int:pool_id>/generate-schedule/undo', methods=['POST'])
@require_role('team_lead', 'admin')
def undo_generate_schedule(pool_id):
    error = _active_entity_error(pool_id)
    if error:
        return error
    generation = (
        ScheduleGeneration.query
        .filter_by(pool_id=pool_id)
        .order_by(ScheduleGeneration.created_at.desc(), ScheduleGeneration.id.desc())
        .first()
    )
    if not generation:
        return jsonify({'error': 'Нет последней генерации для отмены'}), 404
    blocks = ShiftBlock.query.filter_by(pool_id=pool_id, generation_id=generation.id).all()
    block_ids = [block.id for block in blocks]
    if block_ids:
        Signup.query.filter(Signup.block_id.in_(block_ids)).delete(synchronize_session=False)
        ShiftBlock.query.filter(ShiftBlock.id.in_(block_ids)).delete(synchronize_session=False)
    db.session.delete(generation)
    db.session.commit()
    return jsonify({'deleted': len(block_ids), 'message': f'Удалено тайм-блоков: {len(block_ids)}'})


@app.route('/api/blocks/<int:block_id>/signup', methods=['POST'])
@require_auth
def signup_block(block_id):
    block = ShiftBlock.query.filter_by(id=block_id).with_for_update().first()
    if not block:
        abort(404)
    error = _active_entity_error(block.pool_id)
    if error:
        return error
    user = g.user
    # Проверка доступа к бассейну
    pool = db.session.get(Pool, block.pool_id)
    if pool and not can_access_pool(user, pool):
        return jsonify({'error': 'Ты не добавлен на этот бассейн'}), 403
    signup_error = _self_signup_access_error(user, pool)
    if signup_error:
        return signup_error
    data = request.get_json(silent=True) or {}
    target_id = data.get('user_id')
    if target_id and user.role in ('team_lead', 'admin'):
        target_user = get_model_or_404(User, int(target_id))
    elif target_id:
        return jsonify({'error': 'Назначать других людей может только тимлид или админ'}), 403
    else:
        target_user = user

    if not PoolVolunteer.query.filter_by(pool_id=block.pool_id, user_id=target_user.id).first():
        return jsonify({'error': 'Пользователь не назначен на активный бассейн'}), 409

    now_moscow = _moscow_now()
    if block.date < now_moscow.date() or (
        block.date == now_moscow.date() and block.time_end <= now_moscow.strftime('%H:%M')
    ):
        return jsonify({'error': 'Нельзя записаться на завершённую смену'}), 409

    existing = Signup.query.filter_by(block_id=block.id, user_id=target_user.id).first()
    if existing:
        return jsonify({'message': 'Уже записан'}), 200
    if block.capacity is not None:
        current = Signup.query.filter_by(block_id=block.id).count()
        if current >= block.capacity:
            return jsonify({'error': 'Мест больше нет'}), 409
    overlap = (
        Signup.query.join(ShiftBlock, Signup.block_id == ShiftBlock.id)
        .filter(
            Signup.user_id == target_user.id,
            ShiftBlock.pool_id == block.pool_id,
            ShiftBlock.date == block.date,
            ShiftBlock.id != block.id,
            ShiftBlock.time_start < block.time_end,
            ShiftBlock.time_end > block.time_start,
        )
        .first()
    )
    if overlap:
        return jsonify({'error': 'У пользователя уже есть пересекающаяся смена'}), 409
    signup = Signup(block_id=block.id, user_id=target_user.id)
    db.session.add(signup)
    try:
        db.session.flush()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'message': 'Уже записан'}), 200
    log_action(
        'create',
        'signup',
        block.id,
        f'@{target_user.nick} записан на смену {block.date.isoformat()} {block.time_start}-{block.time_end}',
        {
            'target_user_id': target_user.id,
            'target_nick': target_user.nick,
            'date': block.date.isoformat(),
            'time_start': block.time_start,
            'time_end': block.time_end,
            'label': block.label or '',
        },
    )
    _queue_shift_change_notifications(block, target_user, 'create')
    db.session.commit()
    return jsonify({'message': 'Записан на смену'}), 201


@app.route('/api/blocks/<int:block_id>/signup', methods=['DELETE'])
@require_auth
def unsignup_block(block_id):
    user = g.user
    block = get_model_or_404(ShiftBlock, block_id)
    error = _active_entity_error(block.pool_id)
    if error:
        return error
    pool = db.session.get(Pool, block.pool_id)
    signup_error = _self_signup_access_error(user, pool)
    if signup_error:
        return signup_error
    # волонтёр снимает себя; тимлид/админ может снять любого через ?user_id=
    target_id = request.args.get('user_id', type=int)
    if target_id and user.role in ('team_lead', 'admin'):
        target = target_id
    else:
        target = user.id
    signup = Signup.query.filter_by(block_id=block.id, user_id=target).first()
    if not signup:
        return jsonify({'message': 'Записи не было'}), 200
    db.session.delete(signup)
    target_user = db.session.get(User, target)
    log_action(
        'delete',
        'signup',
        block.id,
        f'@{target_user.nick if target_user else target} снят со смены {block.date.isoformat()} {block.time_start}-{block.time_end}',
        {
            'target_user_id': target,
            'target_nick': target_user.nick if target_user else str(target),
            'date': block.date.isoformat(),
            'time_start': block.time_start,
            'time_end': block.time_end,
            'label': block.label or '',
        },
    )
    if target_user:
        _queue_shift_change_notifications(block, target_user, 'delete')
    db.session.commit()
    return jsonify({'message': 'Запись отменена'})


@app.route('/api/me/shifts', methods=['GET'])
@require_auth
def my_shifts():
    pool_id = active_pool_id()
    if not pool_id:
        return jsonify([])
    rows = (
        db.session.query(ShiftBlock)
        .join(Signup, Signup.block_id == ShiftBlock.id)
        .filter(Signup.user_id == g.user.id, ShiftBlock.pool_id == pool_id)
        .order_by(ShiftBlock.date, ShiftBlock.time_start)
        .all()
    )
    return jsonify([{
        'id': b.id,
        'date': b.date.isoformat(),
        'time_start': b.time_start,
        'time_end': b.time_end,
        'label': b.label or '',
    } for b in rows])


# ==================== Статистика ====================


@app.route('/api/stats', methods=['GET'])
@require_auth
def stats():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    pool = db.session.get(Pool, pool_id)
    total_blocks = ShiftBlock.query.filter_by(pool_id=pool_id).count() if pool_id else 0
    volunteers = PoolVolunteer.query.filter_by(pool_id=pool_id).count() if pool_id else 0
    total_signups = (
        db.session.query(Signup).join(ShiftBlock, ShiftBlock.id == Signup.block_id)
        .filter(ShiftBlock.pool_id == pool_id).count() if pool_id else 0
    )
    my = (
        db.session.query(Signup)
        .join(ShiftBlock, ShiftBlock.id == Signup.block_id)
        .filter(Signup.user_id == g.user.id, ShiftBlock.pool_id == pool_id)
        .count()
        if pool_id else 0
    )
    return jsonify({
        'pool': pool.to_dict() if pool else None,
        'totalBlocks': total_blocks,
        'volunteers': volunteers,
        'totalSignups': total_signups,
        'mySignups': my,
    })


@app.route('/api/tribes', methods=['GET'])
@require_auth
def list_tribes():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    tribes = Tribe.query.filter_by(pool_id=pool_id).order_by(Tribe.name).all()
    return jsonify([{'id': t.id, 'name': t.name} for t in tribes])


@app.route('/api/tribes', methods=['POST'])
@require_role('team_lead', 'admin')
def create_tribe():
    data = request.json or {}
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Укажите название трайба'}), 400
    if Tribe.query.filter(Tribe.pool_id == pool_id, db.func.lower(Tribe.name) == name.lower()).first():
        return jsonify({'error': f'Трайб «{name}» уже существует'}), 409
    tribe = Tribe(name=name, pool_id=pool_id)
    db.session.add(tribe)
    db.session.commit()
    return jsonify({'id': tribe.id, 'name': tribe.name, 'pool_id': tribe.pool_id}), 201


@app.route('/api/tribes/<int:tribe_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_tribe(tribe_id):
    tribe = get_model_or_404(Tribe, tribe_id)
    error = _active_entity_error(tribe.pool_id)
    if error:
        return error
    db.session.delete(tribe)
    db.session.commit()
    return jsonify({'message': f'Трайб «{tribe.name}» удалён'})


@app.route('/api/tribes/load-standard', methods=['POST'])
@require_role('team_lead', 'admin')
def load_standard_tribes():
    pool_id = active_pool_id()
    if not pool_id:
        return jsonify({'error': 'Нет активного бассейна'}), 400
    added = 0
    for name in STANDARD_TRIBES_NNV:
        if not Tribe.query.filter(Tribe.pool_id == pool_id, db.func.lower(Tribe.name) == name.lower()).first():
            db.session.add(Tribe(name=name, pool_id=pool_id))
            added += 1
    db.session.commit()
    return jsonify({'message': f'Добавлено {added} трайбов', 'added': added})


@app.route('/api/volunteers', methods=['GET'])
@require_auth
def get_volunteers():
    """Список людей, участвующих в волонтёрской сетке, со статусами и коинами."""
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error

    if pool_id:
        pvs = PoolVolunteer.query.filter_by(pool_id=pool_id).all()
        pv_map = {pv.user_id: pv for pv in pvs}
        users = User.query.filter(User.id.in_(list(pv_map.keys()))).all()
        user_ids = [user.id for user in users]
        pool = db.session.get(Pool, pool_id)
        shift_counts = defaultdict(int)
        reward_buckets_by_user = defaultdict(dict)

        signup_rows = (
            db.session.query(Signup.user_id, ShiftBlock)
            .join(ShiftBlock, ShiftBlock.id == Signup.block_id)
            .filter(Signup.user_id.in_(user_ids), ShiftBlock.pool_id == pool_id)
            .all()
            if user_ids else []
        )
        for user_id, block in signup_rows:
            shift_counts[user_id] += 1
            hours = _block_hours(block)
            reward_type, label, rate = _shift_reward_type(block, pool)
            _add_reward(reward_buckets_by_user[user_id], reward_type, label, hours, int(hours * rate))

        group_rows = (
            db.session.query(
                GroupReview.reviewer_id,
                db.func.coalesce(db.func.sum(GroupReview.quantity), 0),
            )
            .filter(GroupReview.pool_id == pool_id, GroupReview.reviewer_id.in_(user_ids))
            .group_by(GroupReview.reviewer_id)
            .all()
            if user_ids else []
        )
        group_counts = {user_id: int(total or 0) for user_id, total in group_rows}

        tribe_event_rows = (
            db.session.query(TribeEvent.tribe, db.func.count(TribeEvent.id))
            .filter(TribeEvent.pool_id == pool_id)
            .group_by(TribeEvent.tribe)
            .all()
        )
        tribe_event_counts = {normalize_tribe(tribe): int(total or 0) for tribe, total in tribe_event_rows if tribe}

        reward_events = (
            RewardEvent.query
            .filter(
                RewardEvent.pool_id == pool_id,
                RewardEvent.user_id.in_(user_ids),
                RewardEvent.event_type != 'confession',
            )
            .all()
            if user_ids else []
        )
        reward_events_by_user = defaultdict(list)
        for event in reward_events:
            reward_events_by_user[event.user_id].append(event)

        result = []
        for user in users:
            pv = pv_map[user.id]
            role = pv.pool_role or 'volunteer'
            has_conf = bool(pv.has_confession)
            adj = pv.coins_adjustment or 0
            group_cnt = group_counts.get(user.id, 0)
            buckets = reward_buckets_by_user[user.id]
            if group_cnt:
                _add_reward(buckets, 'group_review', 'Проверка групповых', group_cnt, group_cnt * REWARD_RATES['group_review'])
            if has_conf:
                _add_reward(buckets, 'confession', 'Исповедь', 1, REWARD_RATES['confession'])
            normalized_tribe = normalize_tribe(pv.tribe) if pv.tribe else None
            if role == 'tribe_master' and normalized_tribe:
                tribe_events_count = tribe_event_counts.get(normalized_tribe, 0)
                if tribe_events_count:
                    _add_reward(
                        buckets,
                        'tribe_master_event',
                        'Трайб-мастерство',
                        tribe_events_count,
                        tribe_events_count * REWARD_RATES['tribe_master_event'],
                    )
            for event in reward_events_by_user.get(user.id, []):
                meta = REWARD_EVENT_TYPES.get(event.event_type, {'label': event.event_type})
                _add_reward(buckets, event.event_type, meta['label'], event.quantity or 1, event.coins)
            if adj:
                _add_reward(buckets, 'manual', 'Ручная корректировка', 1, adj)
            breakdown = list(buckets.values())
            total = sum(item['coins'] for item in breakdown)
            result.append({
                'id': user.id,
                'nick': user.nick,
                'name': user.name or user.nick,
                'telegram': user.telegram,
                'role': role,
                'tribe': pv.tribe,
                'avatar_url': _avatar_url_for_user(user),
                'is_group_reviewer': bool(group_cnt),
                'group_reviews_count': int(group_cnt),
                'has_confession': has_conf,
                'shifts_count': shift_counts.get(user.id, 0),
                'coins': total,
                'coins_adjustment': adj,
                'coin_breakdown': breakdown,
            })
        order = {'team_lead': 0, 'tribe_master': 1, 'volunteer': 2}
        result.sort(key=lambda x: (order.get(x['role'], 9), x['nick']))
        return jsonify(result)

    # Без pool_id — глобальный список (обратная совместимость)
    users = User.query.filter(User.role.in_(list(VOLUNTEER_PROFILE_ROLES))).order_by(User.nick).all()
    result = []
    for user in users:
        cnt = Signup.query.filter_by(user_id=user.id).count()
        manual = user.coins_adjustment or 0
        rewards = calculate_user_rewards(user, manual)
        group_reviews_count = db.session.query(db.func.coalesce(db.func.sum(GroupReview.quantity), 0)).filter_by(reviewer_id=user.id).scalar() or 0
        result.append({
            'id': user.id,
            'nick': user.nick,
            'name': user.name or user.nick,
            'telegram': user.telegram,
            'avatar_url': _avatar_url_for_user(user),
            'role': user.role,
            'tribe': user.tribe,
            'is_group_reviewer': bool(group_reviews_count),
            'group_reviews_count': int(group_reviews_count),
            'has_confession': bool(user.has_confession),
            'shifts_count': cnt,
            'coins': rewards['total'],
            'coins_adjustment': manual,
            'coin_breakdown': rewards['breakdown'],
        })
    order = {'team_lead': 0, 'tribe_master': 1, 'volunteer': 2}
    result.sort(key=lambda x: (order.get(x['role'], 9), x['nick']))
    return jsonify(result)


def _parse_time(value):
    return datetime.strptime(value, '%H:%M')


def _block_hours(block):
    start = _parse_time(block.time_start)
    end = _parse_time(block.time_end)
    return max(0, (end - start).seconds / 3600)


def _shift_reward_type(block, pool):
    if block.label == 'EXAM':
        return 'exam_hour', 'Экзамен', REWARD_RATES['exam_hour']
    if pool and pool.start_date and block.date == pool.start_date:
        return 'first_day_hour', 'Первый день', REWARD_RATES['first_day_hour']
    if block.date.weekday() >= 5:
        return 'first_week_or_weekend_hour', 'Выходной день', REWARD_RATES['first_week_or_weekend_hour']
    if pool and pool.start_date and block.date < pool.start_date + timedelta(days=7):
        return 'first_week_or_weekend_hour', 'Первая неделя', REWARD_RATES['first_week_or_weekend_hour']
    return 'subsequent_weekday_hour', 'Будний день после первой недели', REWARD_RATES['subsequent_weekday_hour']


def _add_reward(buckets, reward_type, label, count, coins):
    if not count and not coins:
        return
    item = buckets.setdefault(reward_type, {'type': reward_type, 'label': label, 'count': 0, 'coins': 0})
    item['count'] += count
    item['coins'] += coins


def calculate_user_rewards(user, manual_adjustment=0):
    buckets = {}
    rows = (
        db.session.query(Signup, ShiftBlock, Pool)
        .join(ShiftBlock, ShiftBlock.id == Signup.block_id)
        .join(Pool, Pool.id == ShiftBlock.pool_id)
        .filter(Signup.user_id == user.id)
        .all()
    )
    for _, block, pool in rows:
        hours = _block_hours(block)
        reward_type, label, rate = _shift_reward_type(block, pool)
        _add_reward(buckets, reward_type, label, hours, int(hours * rate))

    group_reviews = db.session.query(db.func.coalesce(db.func.sum(GroupReview.quantity), 0)).filter_by(reviewer_id=user.id).scalar() or 0
    if group_reviews:
        group_reviews = int(group_reviews)
        _add_reward(
            buckets,
            'group_review',
            'Проверка групповых',
            group_reviews,
            group_reviews * REWARD_RATES['group_review'],
        )
    if user.has_confession:
        _add_reward(buckets, 'confession', 'Исповедь', 1, REWARD_RATES['confession'])
    if user.role == 'tribe_master':
        _add_reward(buckets, 'tribe_master_event', 'Трайб-мастерство', 1, REWARD_RATES['tribe_master_event'])
    if user.role == 'team_lead':
        _add_reward(buckets, 'team_lead', 'Тимлид команды волонтёров', 1, REWARD_RATES['team_lead'])
    events = RewardEvent.query.filter(
        RewardEvent.user_id == user.id,
        RewardEvent.event_type != 'confession',
    ).all()
    for event in events:
        meta = REWARD_EVENT_TYPES.get(event.event_type, {'label': event.event_type})
        _add_reward(buckets, event.event_type, meta['label'], event.quantity or 1, event.coins)
    if manual_adjustment:
        _add_reward(buckets, 'manual', 'Ручная корректировка', 1, manual_adjustment)

    breakdown = list(buckets.values())
    return {'breakdown': breakdown, 'total': sum(item['coins'] for item in breakdown)}


def calculate_pool_rewards(user, pool_id, has_confession, coins_adjustment, pool_role, pv_tribe=None):
    """Расчёт наград только в рамках одного бассейна."""
    pool = db.session.get(Pool, pool_id)
    buckets = {}
    rows = (
        db.session.query(Signup, ShiftBlock)
        .join(ShiftBlock, ShiftBlock.id == Signup.block_id)
        .filter(Signup.user_id == user.id, ShiftBlock.pool_id == pool_id)
        .all()
    )
    for _, block in rows:
        hours = _block_hours(block)
        reward_type, label, rate = _shift_reward_type(block, pool)
        _add_reward(buckets, reward_type, label, hours, int(hours * rate))

    group_reviews = (
        db.session.query(db.func.coalesce(db.func.sum(GroupReview.quantity), 0))
        .filter(GroupReview.reviewer_id == user.id, GroupReview.pool_id == pool_id)
        .scalar() or 0
    )
    if group_reviews:
        group_reviews = int(group_reviews)
        _add_reward(buckets, 'group_review', 'Проверка групповых', group_reviews, group_reviews * REWARD_RATES['group_review'])

    if has_confession:
        _add_reward(buckets, 'confession', 'Исповедь', 1, REWARD_RATES['confession'])
    if pool_role == 'tribe_master' and pv_tribe:
        tribe_events_count = TribeEvent.query.filter_by(pool_id=pool_id, tribe=pv_tribe).count()
        if tribe_events_count:
            _add_reward(
                buckets,
                'tribe_master_event',
                'Трайб-мастерство',
                tribe_events_count,
                tribe_events_count * REWARD_RATES['tribe_master_event'],
            )

    events = RewardEvent.query.filter(
        RewardEvent.user_id == user.id,
        RewardEvent.pool_id == pool_id,
        RewardEvent.event_type != 'confession',
    ).all()
    for event in events:
        meta = REWARD_EVENT_TYPES.get(event.event_type, {'label': event.event_type})
        _add_reward(buckets, event.event_type, meta['label'], event.quantity or 1, event.coins)

    if coins_adjustment:
        _add_reward(buckets, 'manual', 'Ручная корректировка', 1, coins_adjustment)

    breakdown = list(buckets.values())
    return {'breakdown': breakdown, 'total': sum(item['coins'] for item in breakdown)}


def _volunteer_role_from_payload(data):
    role = data.get('role') or 'volunteer'
    if role not in ('volunteer', 'tribe_master'):
        raise ValueError('Во вкладке волонтёров можно добавлять только волонтёров и трайб-мастеров')
    return role


def _cell_column(cell_ref):
    return ''.join(ch for ch in cell_ref if ch.isalpha())


def _column_index(col):
    index = 0
    for ch in col:
        index = index * 26 + (ord(ch.upper()) - ord('A') + 1)
    return index - 1


def parse_xlsx_rows(file_storage):
    """Прочитать первый лист .xlsx в список строк. Достаточно для простых импорт-шаблонов."""
    ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    with zipfile.ZipFile(file_storage) as archive:
        shared = []
        if 'xl/sharedStrings.xml' in archive.namelist():
            root = ET.fromstring(archive.read('xl/sharedStrings.xml'))
            for item in root.findall('m:si', ns):
                shared.append(''.join(t.text or '' for t in item.findall('.//m:t', ns)))

        sheet_path = 'xl/worksheets/sheet1.xml'
        root = ET.fromstring(archive.read(sheet_path))
        rows = []
        for row in root.findall('.//m:sheetData/m:row', ns):
            values = []
            for cell in row.findall('m:c', ns):
                col_idx = _column_index(_cell_column(cell.attrib.get('r', 'A1')))
                while len(values) <= col_idx:
                    values.append('')
                cell_type = cell.attrib.get('t')
                if cell_type == 'inlineStr':
                    value = ''.join(t.text or '' for t in cell.findall('.//m:t', ns))
                else:
                    raw = cell.find('m:v', ns)
                    value = raw.text if raw is not None and raw.text is not None else ''
                    if cell_type == 's' and value:
                        value = shared[int(value)]
                values[col_idx] = value.strip() if isinstance(value, str) else value
            if any(str(v).strip() for v in values):
                rows.append(values)
        return rows


def rows_to_dicts(rows, columns):
    if not rows:
        return []
    header_map = {
        'nick': 'nick', 'ник': 'nick', 'login': 'nick', 'логин': 'nick', 'ник школьный': 'nick',
        'name': 'name', 'имя': 'name', 'фио': 'name', 'имя фамилия': 'name',
        'role': 'role', 'статус': 'role', 'роль': 'role',
        'tribe': 'tribe', 'группа': 'tribe', 'триб': 'tribe', 'трайб': 'tribe',
        'telegram': 'telegram', 'ник telegram': 'telegram', 'tg': 'telegram',
    }
    first = [header_map.get(str(cell).strip().lower()) for cell in rows[0]]
    has_header = any(first)
    keys = first if has_header else columns
    data_rows = rows[1:] if has_header else rows
    result = []
    for row in data_rows:
        item = {}
        for index, key in enumerate(keys):
            if key and index < len(row):
                item[key] = row[index]
        result.append(item)
    return result


def normalize_tribe(value):
    raw = str(value or '').strip()
    if not raw:
        return None
    return TRIBE_ALIASES.get(raw.lower(), raw)


def tribe_master_for_tribe(pool_id, tribe, exclude_user_id=None):
    if not pool_id or not tribe:
        return None
    query = PoolVolunteer.query.filter_by(
        pool_id=pool_id,
        pool_role='tribe_master',
        tribe=tribe,
    )
    if exclude_user_id:
        query = query.filter(PoolVolunteer.user_id != exclude_user_id)
    return query.first()


def validate_tribe_master_assignment(pool_id, role, tribe, exclude_user_id=None):
    if role != 'tribe_master' or not tribe:
        return None
    owner = tribe_master_for_tribe(pool_id, tribe, exclude_user_id)
    if owner:
        user = db.session.get(User, owner.user_id)
        return f'Трайб "{tribe}" уже назначен @{user.nick if user else owner.user_id}'
    return None


def save_volunteer_rows(rows):
    created = 0
    updated = 0
    skipped = []

    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            skipped.append({'row': index, 'reason': 'Некорректная строка'})
            continue

        nick = (row.get('nick') or '').strip()
        name = (row.get('name') or '').strip() or nick
        telegram = _normalize_telegram_value(row.get('telegram'))
        if not nick:
            skipped.append({'row': index, 'reason': 'Нужен nick'})
            continue

        user = User.query.filter(db.func.lower(User.nick) == nick.lower()).first()
        if telegram and telegram_username_conflict(telegram, user.id if user else None):
            skipped.append({'row': index, 'nick': nick, 'reason': 'Telegram username уже используется'})
            continue
        if user:
            if user.role in ROLES_WITH_PASSWORD:
                skipped.append({'row': index, 'nick': nick, 'reason': 'Ник уже занят тимлидом или админом'})
                continue
            user.name = name
            if telegram:
                user.telegram = telegram
            updated += 1
        else:
            db.session.add(User(
                nick=nick,
                name=name,
                role='volunteer',
                telegram=telegram,
            ))
            created += 1

    log_action(
        'import',
        'volunteer',
        None,
        f'Импорт волонтёров: новых {created}, обновлено {updated}, пропущено {len(skipped)}',
        {'created': created, 'updated': updated, 'skipped': skipped},
    )
    db.session.commit()
    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'message': f'Импортировано: новых {created}, обновлено {updated}, пропущено {len(skipped)}',
    }


@app.route('/api/volunteers', methods=['POST'])
@require_role('team_lead', 'admin')
def create_volunteer_profile():
    data = request.json or {}
    nick = (data.get('nick') or '').strip()
    name = (data.get('name') or '').strip() or nick
    if not nick:
        return jsonify({'error': 'Укажите ник'}), 400
    try:
        role = _volunteer_role_from_payload(data)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    if role != 'volunteer':
        return jsonify({'error': 'Роль назначается отдельно внутри активного бассейна'}), 400

    existing = User.query.filter(db.func.lower(User.nick) == nick.lower()).first()
    if existing:
        if existing.role in ROLES_WITH_PASSWORD:
            return jsonify({'error': 'Такой ник уже есть как тимлид или админ'}), 409
        return jsonify({'error': 'Такой ник уже есть'}), 409

    telegram = _normalize_telegram_value(data.get('telegram'))
    if telegram_username_conflict(telegram):
        return jsonify({'error': 'Этот Telegram username уже используется'}), 409
    user = User(
        nick=nick,
        name=name,
        role=role,
        telegram=telegram,
        tribe=None,
    )
    db.session.add(user)
    db.session.flush()
    log_action(
        'create',
        'volunteer',
        user.id,
        f'Добавлен @{user.nick} как {role}',
        {'target_nick': user.nick, 'role': role},
    )
    db.session.commit()
    return jsonify(user.to_dict()), 201


@app.route('/api/volunteers/import', methods=['POST'])
@require_role('team_lead', 'admin')
def import_volunteer_profiles():
    data = request.json or {}
    rows = data.get('volunteers') or []
    if not isinstance(rows, list) or not rows:
        return jsonify({'error': 'Передайте volunteers: [{nick, name, role?, tribe?}]'}), 400
    return jsonify(save_volunteer_rows(rows))


@app.route('/api/volunteers/import-file', methods=['POST'])
@require_role('team_lead', 'admin')
def import_volunteer_profiles_file():
    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({'error': 'Загрузите файл'}), 400
    try:
        rows = parse_xlsx_rows(uploaded)
        volunteers = rows_to_dicts(rows, ['name', 'nick', 'telegram'])
    except Exception as e:
        return jsonify({'error': f'Не удалось прочитать .xlsx: {e}'}), 400
    return jsonify(save_volunteer_rows(volunteers))


@app.route('/api/volunteers/<int:user_id>/reward-events', methods=['POST'])
@require_role('team_lead', 'admin')
def create_volunteer_reward_event(user_id):
    user = get_model_or_404(User, user_id)
    data = request.json or {}
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    membership = PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=user.id).first()
    if not membership:
        return jsonify({'error': 'Пользователь не назначен на активный бассейн'}), 409
    event_type = data.get('event_type') or 'confession'
    if event_type not in REWARD_EVENT_TYPES:
        return jsonify({'error': 'Неизвестный тип активности'}), 400
    if event_type == 'confession':
        membership.has_confession = True
        RewardEvent.query.filter_by(user_id=user.id, pool_id=pool_id, event_type='confession').delete()
        db.session.commit()
        return jsonify({
            'message': f'{REWARD_EVENT_TYPES[event_type]["label"]}: +{REWARD_RATES["confession"]} коинов для @{user.nick}',
        }), 200

    try:
        quantity = max(1, int(data.get('quantity') or 1))
    except (TypeError, ValueError):
        return jsonify({'error': 'Количество должно быть числом'}), 400

    event_date = None
    if data.get('event_date'):
        try:
            event_date = datetime.fromisoformat(data['event_date']).date()
        except ValueError:
            return jsonify({'error': 'Некорректная дата'}), 400

    coins = quantity * REWARD_EVENT_TYPES[event_type]['coins']
    event = RewardEvent(
        user_id=user.id,
        pool_id=pool_id,
        event_type=event_type,
        event_date=event_date,
        quantity=quantity,
        coins=coins,
        comment=(data.get('comment') or '').strip(),
        created_by=g.user.id,
    )
    db.session.add(event)
    db.session.commit()
    return jsonify({
        'id': event.id,
        'message': f'{REWARD_EVENT_TYPES[event_type]["label"]}: +{coins} коинов для @{user.nick}',
    }), 201


@app.route('/api/volunteers/<int:user_id>', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_volunteer_profile(user_id):
    user = get_model_or_404(User, user_id)
    data = request.json or {}
    changes = {}

    pool_scoped_fields = {'role', 'has_confession', 'tribe', 'coins_adjustment'}
    needs_pool_context = bool(pool_scoped_fields.intersection(data)) or data.get('pool_id') is not None
    pool_id = None
    pv = None
    if needs_pool_context:
        pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
        if error:
            return error
        pv = PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=user_id).first()
        if not pv:
            return jsonify({'error': 'Пользователь не назначен на активный бассейн'}), 409

    if 'role' in data:
        new_role = data.get('role')
        if new_role not in ('volunteer', 'tribe_master'):
            return jsonify({'error': 'На этой странице можно выбрать только волонтёра или трайб-мастера'}), 400
        if user.role in ROLES_WITH_PASSWORD:
            return jsonify({'error': 'Тимлида или админа нельзя сделать трайб-мастером здесь'}), 403
        old = pv.pool_role or 'volunteer'
        if old != new_role:
            changes['role'] = {'from': old, 'to': new_role}
        pv.pool_role = new_role

    if 'name' in data:
        new_name = (data.get('name') or '').strip() or user.nick
        old = user.name or user.nick
        if old != new_name:
            changes['name'] = {'from': old, 'to': new_name}
        user.name = new_name

    if 'nick' in data:
        new_nick = (data.get('nick') or '').strip()
        if not new_nick:
            return jsonify({'error': 'Укажите ник'}), 400
        existing = User.query.filter(
            db.func.lower(User.nick) == new_nick.lower(),
            User.id != user.id,
        ).first()
        if existing:
            return jsonify({'error': 'Такой ник уже есть'}), 409
        if user.nick != new_nick:
            changes['nick'] = {'from': user.nick, 'to': new_nick}
        user.nick = new_nick

    if 'telegram' in data:
        new_telegram = _normalize_telegram_value(data.get('telegram')) or ''
        if telegram_username_conflict(new_telegram, user.id):
            return jsonify({'error': 'Этот Telegram username уже используется'}), 409
        old = user.telegram or ''
        if old != new_telegram:
            changes['telegram'] = {'from': old or None, 'to': new_telegram or None}
        user.telegram = new_telegram or None

    if 'has_confession' in data:
        new_val = bool(data.get('has_confession'))
        if pv:
            old = bool(pv.has_confession)
            if old != new_val:
                changes['has_confession'] = {'from': old, 'to': new_val}
            pv.has_confession = new_val
            if not new_val:
                RewardEvent.query.filter_by(user_id=user.id, pool_id=pool_id, event_type='confession').delete()
        else:
            old = bool(user.has_confession)
            if old != new_val:
                changes['has_confession'] = {'from': old, 'to': new_val}
            user.has_confession = new_val
            if not new_val:
                RewardEvent.query.filter_by(user_id=user.id, event_type='confession').delete()

    if 'tribe' in data:
        new_tribe = normalize_tribe(data.get('tribe'))
        if pv:
            if pv.tribe != new_tribe:
                changes['tribe'] = {'from': pv.tribe, 'to': new_tribe}
            pv.tribe = new_tribe
        else:
            if user.tribe != new_tribe:
                changes['tribe'] = {'from': user.tribe, 'to': new_tribe}
            user.tribe = new_tribe

    if 'coins_adjustment' in data:
        try:
            new_adjustment = int(data.get('coins_adjustment') or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'coins_adjustment должен быть числом'}), 400
        if pv:
            old = pv.coins_adjustment or 0
            if old != new_adjustment:
                changes['coins_adjustment'] = {'from': old, 'to': new_adjustment}
            pv.coins_adjustment = new_adjustment
        else:
            old = user.coins_adjustment or 0
            if old != new_adjustment:
                changes['coins_adjustment'] = {'from': old, 'to': new_adjustment}
            user.coins_adjustment = new_adjustment

    if pv:
        conflict = validate_tribe_master_assignment(
            pool_id,
            pv.pool_role or 'volunteer',
            pv.tribe,
            exclude_user_id=user.id,
        )
        if conflict:
            db.session.rollback()
            return jsonify({'error': conflict}), 409

    if changes:
        log_action(
            'update',
            'volunteer',
            user.id,
            f'Обновлён профиль @{user.nick}',
            {'target_nick': user.nick, 'changes': changes},
        )
    db.session.commit()
    return jsonify(user.to_dict())


# ==================== Групповые проверки ====================


def _group_review_to_dict(review):
    reviewer = db.session.get(User, review.reviewer_id)
    creator = db.session.get(User, review.created_by) if review.created_by else None
    return {
        'id': review.id,
        'pool_id': review.pool_id,
        'date': review.review_date.isoformat(),
        'time_start': review.time_start,
        'quantity': review.quantity or 1,
        'reviewer': _user_public_dict(reviewer) if reviewer else None,
        'created_by': _user_public_dict(creator) if creator else None,
    }


@app.route('/api/group-reviews', methods=['GET'])
@require_auth
def list_group_reviews():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    reviews = (
        GroupReview.query
        .filter_by(pool_id=pool_id)
        .order_by(GroupReview.review_date.desc(), GroupReview.time_start)
        .all()
    )
    return jsonify([_group_review_to_dict(review) for review in reviews])


@app.route('/api/group-reviews', methods=['POST'])
@require_role('team_lead', 'admin')
def create_group_review():
    data = request.json or {}
    try:
        review_date = datetime.fromisoformat(data['date']).date()
    except (KeyError, ValueError):
        return jsonify({'error': 'Нужна дата проверки'}), 400
    reviewer_id = data.get('reviewer_id')
    reviewer = get_model_or_404(User, reviewer_id)
    try:
        quantity = max(1, int(data.get('quantity') or 1))
    except (TypeError, ValueError):
        return jsonify({'error': 'Количество проверок должно быть числом'}), 400
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    if not PoolVolunteer.query.filter_by(pool_id=pool_id, user_id=reviewer.id).first():
        return jsonify({'error': 'Проверяющий не назначен на активный бассейн'}), 409
    review = GroupReview(
        pool_id=pool_id,
        review_date=review_date,
        time_start=(data.get('time_start') or '10:00').strip(),
        flow='',
        quantity=quantity,
        reviewer_id=reviewer.id,
        created_by=g.user.id,
    )
    db.session.add(review)
    db.session.commit()
    return jsonify(_group_review_to_dict(review)), 201


@app.route('/api/group-reviews/<int:review_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_group_review(review_id):
    review = get_model_or_404(GroupReview, review_id)
    error = _active_entity_error(review.pool_id)
    if error:
        return error
    db.session.delete(review)
    db.session.commit()
    return jsonify({'message': 'Групповая проверка удалена'})


# ==================== Ученики ====================


def active_pool_id():
    pool = Pool.query.filter_by(active=True).order_by(Pool.created_at.desc()).first()
    return pool.id if pool else None


def _active_pool_id_for_request(requested_pool_id=None):
    """Resolve all operational requests to the single system-wide active pool."""
    pool_id = active_pool_id()
    if not pool_id:
        return None, (jsonify({'error': 'Нет активного бассейна'}), 400)
    if requested_pool_id is not None and int(requested_pool_id) != pool_id:
        return None, (jsonify({'error': 'Данные можно изменять только в активном бассейне'}), 409)
    request_user = getattr(g, 'user', None)
    if request_user and not _can_access_pool_id(request_user, pool_id):
        return None, (jsonify({'error': 'У тебя нет доступа к активному бассейну'}), 403)
    return pool_id, None


def _active_entity_error(entity_pool_id):
    pool_id = active_pool_id()
    if not pool_id:
        return jsonify({'error': 'Нет активного бассейна'}), 400
    if entity_pool_id != pool_id:
        return jsonify({'error': 'Эта запись относится не к активному бассейну'}), 409
    return None


def _status_counts(pool_id):
    penalties = StudentPenalty.query.filter_by(pool_id=pool_id).all() if pool_id else []
    students_with_penalties = {p.student_name for p in penalties}
    return {
        'students_with_penalties': len(students_with_penalties),
        'pending': len([p for p in penalties if p.workoff_status == 'pending']),
        'overdue': len([p for p in penalties if p.workoff_status == 'overdue']),
        'in_workoff': len([p for p in penalties if p.workoff_status == 'in_workoff']),
        'awaiting_unlock': len([p for p in penalties if p.workoff_status == 'awaiting_unlock']),
    }


def _user_public_dict(user, pool_id=None):
    target_pool_id = pool_id or active_pool_id()
    return {
        'id': user.id,
        'nick': user.nick,
        'name': user.name or user.nick,
        'telegram': user.telegram,
        'role': _effective_access_role(user, target_pool_id),
        'tribe': _effective_access_tribe(user, target_pool_id),
        'avatar_url': _avatar_url_for_user(user),
    }


def _block_with_people(block):
    signups = (
        db.session.query(Signup, User)
        .join(User, User.id == Signup.user_id)
        .filter(Signup.block_id == block.id)
        .order_by(Signup.created_at)
        .all()
    )
    return _block_with_people_from_rows(block, signups)


def _block_with_people_from_rows(block, signups):
    return {
        'id': block.id,
        'date': block.date.isoformat(),
        'time_start': block.time_start,
        'time_end': block.time_end,
        'label': block.label or '',
        'capacity': block.capacity,
        'count': len(signups),
        'volunteers': [_user_public_dict(user) for _, user in signups],
    }


def _signups_index_for_blocks(block_ids):
    if not block_ids:
        return {}
    rows = (
        db.session.query(Signup, User)
        .join(User, User.id == Signup.user_id)
        .filter(Signup.block_id.in_(block_ids))
        .order_by(Signup.created_at)
        .all()
    )
    grouped = defaultdict(list)
    for signup, user in rows:
        grouped[signup.block_id].append((signup, user))
    return grouped


def _tomorrow_blocks(pool_id):
    tomorrow = _moscow_today() + timedelta(days=1)
    blocks = ShiftBlock.query.filter_by(pool_id=pool_id, date=tomorrow).order_by(ShiftBlock.time_start).all()
    signups_by_block = _signups_index_for_blocks([block.id for block in blocks])
    return [_block_with_people_from_rows(block, signups_by_block.get(block.id, [])) for block in blocks]


def _future_my_shifts(user_id, pool_id, limit=5):
    today = _moscow_today()
    rows = (
        db.session.query(ShiftBlock)
        .join(Signup, Signup.block_id == ShiftBlock.id)
        .filter(Signup.user_id == user_id, ShiftBlock.pool_id == pool_id, ShiftBlock.date >= today)
        .order_by(ShiftBlock.date, ShiftBlock.time_start)
        .limit(limit)
        .all()
    )
    signups_by_block = _signups_index_for_blocks([block.id for block in rows])
    return [_block_with_people_from_rows(block, signups_by_block.get(block.id, [])) for block in rows]


def _tribes_for_pool(pool_id):
    defined = [
        normalize_tribe(tribe.name)
        for tribe in Tribe.query.filter_by(pool_id=pool_id).order_by(Tribe.name).all()
        if normalize_tribe(tribe.name)
    ]
    rows = (
        db.session.query(Student.tribe)
        .filter(Student.pool_id == pool_id, Student.tribe.isnot(None))
        .distinct()
        .order_by(Student.tribe)
        .all()
    )
    found = [normalize_tribe(row[0]) for row in rows if row[0]]
    return list(dict.fromkeys([*defined, *found]))


def _resolve_user_tribe(user, pool_id):
    effective_tribe = _effective_access_tribe(user, pool_id)
    if effective_tribe:
        return effective_tribe
    tribes = _tribes_for_pool(pool_id)
    return tribes[0] if tribes else TRIBES[0]


def _tribe_metrics(pool_id, tribe):
    students = Student.query.filter_by(pool_id=pool_id, tribe=tribe).all() if pool_id and tribe else []
    student_ids = [s.id for s in students]
    all_events = (
        StudentEvent.query
        .filter(StudentEvent.student_id.in_(student_ids))
        .all()
        if student_ids else []
    )
    confirmed_events = [event for event in all_events if event.status == 'confirmed']
    entertainment = len([e for e in confirmed_events if e.event_type == 'entertainment'])
    education = len([e for e in confirmed_events if e.event_type == 'education'])
    points_total = sum(e.points or STUDENT_EVENT_POINTS.get(e.event_type, 0) for e in confirmed_events)
    by_student = {}
    for student in students:
        by_student[student.id] = {
            'id': student.id,
            'nick': student.nick,
            'name': student.name,
            'tribe': student.tribe,
            'events_total': 0,
            'entertainment_events': 0,
            'education_events': 0,
        }
    for event in confirmed_events:
        item = by_student.get(event.student_id)
        if not item:
            continue
        item['events_total'] += 1
        item['points'] = item.get('points', 0) + (event.points or STUDENT_EVENT_POINTS.get(event.event_type, 0))
        if event.event_type == 'entertainment':
            item['entertainment_events'] += 1
        if event.event_type == 'education':
            item['education_events'] += 1
    return {
        'tribe': tribe,
        'students_count': len(students),
        'events_total': len(confirmed_events),
        'events_created_total': len(all_events),
        'entertainment_events': entertainment,
        'education_events': education,
        'points_total': points_total,
        'top_students': sorted(by_student.values(), key=lambda s: (-(s.get('points') or 0), -s['events_total'], s['nick']))[:10],
    }


def _tribe_rankings(pool_id):
    rankings = []
    for tribe in _tribes_for_pool(pool_id):
        metrics = _tribe_metrics(pool_id, tribe)
        rankings.append({
            'tribe': tribe,
            'events_total': metrics['events_total'],
            'events_created_total': metrics['events_created_total'],
            'entertainment_events': metrics['entertainment_events'],
            'education_events': metrics['education_events'],
            'points_total': metrics['points_total'],
        })
    rankings.sort(key=lambda row: (-row['events_total'], -row['points_total'], row['tribe']))
    for index, row in enumerate(rankings, start=1):
        row['rank'] = index
    return rankings


def _student_penalty_status(penalties):
    if any(p.workoff_status == 'awaiting_unlock' for p in penalties):
        return 'awaiting_unlock'
    if any(p.workoff_status == 'in_workoff' for p in penalties):
        return 'workoff'
    if any(p.workoff_status in ('pending', 'overdue') for p in penalties):
        return 'received'
    return 'clean'


def _tribe_event_to_dict(event):
    creator = db.session.get(User, event.created_by) if event.created_by else None
    return {
        'id': event.id,
        'pool_id': event.pool_id,
        'tribe': event.tribe,
        'title': event.title,
        'date': event.event_date.isoformat(),
        'time_start': event.time_start or '',
        'location': event.location or '',
        'comment': event.comment or '',
        'created_by': _user_public_dict(creator) if creator else None,
    }


@app.route('/api/dashboard', methods=['GET'])
@require_auth
def dashboard_summary():
    pool = Pool.query.filter_by(active=True).order_by(Pool.created_at.desc()).first()
    pool_id = pool.id if pool else None
    if pool_id and not _can_access_pool_id(g.user, pool_id):
        return jsonify({'error': 'У тебя нет доступа к активному бассейну'}), 403
    tomorrow = _moscow_today() + timedelta(days=1)
    counts = _status_counts(pool_id)
    tomorrow_tribe_events = (
        TribeEvent.query.filter_by(pool_id=pool_id, event_date=tomorrow).order_by(TribeEvent.time_start).all()
        if pool_id else []
    )
    data = {
        'pool': pool.to_dict() if pool else None,
        'role': g.current_role,
        'tomorrow': tomorrow.isoformat(),
        'tomorrow_blocks': _tomorrow_blocks(pool_id) if pool_id else [],
        'penalties': counts,
        'tomorrow_tribe_events': [_tribe_event_to_dict(event) for event in tomorrow_tribe_events],
        'my_shifts': _future_my_shifts(g.user.id, pool_id) if pool_id else [],
        'telegram': _telegram_link_status(g.user),
        'pool_responsibles': _pool_responsibles(pool_id),
        'dashboard_notes': [
            _dashboard_note_to_dict(note)
            for note in DashboardNote.query.filter_by(pool_id=pool_id, is_active=True)
            .order_by(DashboardNote.is_pinned.desc(), DashboardNote.updated_at.desc(), DashboardNote.id.desc())
            .limit(10)
            .all()
        ],
    }
    if g.current_role == 'tribe_master' and pool_id:
        tribe = _resolve_user_tribe(g.user, pool_id)
        rankings = _tribe_rankings(pool_id)
        own_rank = next((row['rank'] for row in rankings if row['tribe'] == tribe), None)
        next_events = (
            TribeEvent.query
            .filter(TribeEvent.pool_id == pool_id, TribeEvent.tribe == tribe, TribeEvent.event_date >= _moscow_today())
            .order_by(TribeEvent.event_date, TribeEvent.time_start)
            .limit(5)
            .all()
        )
        data['tribe'] = {
            **_tribe_metrics(pool_id, tribe),
            'rank': own_rank,
            'rankings': rankings,
            'next_events': [_tribe_event_to_dict(event) for event in next_events],
            'available_tribes': _tribes_for_pool(pool_id),
        }
    return jsonify(data)


@app.route('/api/dashboard-notes', methods=['GET'])
@require_auth
def list_dashboard_notes():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    notes = (
        DashboardNote.query.filter_by(pool_id=pool_id, is_active=True)
        .order_by(DashboardNote.is_pinned.desc(), DashboardNote.updated_at.desc(), DashboardNote.id.desc())
        .all()
    )
    return jsonify([_dashboard_note_to_dict(note) for note in notes])


@app.route('/api/telegram/status', methods=['GET'])
@require_auth
def telegram_status():
    return jsonify(_telegram_link_status(g.user))


@app.route('/api/telegram/bot-avatar', methods=['GET'])
def telegram_bot_avatar():
    if not _telegram_is_configured():
        return jsonify({'error': 'Telegram бот не настроен'}), 503
    try:
        payload, content_type = _telegram_bot_avatar_payload()
    except Exception:
        return jsonify({'error': 'Не удалось загрузить аватар бота'}), 502
    if not payload:
        return jsonify({'error': 'Аватар бота не найден'}), 404
    return send_file(
        BytesIO(payload),
        mimetype=content_type,
        max_age=3600,
        conditional=True,
        download_name='telegram-bot-avatar.jpg',
    )


@app.route('/api/telegram/unlinked-users', methods=['GET'])
@require_role('team_lead', 'admin')
def telegram_unlinked_users():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    return jsonify(_unlinked_users_for_pool(pool_id))


@app.route('/api/telegram/webhook', methods=['POST'])
def telegram_webhook():
    if not _telegram_is_configured():
        return jsonify({'error': 'Telegram бот не настроен'}), 503
    if _is_production_runtime() and not TELEGRAM_WEBHOOK_SECRET:
        return jsonify({'error': 'TELEGRAM_WEBHOOK_SECRET обязателен в production'}), 503
    if TELEGRAM_WEBHOOK_SECRET:
        incoming_secret = request.headers.get('X-Telegram-Bot-Api-Secret-Token', '')
        if incoming_secret != TELEGRAM_WEBHOOK_SECRET:
            return jsonify({'error': 'Неверный webhook secret'}), 403
    payload = request.get_json(silent=True) or {}
    message = payload.get('message')
    callback = payload.get('callback_query')
    if not message and not callback:
        return jsonify({'ok': True, 'ignored': True})
    try:
        result = telegram_handle_message(message) if message else telegram_handle_callback(callback)
        db.session.commit()
        return jsonify({'ok': True, 'result': result})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/telegram/webhook/register', methods=['POST'])
@require_role('team_lead', 'admin')
def register_telegram_webhook():
    if not _telegram_is_configured():
        return jsonify({'error': 'Telegram бот не настроен'}), 503
    data = request.json or {}
    webhook_url = (data.get('webhook_url') or '').strip()
    if not webhook_url:
        return jsonify({'error': 'Передайте webhook_url'}), 400
    payload = {'url': webhook_url}
    secret_token = (data.get('secret_token') or TELEGRAM_WEBHOOK_SECRET or '').strip()
    if secret_token:
        payload['secret_token'] = secret_token
    result = telegram_api('setWebhook', payload)
    commands_result = telegram_sync_commands()
    return jsonify({
        'ok': True,
        'result': result,
        'commands_result': commands_result,
        'webhook_url': webhook_url,
    })


@app.route('/api/telegram/webhook/info', methods=['GET'])
@require_role('team_lead', 'admin')
def telegram_webhook_info():
    if not _telegram_is_configured():
        return jsonify({'error': 'Telegram бот не настроен'}), 503
    result = telegram_get('getWebhookInfo')
    return jsonify({'ok': True, 'result': result})


@app.route('/api/telegram/commands/sync', methods=['POST'])
@require_role('team_lead', 'admin')
def sync_telegram_commands():
    if not _telegram_is_configured():
        return jsonify({'error': 'Telegram бот не настроен'}), 503
    result = telegram_sync_commands()
    log_action('sync', 'telegram_commands', None, 'Синхронизированы команды меню Telegram-бота')
    db.session.commit()
    return jsonify({'ok': True, 'result': result})


@app.route('/api/notifications/dispatch', methods=['GET', 'POST'])
def dispatch_notifications():
    if not verify_internal_api_secret():
        return jsonify({'error': 'Недостаточно прав для dispatch'}), 403
    data = request.get_json(silent=True) or {}
    raw_event_ids = data.get('event_ids')
    event_ids = None
    if raw_event_ids is not None:
        if not isinstance(raw_event_ids, list):
            return jsonify({'error': 'event_ids должен быть списком'}), 400
        try:
            parsed_event_ids = [int(item) for item in raw_event_ids]
        except (TypeError, ValueError):
            return jsonify({'error': 'Некорректный event_ids'}), 400
        event_ids = list(dict.fromkeys(item for item in parsed_event_ids if item > 0))[:100]
    limit = request.args.get('limit', default=20, type=int) or 20
    if event_ids:
        limit = max(limit, len(event_ids))
    try:
        result = process_pending_notifications(
            limit=min(limit, 100),
            event_ids=event_ids,
            schedule_events=event_ids is None,
        )
        db.session.commit()
        return jsonify({'ok': True, **result})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/notifications/dispatch/manual', methods=['POST'])
@require_role('team_lead', 'admin')
def dispatch_notifications_manual():
    limit = request.args.get('limit', default=50, type=int) or 50
    try:
        result = process_pending_notifications(limit=min(limit, 100))
        log_action(
            'dispatch',
            'notification_queue',
            None,
            'Ручной запуск очереди Telegram уведомлений',
            result,
        )
        db.session.commit()
        return jsonify({'ok': True, **result})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 500


def _broadcast_recipient_query(pool_id, filters):
    query = User.query.filter_by(active=True)
    if pool_id:
        allowed_ids = {
            row.user_id for row in PoolVolunteer.query.filter_by(pool_id=pool_id).with_entities(PoolVolunteer.user_id).all()
        }
        if allowed_ids:
            query = query.filter(User.id.in_(allowed_ids))
        else:
            query = query.filter(db.text('0=1'))
    role = (filters or {}).get('role')
    duty_window = ((filters or {}).get('duty_window') or '').strip()
    usernames = [normalize_tg_username(item) for item in ((filters or {}).get('usernames') or []) if item and item.strip()]
    usernames = [item for item in usernames if item]
    if role:
        if role in ('volunteer', 'tribe_master') and pool_id:
            role_user_ids = {
                row.user_id
                for row in PoolVolunteer.query.filter_by(pool_id=pool_id, pool_role=role)
                .with_entities(PoolVolunteer.user_id).all()
            }
            query = query.filter(User.id.in_(role_user_ids)) if role_user_ids else query.filter(db.text('0=1'))
        else:
            query = query.filter_by(role=role)
    if duty_window in {'today', 'tomorrow'} and pool_id:
        target_date = _moscow_now().date() + (timedelta(days=1) if duty_window == 'tomorrow' else timedelta())
        signed_user_ids = {
            row.user_id
            for row in (
                db.session.query(Signup.user_id)
                .join(ShiftBlock, ShiftBlock.id == Signup.block_id)
                .filter(
                    ShiftBlock.pool_id == pool_id,
                    ShiftBlock.date == target_date,
                )
                .distinct()
                .all()
            )
        }
        if signed_user_ids:
            query = query.filter(User.id.in_(signed_user_ids))
        else:
            query = query.filter(db.text('0=1'))
    if usernames:
        telegram_variants = usernames + [f'@{item}' for item in usernames]
        query = query.filter(db.or_(
            db.func.lower(User.telegram).in_(telegram_variants),
            db.func.lower(User.nick).in_(usernames),
        ))
    return query.order_by(User.nick)


@app.route('/api/notifications/overview', methods=['GET'])
@require_role('team_lead', 'admin')
def notifications_overview():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error and active_pool_id() is None:
        settings = get_telegram_settings()
        return jsonify({
            'telegram_settings': settings,
            'test_mode': settings['test_mode'],
            'notes': [],
            'broadcasts': [],
            'recent_deliveries': [],
            'linked_users': [],
            'unlinked_users': [],
            'linked_users_count': 0,
        })
    if error:
        return error
    linked_users = _linked_users_for_pool(pool_id)
    recent_notes = (
        DashboardNote.query.filter_by(pool_id=pool_id)
        .order_by(DashboardNote.is_pinned.desc(), DashboardNote.updated_at.desc(), DashboardNote.id.desc())
        .limit(20)
        .all()
    )
    recent_broadcasts = (
        Broadcast.query.filter_by(pool_id=pool_id)
        .order_by(Broadcast.updated_at.desc(), Broadcast.id.desc())
        .limit(20)
        .all()
    )
    recent_deliveries = (
        NotificationDelivery.query
        .join(NotificationEvent, NotificationEvent.id == NotificationDelivery.notification_id)
        .filter(NotificationEvent.pool_id == pool_id)
        .order_by(NotificationDelivery.created_at.desc(), NotificationDelivery.id.desc())
        .limit(50)
        .all()
    )
    return jsonify({
        'telegram_settings': get_telegram_settings(),
        'test_mode': get_telegram_settings()['test_mode'],
        'notes': [_dashboard_note_to_dict(note) for note in recent_notes],
        'broadcasts': [_broadcast_to_dict(item) for item in recent_broadcasts],
        'recent_deliveries': [{
            'id': delivery.id,
            'notification_id': delivery.notification_id,
            'user_id': delivery.user_id,
            'telegram_chat_id': delivery.telegram_chat_id,
            'delivery_status': delivery.delivery_status,
            'error': delivery.error or '',
            'message_id': delivery.message_id or '',
            'created_at': delivery.created_at.isoformat() if delivery.created_at else None,
        } for delivery in recent_deliveries],
        'linked_users': linked_users,
        'unlinked_users': _unlinked_users_for_pool(pool_id),
        'linked_users_count': len(linked_users),
    })


@app.route('/api/notifications/settings', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_notifications_settings():
    data = request.json or {}
    telegram_payload = data.get('telegram') if isinstance(data.get('telegram'), dict) else data
    try:
        settings, changes = update_telegram_settings(telegram_payload)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    log_action(
        'update',
        'telegram_settings',
        None,
        'Обновлены настройки Telegram уведомлений',
        {'changes': changes, 'settings': settings},
    )
    db.session.commit()
    return jsonify({'telegram_settings': settings, 'changes': changes})


@app.route('/api/notifications/history', methods=['GET'])
@require_role('team_lead', 'admin')
def notifications_history():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    limit = min(request.args.get('limit', default=80, type=int) or 80, 200)
    events = (
        NotificationEvent.query
        .filter(NotificationEvent.pool_id == pool_id)
        .order_by(NotificationEvent.created_at.desc(), NotificationEvent.id.desc())
        .limit(limit)
        .all()
    )
    result = []
    for event in events:
        user = db.session.get(User, event.recipient_user_id) if event.recipient_user_id else None
        deliveries = NotificationDelivery.query.filter_by(notification_id=event.id).order_by(NotificationDelivery.created_at.desc()).all()
        result.append({
            'id': event.id,
            'type': event.type,
            'priority': event.priority,
            'status': event.status,
            'scheduled_for': event.scheduled_for.isoformat() if event.scheduled_for else None,
            'sent_at': event.sent_at.isoformat() if event.sent_at else None,
            'cancelled_at': event.cancelled_at.isoformat() if event.cancelled_at else None,
            'recipient': {
                'id': user.id,
                'nick': user.nick,
                'name': user.name or user.nick,
            } if user else None,
            'payload': json.loads(event.payload or '{}'),
            'source_entity': event.source_entity,
            'source_entity_id': event.source_entity_id,
            'created_at': event.created_at.isoformat() if event.created_at else None,
            'deliveries': [{
                'id': delivery.id,
                'status': delivery.delivery_status,
                'error': delivery.error or '',
                'message_id': delivery.message_id or '',
                'created_at': delivery.created_at.isoformat() if delivery.created_at else None,
            } for delivery in deliveries],
        })
    return jsonify(result)


@app.route('/api/notifications/broadcasts', methods=['POST'])
@require_role('team_lead', 'admin')
def create_broadcast():
    data = request.json or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'Введите текст рассылки'}), 400
    priority = (data.get('priority') or 'normal').strip() or 'normal'
    filters = data.get('filters') or {}
    pool_id = active_pool_id()
    if not pool_id:
        return jsonify({'error': 'Нет активного бассейна'}), 400
    broadcast = Broadcast(
        author_id=g.user.id,
        pool_id=pool_id,
        text=text,
        filters=json.dumps(filters, ensure_ascii=False),
        priority=priority,
        status='queued',
        is_anonymous=bool(data.get('is_anonymous')),
    )
    db.session.add(broadcast)
    db.session.flush()
    recipients = [
        user for user in _broadcast_recipient_query(pool_id, filters).all()
        if _pool_notifications_enabled_for_user(user.id, pool_id)
    ]
    notification_events = []
    for user in recipients:
        payload = {
            'text': text,
            'filters': filters,
            'broadcast_id': broadcast.id,
        }
        event = NotificationEvent(
            type='manual_broadcast',
            priority=priority,
            status='queued',
            recipient_user_id=user.id,
            pool_id=pool_id,
            payload=json.dumps(payload, ensure_ascii=False),
            source_entity='broadcast',
            source_entity_id=broadcast.id,
            created_by=g.user.id,
            dedupe_key=f'broadcast:{broadcast.id}:user:{user.id}',
        )
        db.session.add(event)
        db.session.flush()
        notification_events.append(event)
        account = TelegramAccount.query.filter_by(user_id=user.id, is_linked=True).first()
        db.session.add(NotificationDelivery(
            notification_id=event.id,
            user_id=user.id,
            telegram_chat_id=account.telegram_chat_id if account else None,
            delivery_status='pending' if account else 'skipped',
            error='' if account else 'Telegram не привязан',
        ))
    log_action(
        'create',
        'broadcast',
        broadcast.id,
        'Создана рассылка в разделе уведомлений',
        {'filters': filters, 'priority': priority, 'text': text, 'is_anonymous': bool(broadcast.is_anonymous)},
    )
    db.session.commit()
    queue_notification_dispatch(notification_events)
    return jsonify(_broadcast_to_dict(broadcast)), 201


@app.route('/api/notifications/broadcasts/<int:broadcast_id>', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_broadcast(broadcast_id):
    broadcast = get_model_or_404(Broadcast, broadcast_id)
    error = _active_entity_error(broadcast.pool_id)
    if error:
        return error
    data = request.json or {}
    changes = {}
    if 'text' in data:
        text = (data.get('text') or '').strip()
        if not text:
            return jsonify({'error': 'Введите текст рассылки'}), 400
        changes['text'] = {'from': broadcast.text, 'to': text}
        broadcast.text = text
    if 'priority' in data:
        priority = (data.get('priority') or 'normal').strip() or 'normal'
        changes['priority'] = {'from': broadcast.priority, 'to': priority}
        broadcast.priority = priority
    if 'status' in data:
        status = (data.get('status') or '').strip()
        if status:
            changes['status'] = {'from': broadcast.status, 'to': status}
            broadcast.status = status
    if 'filters' in data:
        new_filters = data.get('filters') or {}
        changes['filters'] = {'from': json.loads(broadcast.filters or '{}'), 'to': new_filters}
        broadcast.filters = json.dumps(new_filters, ensure_ascii=False)
    if 'is_anonymous' in data:
        new_value = bool(data.get('is_anonymous'))
        changes['is_anonymous'] = {'from': bool(broadcast.is_anonymous), 'to': new_value}
        broadcast.is_anonymous = new_value
    queued_events = NotificationEvent.query.filter_by(
        source_entity='broadcast',
        source_entity_id=broadcast.id,
    ).filter(NotificationEvent.status.in_(['queued', 'pending'])).all()
    for event in queued_events:
        payload = json.loads(event.payload or '{}')
        payload['text'] = broadcast.text
        payload['filters'] = json.loads(broadcast.filters or '{}')
        event.payload = json.dumps(payload, ensure_ascii=False)
        event.priority = broadcast.priority
    log_action('update', 'broadcast', broadcast.id, 'Обновлена рассылка', {'changes': changes})
    db.session.commit()
    return jsonify(_broadcast_to_dict(broadcast))


@app.route('/api/notifications/broadcasts/<int:broadcast_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_broadcast(broadcast_id):
    broadcast = get_model_or_404(Broadcast, broadcast_id)
    error = _active_entity_error(broadcast.pool_id)
    if error:
        return error
    event_ids = [
        row.id for row in NotificationEvent.query.with_entities(NotificationEvent.id)
        .filter_by(source_entity='broadcast', source_entity_id=broadcast.id)
        .all()
    ]
    if event_ids:
        NotificationDelivery.query.filter(NotificationDelivery.notification_id.in_(event_ids)).delete(synchronize_session=False)
        NotificationEvent.query.filter(NotificationEvent.id.in_(event_ids)).delete(synchronize_session=False)
    db.session.delete(broadcast)
    log_action('delete', 'broadcast', broadcast_id, 'Удалена рассылка')
    db.session.commit()
    return jsonify({'message': 'Рассылка удалена'})


@app.route('/api/notifications/notes', methods=['POST'])
@require_role('team_lead', 'admin')
def create_notification_note():
    data = request.json or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'Введите текст заметки'}), 400
    pool_id = active_pool_id()
    if not pool_id:
        return jsonify({'error': 'Нет активного бассейна'}), 400
    normalized_text = _normalized_dedupe_text(text)
    _acquire_dedupe_lock('dashboard-note', pool_id, normalized_text)
    recent_notes = DashboardNote.query.filter(
        DashboardNote.pool_id == pool_id,
        DashboardNote.created_at >= _utcnow() - ANNOUNCEMENT_DUPLICATE_WINDOW,
    ).order_by(DashboardNote.created_at.desc()).all()
    duplicate_note = next(
        (note for note in recent_notes if _normalized_dedupe_text(note.text) == normalized_text),
        None,
    )
    if duplicate_note:
        response = _dashboard_note_to_dict(duplicate_note)
        response.update({
            'duplicate': True,
            'telegram_notification_count': 0,
            'message': 'Такое объявление уже опубликовано менее минуты назад',
        })
        db.session.commit()
        return jsonify(response), 200
    note = DashboardNote(
        author_id=g.user.id,
        pool_id=pool_id,
        text=text,
        is_pinned=bool(data.get('is_pinned')),
        is_highlighted=bool(data.get('is_highlighted')),
        is_active=True if data.get('is_active') is None else bool(data.get('is_active')),
        is_anonymous=bool(data.get('is_anonymous')),
    )
    db.session.add(note)
    db.session.flush()
    notification_events = []
    if data.get('notify_telegram'):
        dashboard_url = f'{_frontend_base_url()}/' if _frontend_base_url() else '/'
        notification_text = f'На дашборде появилось новое объявление:\n\n{text}'
        recipients = _broadcast_recipient_query(pool_id, {}).all()
        for user in recipients:
            event = _queue_notification(
                user,
                'dashboard_note_created',
                notification_text,
                f'dashboard-note:{note.id}:user:{user.id}',
                pool_id=pool_id,
                source_entity='dashboard_note',
                source_entity_id=note.id,
                created_by=g.user.id,
                action_buttons=[{'text': 'Открыть дашборд', 'url': dashboard_url}],
            )
            if event:
                notification_events.append(event)
    log_action(
        'create',
        'dashboard_note',
        None,
        'Создана заметка для дашборда',
        {
            'text': text,
            'is_pinned': note.is_pinned,
            'is_highlighted': note.is_highlighted,
            'is_anonymous': note.is_anonymous,
            'notify_telegram': bool(data.get('notify_telegram')),
            'telegram_notification_count': len(notification_events),
        },
    )
    db.session.commit()
    queue_notification_dispatch(notification_events)
    response = _dashboard_note_to_dict(note)
    response['duplicate'] = False
    response['telegram_notification_count'] = len(notification_events)
    return jsonify(response), 201


@app.route('/api/notifications/notes/<int:note_id>', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_notification_note(note_id):
    note = get_model_or_404(DashboardNote, note_id)
    error = _active_entity_error(note.pool_id)
    if error:
        return error
    data = request.json or {}
    changes = {}
    for field in ('text', 'is_pinned', 'is_highlighted', 'is_active', 'is_anonymous'):
        if field not in data:
            continue
        new_value = data.get(field)
        if field == 'text':
            new_value = (new_value or '').strip()
            if not new_value:
                return jsonify({'error': 'Введите текст заметки'}), 400
        old_value = getattr(note, field)
        if old_value != new_value:
            changes[field] = {'from': old_value, 'to': new_value}
            setattr(note, field, new_value)
    log_action('update', 'dashboard_note', note.id, 'Обновлена заметка для дашборда', {'changes': changes})
    db.session.commit()
    return jsonify(_dashboard_note_to_dict(note))


@app.route('/api/notifications/notes/<int:note_id>', methods=['DELETE'])
@require_role('team_lead', 'admin')
def delete_notification_note(note_id):
    note = get_model_or_404(DashboardNote, note_id)
    error = _active_entity_error(note.pool_id)
    if error:
        return error
    db.session.delete(note)
    log_action('delete', 'dashboard_note', note_id, 'Удалена заметка для дашборда')
    db.session.commit()
    return jsonify({'message': 'Заметка удалена'})


def _student_list_payload(pool_id):
    students = Student.query.filter_by(pool_id=pool_id).all()
    penalties_by_student = defaultdict(list)
    events_by_student = defaultdict(list)

    if students:
        student_ids = [student.id for student in students]
        student_nicks = [student.nick for student in students]
        penalties = StudentPenalty.query.filter(StudentPenalty.pool_id == pool_id).filter(
            db.or_(
                StudentPenalty.student_id.in_(student_ids),
                db.and_(StudentPenalty.student_id.is_(None), StudentPenalty.student_name.in_(student_nicks)),
            )
        ).all()
        nick_to_ids = defaultdict(list)
        for student in students:
            nick_to_ids[student.nick].append(student.id)
        for penalty in penalties:
            matched_ids = [penalty.student_id] if penalty.student_id else nick_to_ids.get(penalty.student_name, [])
            for student_id in matched_ids:
                penalties_by_student[student_id].append(penalty)

        all_events = StudentEvent.query.filter(StudentEvent.student_id.in_(student_ids)).all()
        for event in all_events:
            events_by_student[event.student_id].append(event)

    result = []
    for student in students:
        penalties = penalties_by_student.get(student.id, [])
        all_events = events_by_student.get(student.id, [])
        events = [e for e in all_events if e.status == 'confirmed']
        total_hours = sum(
            p.hours * p.multiplier
            for p in penalties
            if p.workoff_status not in ('done', 'awaiting_unlock', 'unlocked')
        )
        pending_count = len([p for p in penalties if p.workoff_status == 'pending'])
        overdue_count = len([p for p in penalties if p.workoff_status == 'overdue'])
        in_workoff_count = len([p for p in penalties if p.workoff_status == 'in_workoff'])
        entertainment_events = len([e for e in events if e.event_type == 'entertainment'])
        education_events = len([e for e in events if e.event_type == 'education'])
        event_points = sum(e.points or STUDENT_EVENT_POINTS.get(e.event_type, 0) for e in events)
        result.append({
            'id': student.id,
            'nick': student.nick,
            'name': student.name,
            'tribe': student.tribe,
            'violations_count': len(penalties),
            'total_penalty_hours': total_hours,
            'pending_penalties': pending_count,
            'overdue_penalties': overdue_count,
            'active_workoff_penalties': in_workoff_count,
            'awaiting_unlock_penalties': len([p for p in penalties if p.workoff_status == 'awaiting_unlock']),
            'penalty_status': _student_penalty_status(penalties),
            'in_workoff': total_hours > 0,
            'events_total': len(events),
            'entertainment_events': entertainment_events,
            'education_events': education_events,
            'event_points': event_points,
            'penalties': [{
                'id': p.id,
                'hours': p.hours * p.multiplier,
                'status': p.workoff_status,
                'volunteer': p.volunteer_name,
                'date': p.date_issued.isoformat(),
            } for p in penalties],
            'events': [{
                'id': e.id,
                'type': e.event_type,
                'date': e.event_date.isoformat() if e.event_date else None,
                'post_url': e.post_url or '',
                'proof_url': e.proof_url or '',
                'points': e.points or STUDENT_EVENT_POINTS.get(e.event_type, 0),
                'status': e.status or 'pending',
                'comment': e.comment or '',
            } for e in events],
        })
    return result


@app.route('/api/students', methods=['GET'])
@require_auth
def get_students():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    return jsonify(_student_list_payload(pool_id))


@app.route('/api/students', methods=['POST'])
@require_role('admin', 'team_lead')
def create_student():
    data = request.json or {}
    nick = (data.get('nick') or '').strip()
    if not nick:
        return jsonify({'error': 'Нужен ник'}), 400
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    existing = Student.query.filter(
        Student.pool_id == pool_id,
        db.func.lower(Student.nick) == nick.lower(),
    ).first()
    if existing:
        return jsonify({'error': 'Ученик с таким ником уже есть в активном бассейне'}), 409
    student = Student(
        nick=nick,
        name=nick,
        tribe=normalize_tribe(data.get('tribe')),
        pool_id=pool_id,
    )
    db.session.add(student)
    db.session.commit()
    return jsonify({'id': student.id, 'message': f'Ученик @{student.nick} добавлен'}), 201


def save_student_rows(rows, pool_id=None):
    pool_id = pool_id or active_pool_id()
    if not pool_id:
        return {
            'created': 0,
            'updated': 0,
            'skipped': [],
            'message': 'Нет активного бассейна',
            'error': 'Нет активного бассейна',
        }
    created = 0
    updated = 0
    skipped = []

    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            skipped.append({'row': index, 'reason': 'Некорректная строка'})
            continue

        nick = (row.get('nick') or '').strip()
        name = (row.get('name') or '').strip() or nick
        tribe = normalize_tribe(row.get('tribe'))
        if not nick:
            skipped.append({'row': index, 'reason': 'Нужен nick'})
            continue

        student = Student.query.filter(
            db.func.lower(Student.nick) == nick.lower(),
            Student.pool_id == pool_id,
        ).first()
        if student:
            student.name = name
            student.tribe = tribe
            updated += 1
        else:
            db.session.add(Student(nick=nick, name=name, tribe=tribe, pool_id=pool_id))
            created += 1

    db.session.commit()
    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'message': f'Импортировано: новых {created}, обновлено {updated}, пропущено {len(skipped)}',
    }


@app.route('/api/students/import', methods=['POST'])
@require_role('admin', 'team_lead')
def import_students():
    data = request.json or {}
    rows = data.get('students') or []
    if not isinstance(rows, list) or not rows:
        return jsonify({'error': 'Передайте students: [{nick, tribe?}]'}), 400

    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    return jsonify(save_student_rows(rows, pool_id=pool_id))


@app.route('/api/students/template', methods=['GET'])
@require_role('admin', 'team_lead')
def students_template():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    pool_id = active_pool_id()
    tribes = [t.name for t in Tribe.query.filter_by(pool_id=pool_id).order_by(Tribe.name).all()]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ученики'
    headers = ['ник школьный', 'трайб']
    ws.append(headers)
    ws.append(['ivanpetrov', tribes[0] if tribes else ''])

    header_fill = PatternFill('solid', fgColor='F3F6FA')
    header_font = Font(bold=True, color='1F2937')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18

    if tribes:
        options = ','.join(tribes)
        validation = DataValidation(type='list', formula1=f'"{options}"', allow_blank=True)
        validation.error = 'Выбери трайб из списка'
        validation.errorTitle = 'Некорректный трайб'
        ws.add_data_validation(validation)
        validation.add('B2:B1000')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='students_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/students/import-file', methods=['POST'])
@require_role('admin', 'team_lead')
def import_students_file():
    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({'error': 'Загрузите файл'}), 400
    try:
        rows = parse_xlsx_rows(uploaded)
        students = rows_to_dicts(rows, ['nick', 'tribe'])
    except Exception as e:
        return jsonify({'error': f'Не удалось прочитать .xlsx: {e}'}), 400
    pool_id, error = _active_pool_id_for_request(request.form.get('pool_id', type=int))
    if error:
        return error
    return jsonify(save_student_rows(students, pool_id=pool_id))


def _build_student_penalties_export(students):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Штрафы'
    headers = ['Ник', 'Количество штрафов']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='F3F6FA')
    header_font = Font(bold=True, color='1F2937')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for student in students:
        ws.append([student['nick'], student['violations_count'] or 0])

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 22
    return wb


def _build_student_events_export(students):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Мероприятия'
    headers = ['Ник', 'Всего мероприятий', 'Развлекательные', 'Обучающие']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='F3F6FA')
    header_font = Font(bold=True, color='1F2937')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for student in students:
        ws.append([
            student['nick'],
            student['events_total'] or 0,
            student['entertainment_events'] or 0,
            student['education_events'] or 0,
        ])

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    return wb


def _students_export_payload(pool_id):
    target_pool_id = pool_id or active_pool_id()
    if not target_pool_id:
        return []
    return _student_list_payload(target_pool_id)


@app.route('/api/students/export-penalties.xlsx', methods=['GET'])
@require_role('admin', 'team_lead')
def export_students_penalties():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    students = _students_export_payload(pool_id)
    wb = _build_student_penalties_export(students)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'students-penalties-{_utcnow().strftime("%Y-%m-%d-%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/students/export-events.xlsx', methods=['GET'])
@require_role('admin', 'team_lead')
def export_students_events():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    students = _students_export_payload(pool_id)
    wb = _build_student_events_export(students)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'students-events-{_utcnow().strftime("%Y-%m-%d-%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/students/<int:student_id>/events', methods=['POST'])
@require_role('tribe_master', 'team_lead', 'admin')
def create_student_event(student_id):
    student = get_model_or_404(Student, student_id)
    error = _active_entity_error(student.pool_id)
    if error:
        return error
    data = request.json or {}
    if (
        g.current_role == 'tribe_master'
        and normalize_tribe(student.tribe) != normalize_tribe(g.current_tribe)
    ):
        return jsonify({'error': 'Можно добавлять мероприятия только ученикам своего трайба'}), 403
    event_type = data.get('event_type')
    if event_type not in ('entertainment', 'education'):
        return jsonify({'error': 'Тип мероприятия должен быть entertainment или education'}), 400
    status = 'confirmed' if g.current_role == 'admin' else 'pending'

    event_date = None
    if data.get('event_date'):
        try:
            event_date = datetime.fromisoformat(data['event_date']).date()
        except ValueError:
            return jsonify({'error': 'Некорректная дата'}), 400

    event = StudentEvent(
        student_id=student.id,
        event_type=event_type,
        event_date=event_date,
        post_url=(data.get('post_url') or '').strip(),
        proof_url=(data.get('proof_url') or '').strip(),
        points=STUDENT_EVENT_POINTS[event_type] if status == 'confirmed' else 0,
        status=status,
        comment=(data.get('comment') or '').strip(),
        created_by=g.user.id,
    )
    db.session.add(event)
    db.session.commit()
    label = 'развлекательное' if event_type == 'entertainment' else 'обучающее'
    suffix = 'и подтверждено' if status == 'confirmed' else 'со статусом "ждет подтверждения АДМ"'
    return jsonify({'id': event.id, 'message': f'Добавлено {label} мероприятие для @{student.nick} {suffix}'}), 201


@app.route('/api/student-events/<int:event_id>', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_student_event(event_id):
    event = get_model_or_404(StudentEvent, event_id)
    student = db.session.get(Student, event.student_id)
    error = _active_entity_error(student.pool_id if student else None)
    if error:
        return error
    data = request.json or {}
    status = data.get('status')
    if status not in STUDENT_EVENT_STATUSES:
        return jsonify({'error': 'Статус должен быть pending, confirmed или rejected'}), 400
    event.status = status
    event.points = STUDENT_EVENT_POINTS.get(event.event_type, 0) if status == 'confirmed' else 0
    db.session.commit()
    return jsonify({'message': 'Статус мероприятия обновлен'})


@app.route('/api/student-events/<int:event_id>', methods=['DELETE'])
@require_role('tribe_master', 'team_lead', 'admin')
def delete_student_event(event_id):
    event = get_model_or_404(StudentEvent, event_id)
    student = db.session.get(Student, event.student_id)
    error = _active_entity_error(student.pool_id if student else None)
    if error:
        return error
    if g.current_role == 'tribe_master' and student and g.current_tribe and normalize_tribe(student.tribe) != normalize_tribe(g.current_tribe):
        return jsonify({'error': 'Можно удалять мероприятия только своего трайба'}), 403
    db.session.delete(event)
    db.session.commit()
    return jsonify({'message': 'Мероприятие ученика удалено'})


@app.route('/api/my-tribe', methods=['GET'])
@require_role('tribe_master', 'team_lead', 'admin')
def my_tribe():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    available_tribes = _tribes_for_pool(pool_id)
    tribe = normalize_tribe(request.args.get('tribe')) or _resolve_user_tribe(g.user, pool_id)
    if g.current_role in ('team_lead', 'admin') and not tribe:
        tribe = available_tribes[0] if available_tribes else ''
    students = Student.query.filter_by(pool_id=pool_id, tribe=tribe).order_by(Student.nick).all()
    rankings = _tribe_rankings(pool_id)
    own_rank = next((row['rank'] for row in rankings if row['tribe'] == tribe), None)
    next_events = (
        TribeEvent.query
        .filter(TribeEvent.pool_id == pool_id, TribeEvent.tribe == tribe, TribeEvent.event_date >= _moscow_today())
        .order_by(TribeEvent.event_date, TribeEvent.time_start)
        .all()
    )
    student_ids = [student.id for student in students]
    event_rows = (
        db.session.query(StudentEvent, Student)
        .join(Student, Student.id == StudentEvent.student_id)
        .filter(StudentEvent.student_id.in_(student_ids))
        .order_by(StudentEvent.event_date.desc(), StudentEvent.created_at.desc())
        .all()
        if student_ids else []
    )
    all_tribe_events = (
        TribeEvent.query
        .filter(TribeEvent.pool_id == pool_id, TribeEvent.event_date >= _moscow_today())
        .order_by(TribeEvent.event_date, TribeEvent.time_start, TribeEvent.tribe)
        .all()
        if g.current_role in ('team_lead', 'admin') else []
    )
    return jsonify({
        **_tribe_metrics(pool_id, tribe),
        'rank': own_rank,
        'rankings': rankings,
        'available_tribes': available_tribes,
        'students': [{
            'id': student.id,
            'nick': student.nick,
            'name': student.name,
            'tribe': student.tribe,
        } for student in students],
        'student_events': [{
            'id': event.id,
            'student_id': student.id,
            'student_nick': student.nick,
            'student_name': student.name,
            'type': event.event_type,
            'date': event.event_date.isoformat() if event.event_date else None,
            'post_url': event.post_url or '',
            'proof_url': event.proof_url or '',
            'points': event.points or 0,
            'status': event.status or 'pending',
            'comment': event.comment or '',
        } for event, student in event_rows],
        'tribe_events': [_tribe_event_to_dict(event) for event in next_events],
        'all_tribe_events': [_tribe_event_to_dict(event) for event in all_tribe_events],
    })


@app.route('/api/tribe-events', methods=['GET'])
@require_auth
def list_tribe_events():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    start = request.args.get('start')
    query = TribeEvent.query.filter_by(pool_id=pool_id)
    if start:
        query = query.filter(TribeEvent.event_date >= datetime.fromisoformat(start).date())
    events = query.order_by(TribeEvent.event_date, TribeEvent.time_start).all()
    return jsonify([_tribe_event_to_dict(event) for event in events])


@app.route('/api/tribe-events', methods=['POST'])
@require_role('tribe_master', 'team_lead', 'admin')
def create_tribe_event():
    data = request.json or {}
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    tribe = normalize_tribe(data.get('tribe') or g.current_tribe)
    if g.current_role == 'tribe_master' and tribe != normalize_tribe(g.current_tribe):
        return jsonify({'error': 'Можно создавать встречи только своего трайба'}), 403
    title = (data.get('title') or '').strip()
    if not tribe or not title or not data.get('event_date'):
        return jsonify({'error': 'Нужны tribe, title и event_date'}), 400
    try:
        event_date = datetime.fromisoformat(data['event_date']).date()
    except ValueError:
        return jsonify({'error': 'Некорректная дата'}), 400
    event = TribeEvent(
        pool_id=pool_id,
        tribe=tribe,
        title=title,
        event_date=event_date,
        time_start=(data.get('time_start') or '').strip(),
        location=(data.get('location') or '').strip(),
        comment=(data.get('comment') or '').strip(),
        created_by=g.user.id,
    )
    db.session.add(event)
    db.session.commit()
    return jsonify(_tribe_event_to_dict(event)), 201


@app.route('/api/tribe-events/<int:event_id>', methods=['DELETE'])
@require_role('tribe_master', 'team_lead', 'admin')
def delete_tribe_event(event_id):
    event = get_model_or_404(TribeEvent, event_id)
    error = _active_entity_error(event.pool_id)
    if error:
        return error
    if g.current_role == 'tribe_master' and normalize_tribe(event.tribe) != normalize_tribe(g.current_tribe):
        return jsonify({'error': 'Можно удалять встречи только своего трайба'}), 403
    _cancel_pending_notifications('tribe_event', event.id, ['tribe_event_tomorrow', 'tribe_event_first_day_shift'])
    db.session.delete(event)
    db.session.commit()
    return jsonify({'message': 'Встреча трайба удалена'})


@app.route('/api/tribe-events/generate-standard', methods=['POST'])
@require_role('tribe_master', 'team_lead', 'admin')
def generate_standard_tribe_events():
    data = request.json or {}
    pool_id = active_pool_id()
    pool = db.session.get(Pool, pool_id) if pool_id else None
    if not pool or not pool.start_date:
        return jsonify({'error': 'Нет активного бассейна с датой старта'}), 400

    available_tribes = _tribes_for_pool(pool_id)
    user_tribe = normalize_tribe(g.current_tribe)
    requested_tribe = normalize_tribe(data.get('tribe'))
    if g.current_role == 'tribe_master':
        tribes = [user_tribe]
    elif requested_tribe:
        if requested_tribe not in available_tribes:
            return jsonify({'error': 'Выбранный трайб не найден'}), 400
        tribes = [requested_tribe]
    else:
        tribes = available_tribes
    tribes = [tribe for tribe in tribes if tribe]
    if not tribes:
        return jsonify({'error': 'Нет заведенных трайбов'}), 400

    first_day = pool.start_date
    first_exam_day = first_day + timedelta(days=3)
    final_day = first_day + timedelta(days=13)
    first_day_times = ['17:30', '18:00', '18:30', '19:00', '19:30', '20:00']
    created = 0
    updated = 0

    def upsert_event(tribe, event_date, time_start, title, comment=''):
        nonlocal created, updated
        existing = TribeEvent.query.filter_by(
            pool_id=pool_id,
            tribe=tribe,
            event_date=event_date,
            time_start=time_start,
            title=title,
        ).first()
        if existing:
            existing.comment = comment or existing.comment
            updated += 1
            return
        db.session.add(TribeEvent(
            pool_id=pool_id,
            tribe=tribe,
            title=title,
            event_date=event_date,
            time_start=time_start,
            comment=comment,
            created_by=g.user.id,
        ))
        created += 1

    for index, tribe in enumerate(tribes):
        upsert_event(
            tribe,
            first_day,
            first_day_times[index % len(first_day_times)],
            'Стартовая встреча трайба',
        )
        upsert_event(
            tribe,
            first_exam_day,
            '17:30',
            'Встреча после первого экзамена',
        )
        upsert_event(
            tribe,
            final_day,
            '17:30',
            'Общая рефлексия всех трайбов',
        )

    db.session.commit()
    tribe_label = tribes[0] if len(tribes) == 1 else 'всех трайбов'
    return jsonify({
        'created': created,
        'updated': updated,
        'message': f'Для {tribe_label}: создано встреч {created}, обновлено {updated}',
    })


@app.route('/api/students/<int:student_id>', methods=['DELETE'])
@require_role('admin', 'team_lead')
def delete_student(student_id):
    student = get_model_or_404(Student, student_id)
    error = _active_entity_error(student.pool_id)
    if error:
        return error
    if StudentPenalty.query.filter_by(student_id=student.id).first():
        return jsonify({'error': 'Сначала удалите связанные штрафы ученика'}), 409
    StudentEvent.query.filter_by(student_id=student.id).delete()
    db.session.delete(student)
    db.session.commit()
    return jsonify({'message': 'Ученик удалён'})


@app.route('/api/students/<int:student_id>', methods=['PATCH'])
@require_role('admin', 'team_lead')
def update_student(student_id):
    student = get_model_or_404(Student, student_id)
    error = _active_entity_error(student.pool_id)
    if error:
        return error
    data = request.json or {}

    if 'tribe' in data:
        student.tribe = normalize_tribe(data.get('tribe'))

    db.session.commit()
    return jsonify({'message': 'Ученик обновлён'})


# ==================== Штрафы (видят все авторизованные) ====================


@app.route('/api/penalties', methods=['GET'])
@require_auth
def get_penalties():
    pool_id, error = _active_pool_id_for_request(request.args.get('pool_id', type=int))
    if error:
        return error
    penalties = StudentPenalty.query.filter_by(pool_id=pool_id).order_by(StudentPenalty.date_issued.desc()).all()
    penalty_ids = [p.id for p in penalties]
    histories = {}
    if penalty_ids:
        for row in PenaltyHistory.query.filter(PenaltyHistory.penalty_id.in_(penalty_ids)).order_by(PenaltyHistory.created_at).all():
            histories.setdefault(row.penalty_id, []).append({
                'id': row.id,
                'old_status': row.old_status,
                'new_status': row.new_status,
                'old_hours': row.old_hours,
                'new_hours': row.new_hours,
                'actor_nick': row.actor_nick,
                'actor_name': row.actor_name,
                'comment': row.comment or '',
                'created_at': row.created_at.isoformat(),
            })
    return jsonify([{
        'id': p.id,
        'student_id': p.student_id,
        'student_name': p.student_name,
        'volunteer_name': p.volunteer_name,
        'hours': p.hours,
        'multiplier': p.multiplier,
        'total_hours': p.hours * p.multiplier,
        'workoff_status': p.workoff_status,
        'description': p.description,
        'date_issued': p.date_issued.isoformat(),
        'date_worked_off': p.date_worked_off.isoformat() if p.date_worked_off else None,
        'workoff_started_at': p.date_worked_off.isoformat() if p.workoff_status == 'in_workoff' and p.date_worked_off else None,
        'history': histories.get(p.id, []),
    } for p in penalties])


@app.route('/api/penalties', methods=['POST'])
@require_auth
def create_penalty():
    data = request.json or {}
    user = g.user
    pool_id, error = _active_pool_id_for_request(data.get('pool_id'))
    if error:
        return error
    student = None
    if data.get('student_id') is not None:
        try:
            student = db.session.get(Student, int(data['student_id']))
        except (TypeError, ValueError):
            return jsonify({'error': 'Некорректный student_id'}), 400
    else:
        student_name = (data.get('student_name') or '').strip()
        if student_name:
            student = Student.query.filter(
                Student.pool_id == pool_id,
                db.func.lower(Student.nick) == student_name.lower(),
            ).first()
    if not student or student.pool_id != pool_id:
        return jsonify({'error': 'Ученик не найден в активном бассейне'}), 404
    _acquire_dedupe_lock('penalty-create', pool_id, student.id)
    duplicate_penalty = StudentPenalty.query.filter(
        StudentPenalty.pool_id == pool_id,
        StudentPenalty.student_id == student.id,
        StudentPenalty.date_issued >= _utcnow() - PENALTY_DUPLICATE_WINDOW,
    ).order_by(StudentPenalty.date_issued.desc()).first()
    if duplicate_penalty:
        db.session.commit()
        return jsonify({
            'id': duplicate_penalty.id,
            'duplicate': True,
            'message': 'Этому ученику уже выдан пенальти менее 3 минут назад',
        }), 200
    penalty = StudentPenalty(
        student_id=student.id,
        student_name=student.nick,
        volunteer_id=user.id,
        volunteer_name=user.name or user.nick,
        hours=2,
        multiplier=1,
        description=data.get('description', ''),
        pool_id=pool_id,
    )
    db.session.add(penalty)
    db.session.flush()
    add_penalty_history(penalty, None, penalty.workoff_status, None, 'Штраф создан')
    log_action(
        'create',
        'penalty',
        penalty.id,
        f'Штраф для {penalty.student_name}: 2h',
        {
            'student': penalty.student_name,
            'volunteer': penalty.volunteer_name,
            'description': penalty.description,
        },
    )
    critical_events = _notify_admins_penalty_created(penalty)
    _queue_penalty_method_question(
        penalty,
        scheduled_for=_utcnow() + timedelta(minutes=5),
        suffix='initial',
    )
    db.session.commit()
    queue_notification_dispatch(critical_events)
    return jsonify({'id': penalty.id, 'duplicate': False, 'message': 'Штраф добавлен'}), 201


@app.route('/api/penalties/<int:penalty_id>', methods=['PATCH'])
@require_auth
def update_penalty_status(penalty_id):
    data = request.json or {}
    penalty = get_model_or_404(StudentPenalty, penalty_id)
    error = _active_entity_error(penalty.pool_id)
    if error:
        return error
    if not _can_access_pool_id(g.user, penalty.pool_id):
        return jsonify({'error': 'У тебя нет доступа к этому бассейну'}), 403
    new_status = data.get('workoff_status', penalty.workoff_status)
    if new_status not in PENALTY_STATUSES:
        return jsonify({'error': 'Некорректный статус штрафа'}), 400
    _acquire_dedupe_lock('penalty-status', penalty.id)
    db.session.refresh(penalty)
    old_status = penalty.workoff_status
    old_hours = penalty.hours * penalty.multiplier
    critical_events = []
    if new_status == old_status:
        db.session.commit()
        return jsonify({
            'message': 'Статус штрафа уже обновлён',
            'duplicate': True,
            'workoff_status': penalty.workoff_status,
        })
    if new_status == 'awaiting_unlock':
        recent_unlock_transition = PenaltyHistory.query.filter(
            PenaltyHistory.penalty_id == penalty.id,
            PenaltyHistory.new_status == 'awaiting_unlock',
            PenaltyHistory.created_at >= _utcnow() - PENALTY_DUPLICATE_WINDOW,
        ).first()
        if recent_unlock_transition:
            db.session.commit()
            return jsonify({
                'message': 'Запрос на разблокировку уже создан менее 3 минут назад',
                'duplicate': True,
                'workoff_status': penalty.workoff_status,
            })
    penalty.workoff_status = new_status
    if new_status == 'overdue' and old_status == 'pending':
        penalty.multiplier *= 2
    if new_status == 'in_workoff':
        penalty.date_worked_off = _utcnow()
    if new_status in ('done', 'awaiting_unlock'):
        penalty.date_worked_off = _utcnow()
    if new_status == 'pending':
        penalty.date_worked_off = None
    if new_status != old_status:
        if new_status == 'in_workoff':
            _cancel_pending_notifications('penalty', penalty.id, ['penalty_method_question'])
        if new_status == 'awaiting_unlock':
            _cancel_pending_notifications('penalty', penalty.id, ['penalty_method_question', 'penalty_workoff_check'])
            critical_events = _notify_admins_penalty_awaiting_unlock(penalty)
        if new_status in ('pending', 'overdue', 'unlocked', 'done'):
            _cancel_pending_notifications('penalty', penalty.id, ['penalty_method_question', 'penalty_workoff_check'])
    if old_status != new_status or old_hours != penalty.hours * penalty.multiplier:
        add_penalty_history(penalty, old_status, new_status, old_hours, data.get('comment') or '')
        log_action(
            'update',
            'penalty',
            penalty.id,
            f'Штраф {penalty.student_name}: {old_status} → {new_status}, {old_hours}h → {penalty.hours * penalty.multiplier}h',
            {
                'student': penalty.student_name,
                'old_status': old_status,
                'new_status': new_status,
                'old_hours': old_hours,
                'new_hours': penalty.hours * penalty.multiplier,
            },
        )
    db.session.commit()
    if critical_events:
        queue_notification_dispatch(critical_events)
    return jsonify({
        'message': 'Штраф обновлён',
        'duplicate': False,
        'workoff_status': penalty.workoff_status,
    })


@app.route('/api/penalties/<int:penalty_id>', methods=['DELETE'])
@require_role('admin', 'team_lead')
def delete_penalty(penalty_id):
    penalty = get_model_or_404(StudentPenalty, penalty_id)
    error = _active_entity_error(penalty.pool_id)
    if error:
        return error
    if not _can_access_pool_id(g.user, penalty.pool_id):
        return jsonify({'error': 'У тебя нет доступа к этому бассейну'}), 403
    student_name = penalty.student_name
    _cancel_pending_notifications('penalty', penalty.id)
    PenaltyHistory.query.filter_by(penalty_id=penalty.id).delete(synchronize_session=False)
    log_action(
        'delete',
        'penalty',
        penalty.id,
        f'Удалён штраф для {student_name}',
        {
            'student': student_name,
            'status': penalty.workoff_status,
            'hours': penalty.hours * penalty.multiplier,
        },
    )
    db.session.delete(penalty)
    db.session.commit()
    return jsonify({'message': f'Штраф для {student_name} удалён'})


# ==================== Админ ====================


def cell_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return 'да' if value else ''
    if value is None:
        return ''
    return value


def export_sheet(headers, rows):
    return {
        'headers': headers,
        'rows': [[cell_value(item) for item in row] for row in rows],
    }


STATUS_EXPORT_LABELS = {
    'pending': 'Ожидает отработки',
    'overdue': 'Не пришёл',
    'awaiting_unlock': 'Ждёт разблокировки',
    'unlocked': 'Разблокирован',
    'done': 'Отработано',
}

EVENT_TYPE_EXPORT_LABELS = {
    'entertainment': 'Развлекательное',
    'education': 'Образовательное',
}
ROLE_EXPORT_LABELS = {
    'volunteer': 'Волонтёр',
    'tribe_master': 'Трайб-мастер',
    'team_lead': 'Тимлид',
    'admin': 'Админ',
}


def _export_event_type_label(event_type):
    return EVENT_TYPE_EXPORT_LABELS.get(event_type, event_type or '')


def build_export_sheets():
    users = {u.id: u for u in User.query.all()}
    blocks = {b.id: b for b in ShiftBlock.query.all()}
    pools = {p.id: p for p in Pool.query.all()}
    students = {s.id: s for s in Student.query.all()}

    sheets = {}
    active_pool = Pool.query.filter_by(active=True).order_by(Pool.created_at.desc()).first()
    start_date = active_pool.start_date if active_pool and active_pool.start_date else None
    all_blocks = ShiftBlock.query.order_by(ShiftBlock.date, ShiftBlock.time_start).all()
    if not start_date and all_blocks:
        start_date = min(block.date for block in all_blocks)

    shift_rows = []
    if start_date:
        shift_rows.append(['СТАРТ БАССЕЙНА:', '', '', '', '', '', start_date, '', '', 'Подсмотреть заработанные коины:', '', ''])
        signups_by_block = {}
        for signup in Signup.query.order_by(Signup.created_at).all():
            signups_by_block.setdefault(signup.block_id, []).append(users.get(signup.user_id))
        blocks_by_date = {}
        for block in all_blocks:
            blocks_by_date.setdefault(block.date, []).append(block)
        current = start_date
        end_date = max((block.date for block in all_blocks), default=start_date)
        while current <= end_date:
            week = [current + timedelta(days=i) for i in range(7)]
            shift_rows.append(week + ['', '', 'никнейм', 'коины'])
            day_lines = []
            max_lines = 1
            for day in week:
                lines = []
                for block in sorted(blocks_by_date.get(day, []), key=lambda item: item.time_start):
                    label = f'{block.time_start} - {block.time_end}{(" " + block.label) if block.label else ""}'
                    lines.append(label)
                    for user in signups_by_block.get(block.id, []):
                        if user:
                            lines.append(user.nick)
                day_lines.append(lines)
                max_lines = max(max_lines, len(lines))
            for index in range(max_lines):
                row = [(lines[index] if index < len(lines) else '') for lines in day_lines]
                shift_rows.append(row + ['', '', '', ''])
            shift_rows.append([''] * 11)
            current += timedelta(days=7)
        shift_rows.append(['Created by app export'])
        coin_totals = []
        for user in User.query.filter(User.role.in_(list(VOLUNTEER_PROFILE_ROLES))).order_by(User.nick).all():
            rewards = calculate_user_rewards(user, user.coins_adjustment or 0)
            coin_totals.append([user.nick, rewards['total']])
        for index, item in enumerate(coin_totals, start=2):
            while len(shift_rows) <= index:
                shift_rows.append([''] * 11)
            while len(shift_rows[index]) < 11:
                shift_rows[index].append('')
            shift_rows[index][9] = item[0]
            shift_rows[index][10] = item[1]
    sheets['shifts'] = {'headers': [], 'rows': [[cell_value(item) for item in row] for row in shift_rows]}

    sheets['site_signups'] = export_sheet(
        ['Когда', 'Действие', 'Дата', 'Время', 'Метка', 'Ник'],
        [[s.created_at, 'create', blocks.get(s.block_id).date if blocks.get(s.block_id) else '',
          f'{blocks.get(s.block_id).time_start} - {blocks.get(s.block_id).time_end}' if blocks.get(s.block_id) else '',
          blocks.get(s.block_id).label if blocks.get(s.block_id) else '', users.get(s.user_id).nick if users.get(s.user_id) else '']
         for s in Signup.query.order_by(Signup.created_at).all()],
    )

    sheets['penalty'] = export_sheet(
        ['Ник нарушителя', 'Дата и время назначения пенальти', 'Причина пенальти', 'Кто зафиксировал пенальти (ник волонтера)',
         'Статус penalty (для отслеживания)', 'Кто занес', 'Дата выдачи пенальти', 'Часы', 'Ждёт разблокировки'],
        [[p.student_name, p.date_issued, p.description, p.volunteer_name,
          STATUS_EXPORT_LABELS.get(p.workoff_status, p.workoff_status), p.volunteer_name, p.date_issued,
          p.hours * p.multiplier, 'да' if p.workoff_status == 'awaiting_unlock' else '']
         for p in StudentPenalty.query.order_by(StudentPenalty.date_issued).all()],
    )
    sheets['tribe_event'] = export_sheet(
        ['Ник участника', 'Трайб', 'Дата проведения', 'Тип мероприятия', 'Краткое описание',
         'Ник трайб-мастера', 'Дата заполнения', 'Баллы', 'Статус', 'Ссылка на пост'],
        [[students.get(e.student_id).nick if students.get(e.student_id) else '', students.get(e.student_id).tribe if students.get(e.student_id) else '',
          e.event_date, _export_event_type_label(e.event_type), e.comment, users.get(e.created_by).nick if users.get(e.created_by) else '',
          e.created_at, e.points or STUDENT_EVENT_POINTS.get(e.event_type, 0), e.status or '', e.post_url]
         for e in StudentEvent.query.order_by(StudentEvent.event_date, StudentEvent.created_at).all()],
    )

    reward_rows = []
    for user in User.query.filter(User.role.in_(list(VOLUNTEER_PROFILE_ROLES))).order_by(User.nick).all():
        rewards = calculate_user_rewards(user, user.coins_adjustment or 0)
        by_type = {item['type']: item for item in rewards['breakdown']}
        first_day = by_type.get('first_day_hour', {'count': 0, 'coins': 0})
        first_week = by_type.get('first_week_or_weekend_hour', {'count': 0, 'coins': 0})
        weekdays = by_type.get('subsequent_weekday_hour', {'count': 0, 'coins': 0})
        exams = by_type.get('exam_hour', {'count': 0, 'coins': 0})
        group_review = by_type.get('group_review', {'count': 0, 'coins': 0})
        tribe_master = by_type.get('tribe_master_event', {'count': 0, 'coins': 0})
        confession = by_type.get('confession', {'count': 0, 'coins': 0})
        manual = by_type.get('manual', {'count': 0, 'coins': 0})
        reward_rows.append([
            user.nick,
            first_day.get('count', 0),
            first_week.get('count', 0),
            weekdays.get('count', 0),
            exams.get('count', 0),
            group_review.get('count', 0),
            tribe_master.get('count', 0),
            confession.get('count', 0),
            '',
            first_day.get('coins', 0),
            first_week.get('coins', 0),
            weekdays.get('coins', 0),
            exams.get('coins', 0),
            group_review.get('coins', 0),
            tribe_master.get('coins', 0),
            confession.get('coins', 0),
            manual.get('coins', 0),
            rewards['total'],
        ])
    sheets['reward_calc'] = export_sheet(
        ['Никнейм', 'Первый день', 'Первая неделя и выходные', 'Будние дни', 'Экзамены',
         'Групповые', 'Трайб-ивенты', 'Исповедь', 'ИТОГ', 'Первый день', 'Первая неделя и выходные',
         'Будние дни', 'Экзамены', 'Групповые', 'Трайб-ивенты', 'Исповедь', 'Допы', 'ИТОГО'],
        reward_rows,
    )

    group_reviewers = db.session.query(User).join(GroupReview, GroupReview.reviewer_id == User.id).distinct().order_by(User.nick).all()
    volunteer_rows = []
    volunteer_list = User.query.filter(User.role.in_(list(VOLUNTEER_PROFILE_ROLES))).order_by(User.nick).all()
    tribe_masters = User.query.filter_by(role='tribe_master').order_by(User.tribe, User.nick).all()
    max_volunteer_rows = max(len(volunteer_list), len(tribe_masters), len(group_reviewers), 1)
    for index in range(max_volunteer_rows):
        person = volunteer_list[index] if index < len(volunteer_list) else None
        master = tribe_masters[index] if index < len(tribe_masters) else None
        reviewer = group_reviewers[index] if index < len(group_reviewers) else None
        volunteer_rows.append([
            person.name or person.nick if person else '',
            person.nick if person else '',
            f'{master.tribe} ->' if master and master.tribe else '',
            master.nick if master else '',
            '',
            reviewer.nick if reviewer else '',
            '',
            person.nick if person else '',
            person.name or person.nick if person else '',
        ])
    sheets['volunteers'] = export_sheet(
        ['ФИО', 'Никнейм', '', 'Трайб мастера', '', 'Проверка групповых', '', '', ''],
        volunteer_rows,
    )

    review_grid_rows = []
    reviews_by_date = {}
    for review in GroupReview.query.order_by(GroupReview.review_date, GroupReview.time_start).all():
        reviews_by_date.setdefault(review.review_date, []).append(review)
    for review_date, reviews in reviews_by_date.items():
        row = [review_date, 'проверки']
        for review in reviews:
            reviewer = users.get(review.reviewer_id)
            row.extend([reviewer.nick if reviewer else ''] * max(1, review.quantity or 1))
        review_grid_rows.append(row)
    sheets['group_review'] = export_sheet(
        ['проверка групповых'],
        review_grid_rows,
    )

    confession_rows = [['дата проведения', start_date + timedelta(days=4) if start_date else '']]
    for index, user in enumerate(User.query.filter_by(has_confession=True).order_by(User.nick).all(), start=1):
        confession_rows.append([index, user.nick])
    sheets['confession'] = export_sheet(
        ['Иcповедь волонтёра'],
        confession_rows,
    )

    sheets['tribe_events'] = export_sheet(
        ['Дата', 'Время', 'Трайб', 'Название', 'Место', 'Комментарий', 'Кто занёс'],
        [[e.event_date, e.time_start, e.tribe, e.title, e.location, e.comment, users.get(e.created_by).nick if users.get(e.created_by) else '']
         for e in TribeEvent.query.order_by(TribeEvent.event_date, TribeEvent.time_start).all()],
    )

    sheets['reward'] = export_sheet(
        ['Type', 'Coins'],
        [
            ['Дежурство в первую неделю интенсива и в выходные дни (за 1 час)', REWARD_RATES['first_week_or_weekend_hour']],
            ['Дежурство в последующие недели интенсива | Будние дни (за 1 час)', REWARD_RATES['subsequent_weekday_hour']],
            ['Проверка групповых (1 проверка)', REWARD_RATES['group_review']],
            ['Участие в проведении экзамена (за 1 час)', REWARD_RATES['exam_hour']],
            ['Трайб мастер на отборочном интенсиве (1 мероприятие)', REWARD_RATES['tribe_master_event']],
            ['Тим лидер команды волонтеров', REWARD_RATES['team_lead']],
            ['Волонтерство в первый день отборочного интенсива (1 час)', REWARD_RATES['first_day_hour']],
            ['Участие в исповеди', REWARD_RATES['confession']],
        ],
    )

    sheets['penalty_history'] = export_sheet(
        ['ID штрафа', 'Было', 'Стало', 'Было часов', 'Стало часов', 'Кто изменил', 'Комментарий', 'Когда'],
        [[h.penalty_id, STATUS_EXPORT_LABELS.get(h.old_status, h.old_status), STATUS_EXPORT_LABELS.get(h.new_status, h.new_status),
          h.old_hours, h.new_hours, h.actor_nick, h.comment, h.created_at]
         for h in PenaltyHistory.query.order_by(PenaltyHistory.created_at).all()],
    )
    sheets['action_log'] = export_sheet(
        ['Когда', 'Кто', 'Действие', 'Сущность', 'Описание'],
        [[a.created_at, a.actor_nick, a.action, a.entity, a.description]
         for a in ActionLog.query.order_by(ActionLog.created_at).all()],
    )
    return sheets


def build_export_workbook():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    sheets = build_export_sheets()
    template = load_google_sheet_template()
    if template:
        return build_template_export_workbook(template, sheets)

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill('solid', fgColor='1F1F1F')
    header_font = Font(bold=True, color='00FF00')
    for name, data in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if data['headers']:
            ws.append(data['headers'])
        for row in data['rows']:
            ws.append(row)
        if ws.max_row:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2' if data['headers'] else None
        for column in ws.columns:
            letter = column[0].column_letter
            width = min(36, max(10, max(len(str(cell.value or '')) for cell in column[:200]) + 2))
            ws.column_dimensions[letter].width = width
    return wb


def load_google_sheet_template():
    sheet_id = os.getenv('GOOGLE_SHEETS_ID', '')
    if not sheet_id:
        return None
    url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
    try:
        import requests
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        from openpyxl import load_workbook
        return load_workbook(BytesIO(response.content))
    except Exception as e:
        print('[export] template download failed:', e)
        return None


def write_matrix(ws, values, start_row=1, start_col=1, preserve_first_row=False):
    if not values:
        return
    first_row = start_row + (1 if preserve_first_row else 0)
    max_rows = max(0, ws.max_row - first_row + 1)
    max_cols = max(ws.max_column or 0, max((len(row) for row in values), default=0))
    if max_rows and max_cols:
        for row in ws.iter_rows(min_row=first_row, max_row=ws.max_row, min_col=start_col, max_col=max_cols):
            for cell in row:
                cell.value = None
    for r_index, row in enumerate(values, start=start_row):
        if preserve_first_row and r_index == start_row:
            continue
        for c_index, value in enumerate(row, start=start_col):
            ws.cell(r_index, c_index).value = value


def export_values_for_sheet(name, data):
    return [data['headers'], *data['rows']] if data['headers'] else data['rows']


def build_template_export_workbook(wb, sheets):
    source_names = list(sheets.keys())
    for name in list(wb.sheetnames):
        if name.startswith('export_'):
            del wb[name]
    for name in source_names:
        if name in wb.sheetnames:
            ws = wb.copy_worksheet(wb[name])
            ws.title = f'export_{name}'[:31]
        else:
            ws = wb.create_sheet(title=f'export_{name}'[:31])
        values = export_values_for_sheet(name, sheets[name])
        preserve_first_row = name not in ('shifts', 'tribe_events')
        write_matrix(ws, values, preserve_first_row=preserve_first_row)
    meta = wb.create_sheet(title='export_meta') if 'export_meta' not in wb.sheetnames else wb['export_meta']
    write_matrix(meta, [
        ['key', 'value'],
        ['exported_at', _utcnow().isoformat()],
        ['sheets', ', '.join([f'export_{name}' for name in source_names])],
    ])
    return wb


def save_export_workbook(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    workbook = build_export_workbook()
    workbook.save(path)
    return path


def latest_backup_file():
    if not os.path.isdir(BACKUP_DIR):
        return None
    files = [
        os.path.join(BACKUP_DIR, name)
        for name in os.listdir(BACKUP_DIR)
        if name.endswith('.xlsx')
    ]
    return max(files, key=os.path.getmtime) if files else None


def backup_filename_for_today():
    return os.path.join(BACKUP_DIR, f'pool-backup-{_moscow_today().isoformat()}.xlsx')


def create_daily_backup(force=False):
    if not _backup_lock.acquire(blocking=False):
        return None
    try:
        path = backup_filename_for_today()
        if not force and os.path.exists(path):
            return path
        with app.app_context():
            save_export_workbook(path)
        print(f'[backup] сохранён резерв: {path}')
        return path
    finally:
        _backup_lock.release()


def backup_worker_loop():
    while True:
        try:
            create_daily_backup(force=False)
        except Exception as e:
            print('[backup] error:', e)
        time.sleep(BACKUP_INTERVAL)


def start_backup_worker():
    t = threading.Thread(target=backup_worker_loop, daemon=True)
    t.start()
    print(f'[backup] воркер запущен, папка={BACKUP_DIR}')


def start_runtime_services():
    global _runtime_started
    if _runtime_started:
        return
    _runtime_started = True
    start_backup_worker()


@app.route('/api/admin/export.xlsx', methods=['GET'])
@require_role('team_lead', 'admin')
def export_xlsx():
    output = BytesIO()
    workbook = build_export_workbook()
    workbook.save(output)
    output.seek(0)
    filename = f'pool-export-{_utcnow().strftime("%Y-%m-%d-%H%M")}.xlsx'
    log_action('export', 'workbook', None, 'Скачан Excel-экспорт', {'filename': filename})
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


GOOGLE_TEMPLATE_SHEETS = ['volunteers', 'shifts', 'penalty', 'tribe_event', 'tribe_events']


def _template_penalty_status(status):
    if status in {'done', 'unlocked'}:
        return 'Отработано'
    if status in {'in_workoff', 'awaiting_unlock'}:
        return 'Назначено'
    return 'Ожидает'


def build_google_sheets_template_payload(pool_id):
    """Build a pool-scoped snapshot for the existing formatted Sheets template."""
    pool = get_model_or_404(Pool, pool_id)
    memberships = (
        db.session.query(PoolVolunteer, User)
        .join(User, User.id == PoolVolunteer.user_id)
        .filter(PoolVolunteer.pool_id == pool_id)
        .order_by(User.nick)
        .all()
    )
    volunteers = sorted(({
        'nick': user.nick,
        'name': user.name or user.nick,
        'role': membership.pool_role or 'volunteer',
        'tribe': membership.tribe or '',
        'is_group_reviewer': bool(user.is_group_reviewer),
    }
        for membership, user in memberships
        if membership.pool_role in {'volunteer', 'tribe_master'}
    ), key=lambda item: item['nick'].lower())
    block_rows = ShiftBlock.query.filter_by(pool_id=pool_id).order_by(
        ShiftBlock.date,
        ShiftBlock.time_start,
        ShiftBlock.id,
    ).all()
    block_ids = [block.id for block in block_rows]
    signup_rows = (
        db.session.query(Signup, User)
        .join(User, User.id == Signup.user_id)
        .filter(Signup.block_id.in_(block_ids))
        .order_by(Signup.created_at, Signup.id)
        .all()
        if block_ids else []
    )
    signups_by_block = defaultdict(list)
    for signup, user in signup_rows:
        signups_by_block[signup.block_id].append(user.nick)

    students = {
        student.id: student
        for student in Student.query.filter_by(pool_id=pool_id).all()
    }
    student_ids = list(students)
    student_events = (
        StudentEvent.query
        .filter(StudentEvent.student_id.in_(student_ids))
        .order_by(StudentEvent.event_date, StudentEvent.created_at, StudentEvent.id)
        .all()
        if student_ids else []
    )
    creators = {
        user.id: user
        for user in User.query.filter(User.id.in_({
            event.created_by for event in student_events if event.created_by
        })).all()
    }
    penalties = StudentPenalty.query.filter_by(pool_id=pool_id).order_by(
        StudentPenalty.date_issued,
        StudentPenalty.id,
    ).all()
    tribe_events = TribeEvent.query.filter_by(pool_id=pool_id).order_by(
        TribeEvent.event_date,
        TribeEvent.time_start,
        TribeEvent.id,
    ).all()

    return {
        'secret': SYNC_SECRET,
        'mode': 'template_snapshot_v1',
        'exported_at': _utcnow().isoformat(),
        'pool': {
            'id': pool.id,
            'name': pool.name,
            'start_date': pool.start_date.isoformat() if pool.start_date else None,
        },
        'volunteers': volunteers,
        'shifts': [{
            'id': block.id,
            'date': block.date.isoformat(),
            'time_start': block.time_start,
            'time_end': block.time_end,
            'label': block.label or '',
            'capacity': block.capacity,
            'volunteers': signups_by_block.get(block.id, []),
        } for block in block_rows],
        'penalties': [{
            'id': penalty.id,
            'student': penalty.student_name,
            'issued_at': penalty.date_issued.isoformat() if penalty.date_issued else None,
            'description': penalty.description or '',
            'volunteer': penalty.volunteer_name or '',
            'status': _template_penalty_status(penalty.workoff_status),
            'entered_by': penalty.volunteer_name or '',
            'assigned_at': penalty.date_issued.isoformat() if penalty.date_issued else None,
        } for penalty in penalties],
        'student_events': [{
            'id': event.id,
            'student': students[event.student_id].nick,
            'tribe': students[event.student_id].tribe or '',
            'event_date': event.event_date.isoformat() if event.event_date else None,
            'event_type': _export_event_type_label(event.event_type),
            'description': event.comment or '',
            'tribe_master': creators[event.created_by].nick if event.created_by in creators else '',
            'created_at': event.created_at.isoformat() if event.created_at else None,
            'points': event.points or STUDENT_EVENT_POINTS.get(event.event_type, 0),
            'status': event.status or '',
            'post_url': event.post_url or '',
        } for event in student_events],
        'tribe_events': [{
            'id': event.id,
            'date': event.event_date.isoformat(),
            'time_start': event.time_start or '',
            'tribe': event.tribe or '',
            'title': event.title or '',
            'location': event.location or '',
            'comment': event.comment or '',
        } for event in tribe_events],
    }


def _google_sheet_connection_payload(pool):
    return {
        'pool_id': pool.id,
        'sheet_url': pool.google_sheet_url or '',
        'webhook_url': pool.google_sheet_webhook_url or '',
        'enabled': bool(pool.google_sheet_enabled),
        'last_export_at': pool.google_sheet_last_export_at.isoformat() if pool.google_sheet_last_export_at else None,
        'last_error': pool.google_sheet_last_error or '',
    }


def export_pool_to_google_sheets(pool):
    if not pool.google_sheet_webhook_url:
        raise ValueError('Не указан URL Apps Script')
    if not SYNC_SECRET:
        raise ValueError('На сервере не настроен SYNC_SECRET')
    payload = build_google_sheets_template_payload(pool.id)
    try:
        import requests
        response = requests.post(pool.google_sheet_webhook_url, json=payload, timeout=60)
        data = response.json() if response.content else {}
    except Exception as exc:
        pool.google_sheet_last_error = str(exc)[:1000]
        raise RuntimeError(f'Не удалось отправить данные: {exc}') from exc
    if response.status_code != 200 or not isinstance(data, dict) or not data.get('ok'):
        error = data.get('error') if isinstance(data, dict) else None
        pool.google_sheet_last_error = str(error or f'HTTP {response.status_code}')[:1000]
        raise RuntimeError(pool.google_sheet_last_error)
    pool.google_sheet_last_export_at = _utcnow()
    pool.google_sheet_last_error = ''
    return data


def _active_google_sheets_pool_or_error(pool_id):
    pool = get_model_or_404(Pool, pool_id)
    if not pool.active or pool.archived:
        return None, (jsonify({'error': 'Google Sheets можно настраивать только для активного бассейна'}), 409)
    return pool, None


@app.route('/api/pools/<int:pool_id>/google-sheets', methods=['GET'])
@require_role('team_lead', 'admin')
def get_pool_google_sheets(pool_id):
    pool, error = _active_google_sheets_pool_or_error(pool_id)
    if error:
        return error
    return jsonify(_google_sheet_connection_payload(pool))


@app.route('/api/pools/<int:pool_id>/google-sheets', methods=['PATCH'])
@require_role('team_lead', 'admin')
def update_pool_google_sheets(pool_id):
    pool, error = _active_google_sheets_pool_or_error(pool_id)
    if error:
        return error
    data = request.json or {}
    sheet_url = (data.get('sheet_url') or '').strip()
    webhook_url = (data.get('webhook_url') or '').strip()
    enabled = bool(data.get('enabled'))
    if sheet_url and not re.match(r'^https://docs\.google\.com/spreadsheets/d/[\w-]+', sheet_url):
        return jsonify({'error': 'Некорректная ссылка на Google Sheets'}), 400
    if webhook_url and not re.match(r'^https://script\.google\.com/macros/s/[\w-]+/exec$', webhook_url):
        return jsonify({'error': 'Некорректный URL веб-приложения Apps Script'}), 400
    if enabled and (not sheet_url or not webhook_url):
        return jsonify({'error': 'Для включения нужны ссылка на таблицу и URL Apps Script'}), 400
    pool.google_sheet_url = sheet_url or None
    pool.google_sheet_webhook_url = webhook_url or None
    pool.google_sheet_enabled = enabled
    pool.google_sheet_last_error = ''
    log_action(
        'update',
        'google_sheets_export',
        pool.id,
        f'Настроена автовыгрузка Google Sheets для бассейна «{pool.name}»',
        {'enabled': enabled, 'sheet_url': sheet_url},
    )
    db.session.commit()
    return jsonify(_google_sheet_connection_payload(pool))


@app.route('/api/pools/<int:pool_id>/google-sheets/export', methods=['POST'])
@require_role('team_lead', 'admin')
def export_pool_google_sheets_now(pool_id):
    pool, error = _active_google_sheets_pool_or_error(pool_id)
    if error:
        return error
    try:
        result = export_pool_to_google_sheets(pool)
        log_action(
            'export',
            'google_sheets_export',
            pool.id,
            f'Данные бассейна «{pool.name}» выгружены в Google Sheets',
            {'sheets': result.get('sheets') or GOOGLE_TEMPLATE_SHEETS, 'warnings': result.get('warnings') or []},
        )
        db.session.commit()
        return jsonify({
            'message': 'Таблица обновлена',
            'sheets': result.get('sheets') or GOOGLE_TEMPLATE_SHEETS,
            'warnings': result.get('warnings') or [],
            **_google_sheet_connection_payload(pool),
        })
    except (ValueError, RuntimeError) as exc:
        db.session.commit()
        return jsonify({'error': str(exc), **_google_sheet_connection_payload(pool)}), 502


@app.route('/api/internal/google-sheets/export', methods=['POST'])
def export_configured_google_sheets():
    if not verify_internal_api_secret():
        return jsonify({'error': 'Недостаточно прав для Google Sheets export'}), 403
    pools = Pool.query.filter_by(
        active=True,
        archived=False,
        google_sheet_enabled=True,
    ).order_by(Pool.id).all()
    exported = []
    errors = []
    for pool in pools:
        try:
            result = export_pool_to_google_sheets(pool)
            exported.append({'pool_id': pool.id, 'sheets': result.get('sheets') or GOOGLE_TEMPLATE_SHEETS})
        except (ValueError, RuntimeError) as exc:
            errors.append({'pool_id': pool.id, 'error': str(exc)})
    db.session.commit()
    status = 502 if errors else 200
    return jsonify({'ok': not errors, 'exported': exported, 'errors': errors}), status


@app.route('/api/admin/backup-status', methods=['GET'])
@require_role('team_lead', 'admin')
def backup_status():
    latest = latest_backup_file()
    return jsonify({
        'persistent': not _is_production_runtime(),
        'backup_dir': BACKUP_DIR,
        'latest_file': latest,
        'latest_at': datetime.fromtimestamp(os.path.getmtime(latest)).isoformat() if latest else None,
        'today_exists': os.path.exists(backup_filename_for_today()),
    })


@app.route('/api/admin/backup-now', methods=['POST'])
@require_role('team_lead', 'admin')
def backup_now():
    if _is_production_runtime():
        return jsonify({
            'error': 'Локальные резервные копии недоступны в serverless. Используйте Excel-экспорт или внешнее хранилище.'
        }), 409
    path = create_daily_backup(force=True)
    log_action('backup', 'workbook', None, 'Создан ручной резерв Excel', {'path': path})
    db.session.commit()
    return jsonify({'message': 'Резерв создан', 'path': path})


@app.route('/api/admin/action-log', methods=['GET'])
@require_role('team_lead', 'admin')
def action_log():
    limit = min(request.args.get('limit', default=80, type=int) or 80, 200)
    rows = ActionLog.query.order_by(ActionLog.created_at.desc(), ActionLog.id.desc()).limit(limit).all()
    return jsonify([{
        'id': row.id,
        'actor_nick': row.actor_nick,
        'actor_name': row.actor_name,
        'action': row.action,
        'entity': row.entity,
        'entity_id': row.entity_id,
        'description': row.description or '',
        'payload': json.loads(row.payload or '{}'),
        'created_at': row.created_at.isoformat(),
    } for row in rows])


@app.route('/api/admin/reset', methods=['POST'])
@require_role('admin')
def reset_database():
    data = request.json or {}
    if data.get('confirm') != 'RESET':
        return jsonify({'error': 'Нужно подтверждение confirm=RESET'}), 400
    db.drop_all()
    db.create_all()
    seed_admin()
    return jsonify({'message': 'База сброшена'})


@app.route('/api/admin/seed', methods=['POST'])
@require_role('admin')
def seed_database_endpoint():
    """Заполнить БД тестовыми данными для бассейна."""
    try:
        from seed_pool import run
        run()
        return jsonify({'message': 'База заполнена данными бассейна'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Инициализация ====================


def seed_admin():
    """Создать стартового админа, если его нет."""
    admin_nick = os.getenv('ADMIN_NICK', 'admin')
    admin_pass = os.getenv('ADMIN_PASSWORD')
    if not admin_pass:
        if _is_production_runtime():
            raise RuntimeError('ADMIN_PASSWORD обязателен для первичного production seed')
        admin_pass = 'admin123'
    if not User.query.filter(db.func.lower(User.nick) == admin_nick.lower()).first():
        admin = User(
            nick=admin_nick,
            name='Администратор',
            role='admin',
            password_hash=generate_password_hash(admin_pass),
        )
        db.session.add(admin)
        db.session.commit()
        print(f'[seed] admin создан: ник="{admin_nick}" пароль="{admin_pass}"')


def seed_pool_data():
    """Автоматически загрузить данные бассейна при первом запуске."""
    # Проверяем есть ли уже данные в базе
    if ShiftBlock.query.first() is not None:
        return  # Данные уже есть, не загружаем

    try:
        from seed_pool import run
        run()
        print('[seed] тестовые данные бассейна загружены автоматически')
    except Exception as e:
        print(f'[seed] ошибка при загрузке тестовых данных: {e}')


def ensure_user_profile_columns():
    """Лёгкая SQLite-миграция для новых полей профиля волонтёра."""
    if db.engine.dialect.name != 'sqlite':
        return
    with db.engine.connect() as conn:
        tables = {row[0] for row in conn.exec_driver_sql(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()}
        columns = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(users)').fetchall()}
        if 'is_group_reviewer' not in columns:
            conn.exec_driver_sql('ALTER TABLE users ADD COLUMN is_group_reviewer BOOLEAN DEFAULT 0')
        if 'has_confession' not in columns:
            conn.exec_driver_sql('ALTER TABLE users ADD COLUMN has_confession BOOLEAN DEFAULT 0')
        if 'coins_adjustment' not in columns:
            conn.exec_driver_sql('ALTER TABLE users ADD COLUMN coins_adjustment INTEGER DEFAULT 0')
        if 'tribe' not in columns:
            conn.exec_driver_sql('ALTER TABLE users ADD COLUMN tribe VARCHAR(50)')
        shift_columns = set()
        if 'shift_blocks' in tables:
            shift_columns = {
                row[1] for row in conn.exec_driver_sql('PRAGMA table_info(shift_blocks)').fetchall()
            }
            if 'generation_id' not in shift_columns:
                conn.exec_driver_sql('ALTER TABLE shift_blocks ADD COLUMN generation_id INTEGER')
        student_event_columns = set()
        if 'student_events' in tables:
            student_event_columns = {
                row[1] for row in conn.exec_driver_sql('PRAGMA table_info(student_events)').fetchall()
            }
        if 'reward_events' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE reward_events (
                    id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL,
                    event_type VARCHAR(50) NOT NULL,
                    event_date DATE,
                    quantity INTEGER DEFAULT 1,
                    coins INTEGER NOT NULL,
                    comment TEXT,
                    created_by INTEGER,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(user_id) REFERENCES users (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        if 'student_events' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE student_events (
                    id INTEGER NOT NULL,
                    student_id INTEGER NOT NULL,
                    event_type VARCHAR(30) NOT NULL,
                    event_date DATE,
                    post_url VARCHAR(500),
                    proof_url VARCHAR(500),
                    points INTEGER DEFAULT 0,
                    status VARCHAR(20) DEFAULT 'pending',
                    comment TEXT,
                    created_by INTEGER,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(student_id) REFERENCES students (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        else:
            if 'post_url' not in student_event_columns:
                conn.exec_driver_sql('ALTER TABLE student_events ADD COLUMN post_url VARCHAR(500)')
            if 'proof_url' not in student_event_columns:
                conn.exec_driver_sql('ALTER TABLE student_events ADD COLUMN proof_url VARCHAR(500)')
            if 'points' not in student_event_columns:
                conn.exec_driver_sql('ALTER TABLE student_events ADD COLUMN points INTEGER DEFAULT 0')
            if 'status' not in student_event_columns:
                conn.exec_driver_sql("ALTER TABLE student_events ADD COLUMN status VARCHAR(20) DEFAULT 'confirmed'")
            conn.exec_driver_sql("UPDATE student_events SET status = 'confirmed' WHERE status IS NULL OR status = ''")
            conn.exec_driver_sql("UPDATE student_events SET points = 2 WHERE event_type = 'entertainment' AND (points IS NULL OR points = 0)")
            conn.exec_driver_sql("UPDATE student_events SET points = 4 WHERE event_type = 'education' AND (points IS NULL OR points = 0)")
        if 'tribe_events' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE tribe_events (
                    id INTEGER NOT NULL,
                    pool_id INTEGER,
                    tribe VARCHAR(50) NOT NULL,
                    title VARCHAR(200) NOT NULL,
                    event_date DATE NOT NULL,
                    time_start VARCHAR(5),
                    location VARCHAR(200),
                    comment TEXT,
                    created_by INTEGER,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        if 'schedule_generations' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE schedule_generations (
                    id INTEGER NOT NULL,
                    pool_id INTEGER NOT NULL,
                    end_date DATE,
                    created_by INTEGER,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        if 'group_reviews' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE group_reviews (
                    id INTEGER NOT NULL,
                    pool_id INTEGER,
                    review_date DATE NOT NULL,
                    time_start VARCHAR(5) NOT NULL,
                    flow VARCHAR(50) NOT NULL,
                    quantity INTEGER DEFAULT 1,
                    reviewer_id INTEGER NOT NULL,
                    created_by INTEGER,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(reviewer_id) REFERENCES users (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        else:
            group_review_columns = {
                row[1] for row in conn.exec_driver_sql('PRAGMA table_info(group_reviews)').fetchall()
            }
            if 'quantity' not in group_review_columns:
                conn.exec_driver_sql('ALTER TABLE group_reviews ADD COLUMN quantity INTEGER DEFAULT 1')
            conn.exec_driver_sql('UPDATE group_reviews SET quantity = 1 WHERE quantity IS NULL OR quantity < 1')
        if 'penalty_history' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE penalty_history (
                    id INTEGER NOT NULL,
                    penalty_id INTEGER NOT NULL,
                    old_status VARCHAR(20),
                    new_status VARCHAR(20) NOT NULL,
                    old_hours INTEGER,
                    new_hours INTEGER,
                    actor_id INTEGER,
                    actor_nick VARCHAR(100),
                    actor_name VARCHAR(100),
                    comment TEXT,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(actor_id) REFERENCES users (id)
                )
            """)
        if 'action_logs' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE action_logs (
                    id INTEGER NOT NULL,
                    actor_id INTEGER,
                    actor_nick VARCHAR(100),
                    actor_name VARCHAR(100),
                    action VARCHAR(80) NOT NULL,
                    entity VARCHAR(50) NOT NULL,
                    entity_id INTEGER,
                    description TEXT,
                    payload TEXT,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(actor_id) REFERENCES users (id)
                )
            """)
        if 'telegram_accounts' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE telegram_accounts (
                    id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL UNIQUE,
                    telegram_username VARCHAR(100) NOT NULL,
                    telegram_user_id VARCHAR(100) UNIQUE,
                    telegram_chat_id VARCHAR(100),
                    is_linked BOOLEAN DEFAULT 0,
                    linked_at DATETIME,
                    last_photo_sync_at DATETIME,
                    photo_file_id VARCHAR(255),
                    photo_url VARCHAR(500),
                    delivery_enabled BOOLEAN DEFAULT 1,
                    last_delivery_at DATETIME,
                    created_at DATETIME,
                    updated_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(user_id) REFERENCES users (id)
                )
            """)
        if 'notification_events' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE notification_events (
                    id INTEGER NOT NULL,
                    type VARCHAR(50) NOT NULL,
                    priority VARCHAR(20) DEFAULT 'normal',
                    status VARCHAR(20) DEFAULT 'draft',
                    scheduled_for DATETIME,
                    sent_at DATETIME,
                    cancelled_at DATETIME,
                    processing_started_at DATETIME,
                    recipient_user_id INTEGER,
                    pool_id INTEGER,
                    payload TEXT,
                    dedupe_key VARCHAR(255),
                    source_entity VARCHAR(50),
                    source_entity_id INTEGER,
                    created_by INTEGER,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(recipient_user_id) REFERENCES users (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        else:
            notification_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(notification_events)').fetchall()}
            if 'processing_started_at' not in notification_cols:
                conn.exec_driver_sql('ALTER TABLE notification_events ADD COLUMN processing_started_at DATETIME')
        if 'notification_deliveries' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE notification_deliveries (
                    id INTEGER NOT NULL,
                    notification_id INTEGER NOT NULL,
                    user_id INTEGER,
                    telegram_chat_id VARCHAR(100),
                    delivery_status VARCHAR(20) DEFAULT 'pending',
                    error TEXT,
                    message_id VARCHAR(100),
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(notification_id) REFERENCES notification_events (id),
                    FOREIGN KEY(user_id) REFERENCES users (id)
                )
            """)
        if 'broadcasts' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE broadcasts (
                    id INTEGER NOT NULL,
                    author_id INTEGER NOT NULL,
                    pool_id INTEGER,
                    text TEXT NOT NULL,
                    filters TEXT,
                    priority VARCHAR(20) DEFAULT 'normal',
                    status VARCHAR(20) DEFAULT 'draft',
                    created_at DATETIME,
                    updated_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(author_id) REFERENCES users (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id)
                )
            """)
        else:
            broadcast_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(broadcasts)').fetchall()}
            if 'pool_id' not in broadcast_cols:
                conn.exec_driver_sql('ALTER TABLE broadcasts ADD COLUMN pool_id INTEGER')
            if 'is_anonymous' not in broadcast_cols:
                conn.exec_driver_sql('ALTER TABLE broadcasts ADD COLUMN is_anonymous BOOLEAN DEFAULT 0')
        if 'dashboard_notes' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE dashboard_notes (
                    id INTEGER NOT NULL,
                    author_id INTEGER NOT NULL,
                    pool_id INTEGER,
                    text TEXT NOT NULL,
                    is_pinned BOOLEAN DEFAULT 0,
                    is_highlighted BOOLEAN DEFAULT 0,
                    is_active BOOLEAN DEFAULT 1,
                    is_anonymous BOOLEAN DEFAULT 0,
                    created_at DATETIME,
                    updated_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(author_id) REFERENCES users (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id)
                )
            """)
        else:
            note_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(dashboard_notes)').fetchall()}
            if 'pool_id' not in note_cols:
                conn.exec_driver_sql('ALTER TABLE dashboard_notes ADD COLUMN pool_id INTEGER')
            if 'is_anonymous' not in note_cols:
                conn.exec_driver_sql('ALTER TABLE dashboard_notes ADD COLUMN is_anonymous BOOLEAN DEFAULT 0')
        if 'app_settings' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE app_settings (
                    key VARCHAR(100) NOT NULL,
                    value TEXT,
                    updated_at DATETIME,
                    PRIMARY KEY (key)
                )
            """)
        # Pool.archived column
        if 'pools' in tables:
            pool_columns = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(pools)').fetchall()}
            if 'archived' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN archived BOOLEAN DEFAULT 0')
            if 'google_sheet_url' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN google_sheet_url TEXT')
            if 'google_sheet_webhook_url' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN google_sheet_webhook_url TEXT')
            if 'google_sheet_enabled' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN google_sheet_enabled BOOLEAN DEFAULT 0')
            if 'google_sheet_last_export_at' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN google_sheet_last_export_at DATETIME')
            if 'google_sheet_last_error' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN google_sheet_last_error TEXT')
            if 'self_signup_enabled' not in pool_columns:
                conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN self_signup_enabled BOOLEAN DEFAULT 1 NOT NULL')
        # PoolVolunteer table
        if 'pool_volunteers' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE pool_volunteers (
                    id INTEGER NOT NULL,
                    pool_id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL,
                    tribe VARCHAR(50),
                    pool_role VARCHAR(40) DEFAULT 'volunteer',
                    notifications_enabled BOOLEAN DEFAULT 1 NOT NULL,
                    has_confession BOOLEAN DEFAULT 0,
                    coins_adjustment INTEGER DEFAULT 0,
                    assigned_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id),
                    FOREIGN KEY(user_id) REFERENCES users (id),
                    UNIQUE (pool_id, user_id)
                )
            """)
        else:
            pv_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(pool_volunteers)').fetchall()}
            if 'pool_role' not in pv_cols:
                conn.exec_driver_sql("ALTER TABLE pool_volunteers ADD COLUMN pool_role VARCHAR(40) DEFAULT 'volunteer'")
            if 'notifications_enabled' not in pv_cols:
                conn.exec_driver_sql('ALTER TABLE pool_volunteers ADD COLUMN notifications_enabled BOOLEAN DEFAULT 1 NOT NULL')
            if 'has_confession' not in pv_cols:
                conn.exec_driver_sql('ALTER TABLE pool_volunteers ADD COLUMN has_confession BOOLEAN DEFAULT 0')
            if 'coins_adjustment' not in pv_cols:
                conn.exec_driver_sql('ALTER TABLE pool_volunteers ADD COLUMN coins_adjustment INTEGER DEFAULT 0')
        if 'pool_invite_links' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE pool_invite_links (
                    id INTEGER NOT NULL,
                    pool_id INTEGER NOT NULL UNIQUE,
                    token VARCHAR(128) NOT NULL UNIQUE,
                    created_by INTEGER NOT NULL,
                    is_active BOOLEAN DEFAULT 1,
                    uses_count INTEGER DEFAULT 0,
                    max_uses INTEGER,
                    expires_at DATETIME,
                    created_at DATETIME,
                    updated_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id),
                    FOREIGN KEY(created_by) REFERENCES users (id)
                )
            """)
        else:
            pool_invite_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(pool_invite_links)').fetchall()}
            if 'max_uses' not in pool_invite_cols:
                conn.exec_driver_sql('ALTER TABLE pool_invite_links ADD COLUMN max_uses INTEGER')
            if 'expires_at' not in pool_invite_cols:
                conn.exec_driver_sql('ALTER TABLE pool_invite_links ADD COLUMN expires_at DATETIME')
        if 'telegram_accounts' in tables:
            telegram_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(telegram_accounts)').fetchall()}
            if 'telegram_user_id' not in telegram_cols:
                conn.exec_driver_sql('ALTER TABLE telegram_accounts ADD COLUMN telegram_user_id VARCHAR(100)')
            conn.exec_driver_sql(
                'CREATE UNIQUE INDEX IF NOT EXISTS uq_telegram_account_user_id '
                'ON telegram_accounts (telegram_user_id) WHERE telegram_user_id IS NOT NULL'
            )
        if 'student_penalties' in tables:
            penalty_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(student_penalties)').fetchall()}
            if 'student_id' not in penalty_cols:
                conn.exec_driver_sql('ALTER TABLE student_penalties ADD COLUMN student_id INTEGER')
            conn.exec_driver_sql("""
                UPDATE student_penalties
                SET student_id = (
                    SELECT students.id FROM students
                    WHERE students.pool_id = student_penalties.pool_id
                      AND lower(students.nick) = lower(student_penalties.student_name)
                    LIMIT 1
                )
                WHERE student_id IS NULL
            """)
        # RewardEvent.pool_id
        if 'reward_events' in tables:
            re_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(reward_events)').fetchall()}
            if 'pool_id' not in re_cols:
                conn.exec_driver_sql('ALTER TABLE reward_events ADD COLUMN pool_id INTEGER')
        if 'tribes' not in tables:
            conn.exec_driver_sql("""
                CREATE TABLE tribes (
                    id INTEGER NOT NULL,
                    pool_id INTEGER,
                    name VARCHAR(100) NOT NULL,
                    created_at DATETIME,
                    PRIMARY KEY (id),
                    FOREIGN KEY(pool_id) REFERENCES pools (id)
                )
            """)
        else:
            tribe_cols = {row[1] for row in conn.exec_driver_sql('PRAGMA table_info(tribes)').fetchall()}
            if 'pool_id' not in tribe_cols:
                conn.exec_driver_sql('ALTER TABLE tribes RENAME TO tribes_legacy')
                conn.exec_driver_sql("""
                    CREATE TABLE tribes (
                        id INTEGER NOT NULL,
                        pool_id INTEGER,
                        name VARCHAR(100) NOT NULL,
                        created_at DATETIME,
                        PRIMARY KEY (id),
                        FOREIGN KEY(pool_id) REFERENCES pools (id)
                    )
                """)
                active_pool = conn.exec_driver_sql(
                    'SELECT id FROM pools WHERE active = 1 ORDER BY created_at DESC LIMIT 1'
                ).fetchone()
                active_pool_id_value = active_pool[0] if active_pool else None
                conn.exec_driver_sql(
                    'INSERT INTO tribes (id, pool_id, name, created_at) SELECT id, ?, name, created_at FROM tribes_legacy',
                    (active_pool_id_value,),
                )
                conn.exec_driver_sql('DROP TABLE tribes_legacy')
        active_pool = conn.exec_driver_sql(
            'SELECT id FROM pools WHERE active = 1 ORDER BY created_at DESC LIMIT 1'
        ).fetchone()
        active_pool_id_value = active_pool[0] if active_pool else None
        if active_pool_id_value is not None:
            conn.exec_driver_sql("""
                UPDATE broadcasts
                SET pool_id = (
                    SELECT ne.pool_id
                    FROM notification_events ne
                    WHERE ne.source_entity = 'broadcast'
                      AND ne.source_entity_id = broadcasts.id
                      AND ne.pool_id IS NOT NULL
                    ORDER BY ne.id ASC
                    LIMIT 1
                )
                WHERE pool_id IS NULL
            """)
            conn.exec_driver_sql('UPDATE dashboard_notes SET pool_id = ? WHERE pool_id IS NULL', (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE broadcasts SET pool_id = ? WHERE pool_id IS NULL', (active_pool_id_value,))
        for table in ('users', 'students', 'tribe_events'):
            if table in tables:
                conn.exec_driver_sql(f"UPDATE {table} SET tribe = 'Ленты' WHERE lower(tribe) IN ('a', '1')")
                conn.exec_driver_sql(f"UPDATE {table} SET tribe = 'Короны' WHERE lower(tribe) IN ('b', '2')")
                conn.exec_driver_sql(f"UPDATE {table} SET tribe = 'Олени' WHERE lower(tribe) IN ('c', '3')")
        conn.commit()


def ensure_postgres_profile_columns():
    """Лёгкая Postgres-миграция для старых production-баз без Alembic."""
    if db.engine.dialect.name != 'postgresql':
        return
    with db.engine.connect() as conn:
        conn.exec_driver_sql('ALTER TABLE users ADD COLUMN IF NOT EXISTS is_group_reviewer BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE users ADD COLUMN IF NOT EXISTS has_confession BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE users ADD COLUMN IF NOT EXISTS coins_adjustment INTEGER DEFAULT 0')
        conn.exec_driver_sql('ALTER TABLE users ADD COLUMN IF NOT EXISTS tribe VARCHAR(50)')

        conn.exec_driver_sql('ALTER TABLE shift_blocks ADD COLUMN IF NOT EXISTS generation_id INTEGER')

        conn.exec_driver_sql('ALTER TABLE student_events ADD COLUMN IF NOT EXISTS post_url VARCHAR(500)')
        conn.exec_driver_sql('ALTER TABLE student_events ADD COLUMN IF NOT EXISTS proof_url VARCHAR(500)')
        conn.exec_driver_sql('ALTER TABLE student_events ADD COLUMN IF NOT EXISTS points INTEGER DEFAULT 0')
        conn.exec_driver_sql("ALTER TABLE student_events ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'confirmed'")
        conn.exec_driver_sql("UPDATE student_events SET status = 'confirmed' WHERE status IS NULL OR status = ''")
        conn.exec_driver_sql("UPDATE student_events SET points = 2 WHERE event_type = 'entertainment' AND (points IS NULL OR points = 0)")
        conn.exec_driver_sql("UPDATE student_events SET points = 4 WHERE event_type = 'education' AND (points IS NULL OR points = 0)")

        conn.exec_driver_sql('ALTER TABLE group_reviews ADD COLUMN IF NOT EXISTS quantity INTEGER DEFAULT 1')
        conn.exec_driver_sql('UPDATE group_reviews SET quantity = 1 WHERE quantity IS NULL OR quantity < 1')

        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS archived BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS google_sheet_url TEXT')
        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS google_sheet_webhook_url TEXT')
        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS google_sheet_enabled BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS google_sheet_last_export_at TIMESTAMP')
        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS google_sheet_last_error TEXT')
        conn.exec_driver_sql('ALTER TABLE pools ADD COLUMN IF NOT EXISTS self_signup_enabled BOOLEAN DEFAULT TRUE NOT NULL')
        conn.exec_driver_sql('ALTER TABLE broadcasts ADD COLUMN IF NOT EXISTS pool_id INTEGER')
        conn.exec_driver_sql('ALTER TABLE broadcasts ADD COLUMN IF NOT EXISTS is_anonymous BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE dashboard_notes ADD COLUMN IF NOT EXISTS pool_id INTEGER')
        conn.exec_driver_sql('ALTER TABLE dashboard_notes ADD COLUMN IF NOT EXISTS is_anonymous BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE tribes ADD COLUMN IF NOT EXISTS pool_id INTEGER')
        conn.exec_driver_sql('ALTER TABLE tribes DROP CONSTRAINT IF EXISTS tribes_name_key')

        conn.exec_driver_sql("ALTER TABLE pool_volunteers ADD COLUMN IF NOT EXISTS pool_role VARCHAR(40) DEFAULT 'volunteer'")
        conn.exec_driver_sql('ALTER TABLE pool_volunteers ALTER COLUMN pool_role TYPE VARCHAR(40)')
        conn.exec_driver_sql('ALTER TABLE pool_volunteers ADD COLUMN IF NOT EXISTS notifications_enabled BOOLEAN DEFAULT TRUE NOT NULL')
        conn.exec_driver_sql('ALTER TABLE pool_volunteers ADD COLUMN IF NOT EXISTS has_confession BOOLEAN DEFAULT FALSE')
        conn.exec_driver_sql('ALTER TABLE pool_volunteers ADD COLUMN IF NOT EXISTS coins_adjustment INTEGER DEFAULT 0')
        conn.exec_driver_sql('ALTER TABLE pool_invite_links ADD COLUMN IF NOT EXISTS max_uses INTEGER')
        conn.exec_driver_sql('ALTER TABLE pool_invite_links ADD COLUMN IF NOT EXISTS expires_at TIMESTAMP')

        conn.exec_driver_sql('ALTER TABLE reward_events ADD COLUMN IF NOT EXISTS pool_id INTEGER')
        conn.exec_driver_sql('ALTER TABLE telegram_accounts ALTER COLUMN photo_url TYPE TEXT')
        conn.exec_driver_sql('ALTER TABLE telegram_accounts ADD COLUMN IF NOT EXISTS telegram_user_id VARCHAR(100)')
        conn.exec_driver_sql('ALTER TABLE notification_events ADD COLUMN IF NOT EXISTS processing_started_at TIMESTAMP')
        conn.exec_driver_sql('ALTER TABLE student_penalties ADD COLUMN IF NOT EXISTS student_id INTEGER')

        conn.exec_driver_sql("""
            UPDATE student_penalties sp
            SET student_id = s.id
            FROM students s
            WHERE sp.student_id IS NULL
              AND s.pool_id = sp.pool_id
              AND lower(s.nick) = lower(sp.student_name)
        """)
        conn.exec_driver_sql('ALTER TABLE students DROP CONSTRAINT IF EXISTS students_nick_key')
        conn.exec_driver_sql(
            'CREATE UNIQUE INDEX IF NOT EXISTS uq_student_pool_nick_ci '
            'ON students (pool_id, lower(nick))'
        )
        conn.exec_driver_sql(
            'CREATE UNIQUE INDEX IF NOT EXISTS uq_telegram_account_user_id '
            'ON telegram_accounts (telegram_user_id) WHERE telegram_user_id IS NOT NULL'
        )

        active_pool = conn.exec_driver_sql(
            'SELECT id FROM pools WHERE active = TRUE ORDER BY created_at DESC LIMIT 1'
        ).fetchone()
        active_pool_id_value = active_pool[0] if active_pool else None
        if active_pool_id_value is not None:
            conn.exec_driver_sql(
                'UPDATE pools SET active = FALSE WHERE active = TRUE AND id != %s',
                (active_pool_id_value,),
            )
            conn.exec_driver_sql("""
                UPDATE broadcasts
                SET pool_id = COALESCE(
                    pool_id,
                    (
                        SELECT ne.pool_id
                        FROM notification_events ne
                        WHERE ne.source_entity = 'broadcast'
                          AND ne.source_entity_id = broadcasts.id
                          AND ne.pool_id IS NOT NULL
                        ORDER BY ne.id ASC
                        LIMIT 1
                    ),
                    %s
                )
                WHERE pool_id IS NULL
            """, (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE dashboard_notes SET pool_id = %s WHERE pool_id IS NULL', (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE tribes SET pool_id = %s WHERE pool_id IS NULL', (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE students SET pool_id = %s WHERE pool_id IS NULL', (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE group_reviews SET pool_id = %s WHERE pool_id IS NULL', (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE tribe_events SET pool_id = %s WHERE pool_id IS NULL', (active_pool_id_value,))
            conn.exec_driver_sql('UPDATE student_penalties SET pool_id = %s WHERE pool_id IS NULL', (active_pool_id_value,))

        conn.exec_driver_sql(
            'CREATE UNIQUE INDEX IF NOT EXISTS uq_single_active_pool '
            'ON pools (active) WHERE active = TRUE'
        )
        conn.exec_driver_sql("""
            DO $$ BEGIN
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_student_penalty_student') THEN
                    ALTER TABLE student_penalties
                    ADD CONSTRAINT fk_student_penalty_student FOREIGN KEY (student_id) REFERENCES students(id) NOT VALID;
                END IF;
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_student_penalty_pool') THEN
                    ALTER TABLE student_penalties
                    ADD CONSTRAINT fk_student_penalty_pool FOREIGN KEY (pool_id) REFERENCES pools(id) NOT VALID;
                END IF;
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_penalty_history_penalty') THEN
                    ALTER TABLE penalty_history
                    ADD CONSTRAINT fk_penalty_history_penalty FOREIGN KEY (penalty_id)
                    REFERENCES student_penalties(id) ON DELETE CASCADE NOT VALID;
                END IF;
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_student_pool') THEN
                    ALTER TABLE students
                    ADD CONSTRAINT fk_student_pool FOREIGN KEY (pool_id) REFERENCES pools(id) NOT VALID;
                END IF;
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_group_review_pool') THEN
                    ALTER TABLE group_reviews
                    ADD CONSTRAINT fk_group_review_pool FOREIGN KEY (pool_id) REFERENCES pools(id) NOT VALID;
                END IF;
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_tribe_event_pool') THEN
                    ALTER TABLE tribe_events
                    ADD CONSTRAINT fk_tribe_event_pool FOREIGN KEY (pool_id) REFERENCES pools(id) NOT VALID;
                END IF;
            END $$;
        """)

        for table in ('users', 'students', 'tribe_events'):
            conn.exec_driver_sql(f"UPDATE {table} SET tribe = 'Ленты' WHERE lower(tribe) IN ('a', '1')")
            conn.exec_driver_sql(f"UPDATE {table} SET tribe = 'Короны' WHERE lower(tribe) IN ('b', '2')")
            conn.exec_driver_sql(f"UPDATE {table} SET tribe = 'Олени' WHERE lower(tribe) IN ('c', '3')")
        conn.commit()


def should_auto_init_db():
    value = os.getenv('AUTO_INIT_DB')
    if value is not None:
        return value.lower() == 'true'
    return not _is_production_runtime()


def should_run_schema_migrations():
    return _env_flag('RUN_SCHEMA_MIGRATIONS', 'false')


def run_database_migrations():
    """Apply schema changes explicitly, never as a production import side effect."""
    if db.engine.dialect.name != 'postgresql':
        db.create_all()
        ensure_user_profile_columns()
        return

    lock_connection = db.engine.connect()
    try:
        lock_connection.exec_driver_sql(
            "SELECT pg_advisory_lock(hashtext('school21-pool-management-schema'))"
        )
        db.create_all()
        ensure_postgres_profile_columns()
    finally:
        try:
            lock_connection.exec_driver_sql(
                "SELECT pg_advisory_unlock(hashtext('school21-pool-management-schema'))"
            )
        finally:
            lock_connection.close()


def validate_runtime_config():
    if _is_production_runtime() and app.config['SECRET_KEY'] == 'dev-secret-key-change-me':
        raise RuntimeError('SECRET_KEY должен быть задан безопасным значением в production')


def should_start_runtime_services():
    return (
        os.getenv('AUTO_START_WORKERS', 'false').lower() == 'true'
        and os.getenv('SKIP_APP_WORKERS', 'false').lower() != 'true'
    )


def should_auto_sync_telegram_commands():
    default = 'false' if _is_production_runtime() else 'true'
    return os.getenv('AUTO_SYNC_TELEGRAM_COMMANDS', default).lower() == 'true'


def sync_telegram_commands_on_startup():
    if not should_auto_sync_telegram_commands() or not _telegram_is_configured():
        return
    try:
        telegram_sync_commands()
    except Exception as exc:
        app.logger.warning('Telegram commands sync skipped: %s', exc)


validate_runtime_config()

if should_auto_init_db():
    with app.app_context():
        run_database_migrations()
        seed_admin()
        seed_pool_data()
elif should_run_schema_migrations():
    with app.app_context():
        run_database_migrations()


sync_telegram_commands_on_startup()


if __name__ != '__main__' and should_start_runtime_services():
    start_runtime_services()


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5001))
    debug = os.getenv('FLASK_DEBUG', 'true').lower() == 'true'
    # запускаем воркер один раз: в debug-режиме код стартует дважды (reloader),
    # поэтому стартуем только в дочернем процессе reloader или когда debug выключен
    if should_start_runtime_services() and (not debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true'):
        start_runtime_services()
    app.run(host='0.0.0.0', port=port, debug=debug)
